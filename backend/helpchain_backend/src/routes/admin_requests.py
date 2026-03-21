from __future__ import annotations

import csv
import secrets
from datetime import UTC, datetime, timedelta
from io import BytesIO, StringIO

from flask import (
    Response,
    abort,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_babel import gettext as _
from flask_login import current_user, login_required
from sqlalchemy import func, or_, select, tuple_
from sqlalchemy.orm import joinedload

from backend.audit import log_activity
from backend.extensions import db
from ..constants.categories import (
    REQUEST_CATEGORY_CODES,
    normalize_request_category,
    request_category_choices,
)
from ..models import (
    AdminUser,
    Case,
    Notification,
    Request,
    RequestActivity,
    RequestMetric,
    Structure,
    User,
    Volunteer,
    VolunteerAction,
    VolunteerInterest,
    VolunteerRequestState,
    utc_now,
)
from ..notifications.inapp import NUDGE_COOLDOWN_HOURS, send_nudge_notification
from ..services.case_risk import risk_label_from_score, score_request_risk
from ..services.case_summary import build_case_summary
from ..services.geocoding import request_address_display_text
from ..services.recommendation_engine import compute_recommendation
from ..services.risk_alerts import evaluate_case_alerts
from ..services.risk_engine import update_case_risk
from ..statuses import REQUEST_STATUS_ALLOWED, normalize_request_status
from .admin import (
    ASSIGN_SLA_HOURS,
    CASE_PRIORITY_RANK,
    RESOLVE_SLA_DAYS,
    SLA_QUEUE_KINDS,
    STATUS_LABELS,
    STATUS_LABELS_BG,
    VOLUNTEER_ASSIGN_SLA_HOURS,
    _admin_id,
    _admin_role_value,
    _append_case_event,
    _apply_sla_queue_filter,
    _audit_denied_action,
    _audit_request,
    _build_helpchain_recommendation,
    _build_notseen_subquery,
    _build_operational_blockages,
    _build_risk_ai_suggestion,
    _cases_enabled,
    _compute_case_signals,
    _current_structure_id,
    _is_global_admin,
    _lock_expired,
    _locked_by_other,
    _normalize_sla_days,
    _normalize_sla_kind,
    _notseen_hours_from_risk,
    _now_utc,
    _scope_requests,
    _table_exists,
    _table_has_column,
    admin_bp,
    admin_required,
    admin_required_404,
    admin_role_required,
    audit_admin_action,
    can_edit_request,
    get_volunteer_engagement_score,
    is_stale,
    is_safe_url,
    log_request_activity,
    suggest_best_professional,
)


def _log_status_change_once(
    req_id: int,
    old_status: str | None,
    new_status: str | None,
    actor_admin_id: int | None,
):
    """Add a single status_change activity only when there is a real change."""
    if not _table_has_column("request_activities", "volunteer_id"):
        return
    if (old_status or "") == (new_status or ""):
        return
    db.session.add(
        RequestActivity(
            request_id=req_id,
            actor_admin_id=actor_admin_id,
            action="status_change",
            old_value=old_status,
            new_value=new_status,
        )
    )


def _is_request_locked(req) -> bool:
    """Consider a request locked when status is done or cancelled (canonical)."""
    s = normalize_request_status(getattr(req, "status", None))
    return s in ("done", "cancelled")


@admin_bp.route("/emergency-requests", methods=["GET"])
@admin_required
def emergency_requests():
    admin_required_404()
    if not getattr(current_user, "is_admin", False):
        flash(_("Access denied."), "error")
        return redirect(url_for("main.index"))

    q = (request.args.get("q") or "").strip()
    days = int(request.args.get("days") or 7)
    days = max(1, min(days, 90))
    page = int(request.args.get("page") or 1)
    page = max(page, 1)
    per_page = int(request.args.get("per_page") or 25)
    per_page = max(10, min(per_page, 100))
    since = datetime.now(UTC).replace(tzinfo=None) - timedelta(days=days)

    query = _scope_requests(Request.query).filter(
        Request.created_at >= since, Request.category == "emergency"
    ).order_by(Request.created_at.desc())

    if q:
        query = query.filter(
            (Request.city.ilike(f"%{q}%"))
            | (Request.email.ilike(f"%{q}%"))
            | (Request.phone.ilike(f"%{q}%"))
            | (Request.priority.ilike(f"%{q}%"))
        )

    total = query.count()
    items = query.offset((page - 1) * per_page).limit(per_page).all()

    return render_template(
        "admin_emergency_requests.html",
        items=items,
        total=total,
        page=page,
        per_page=per_page,
        q=q,
        days=days,
    )


@admin_bp.get("/api/suggest-assignee/<int:req_id>")
@admin_required
def admin_suggest_assignee_api(req_id: int):
    admin_required_404()
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    try:
        return jsonify(
            {
                "status": "ok",
                "suggestion": suggest_best_professional(req),
            }
        )
    except Exception:
        current_app.logger.exception("admin_suggest_assignee_api_failed")
        return jsonify({"status": "error", "suggestion": None}), 500


@admin_bp.route("/update_status/<int:req_id>", methods=["POST"])
@admin_required
def update_status(req_id):
    admin_required_404()
    from flask import current_app

    if not current_app.config.get("TESTING", False):
        if not getattr(current_user, "is_admin", False):
            return jsonify({"error": "Unauthorized"}), 403

    req = db.session.get(Request, req_id)
    if not req:
        return jsonify({"error": "Request not found"}), 404
    if not can_edit_request(req, current_user):
        role = getattr(
            getattr(current_user, "role", None),
            "value",
            getattr(current_user, "role", None),
        )
        role = (role or "").strip().lower()
        if role not in {"ops", "admin", "superadmin", "super_admin"}:
            abort(403)

    new_status = (request.form.get("status") or "").strip()
    old_raw_status = req.status
    old_status = normalize_request_status(old_raw_status)
    new_status = normalize_request_status(new_status)

    if new_status not in REQUEST_STATUS_ALLOWED:
        current_app.logger.warning(
            "ADMIN update_status blocked invalid new_status=%r for request_id=%s",
            new_status,
            req_id,
        )
        flash("Invalid status.", "warning")
        return redirect(url_for("admin.admin_request_details", req_id=req_id))

    if not new_status or new_status == old_status:
        if request.is_json or (
            request.accept_mimetypes
            and request.accept_mimetypes.best == "application/json"
        ):
            return (
                jsonify(
                    {
                        "success": False,
                        "status": old_status,
                        "message": "No status change.",
                    }
                ),
                200,
            )
        flash("No status change.", "info")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    req.status = new_status
    closing_statuses = {"done", "cancelled"}
    if new_status in closing_statuses:
        req.completed_at = utc_now()
    else:
        req.completed_at = None
    log_request_activity(
        req,
        "status_change",
        old=old_status,
        new=new_status,
        actor_admin_id=getattr(current_user, "id", None),
    )
    _audit_request(
        req.id,
        action="status_change",
        message="Status updated",
        old=old_status,
        new=new_status,
    )
    if _table_exists("request_metrics"):
        metric = db.session.query(RequestMetric).filter_by(request_id=req.id).first()
        if metric is None:
            metric = RequestMetric(request_id=req.id)
            db.session.add(metric)
        if new_status == "done" and metric.time_to_complete is None and req.created_at:
            try:
                metric.time_to_complete = int(
                    (utc_now() - req.created_at).total_seconds()
                )
            except Exception:
                pass

    if new_status == "in_progress":
        assigned_volunteer_id = getattr(req, "assigned_volunteer_id", None)
        if not assigned_volunteer_id:
            current_app.logger.warning(
                "Interest sync skipped: request_id=%s set to in_progress without assigned_volunteer_id",
                req.id,
            )
        else:
            assigned_volunteer = db.session.get(Volunteer, int(assigned_volunteer_id))
            if assigned_volunteer is None:
                current_app.logger.warning(
                    "Interest sync skipped: request_id=%s assigned_volunteer_id=%s not found in volunteers",
                    req.id,
                    assigned_volunteer_id,
                )
                assigned_volunteer_id = None

        if assigned_volunteer_id:
            q = VolunteerInterest.query.filter_by(request_id=req.id)
            owner_latest = (
                q.filter_by(volunteer_id=assigned_volunteer_id)
                .order_by(VolunteerInterest.id.desc())
                .first()
            )
            if owner_latest is None:
                owner_latest = VolunteerInterest(
                    request_id=req.id,
                    volunteer_id=assigned_volunteer_id,
                    status="approved",
                )
                db.session.add(owner_latest)
                current_app.logger.info(
                    "Interest sync: created approved assigned-volunteer interest (request_id=%s, volunteer_id=%s)",
                    req.id,
                    assigned_volunteer_id,
                )
            elif owner_latest.status != "approved":
                owner_latest.status = "approved"
                db.session.add(owner_latest)
                current_app.logger.info(
                    "Interest sync: set assigned-volunteer interest to approved (request_id=%s, volunteer_id=%s)",
                    req.id,
                    assigned_volunteer_id,
                )

            pending_others = (
                q.filter(VolunteerInterest.status == "pending")
                .filter(VolunteerInterest.volunteer_id != assigned_volunteer_id)
                .all()
            )
            for vi_row in pending_others:
                vi_row.status = "rejected"
                db.session.add(vi_row)

            if pending_others:
                current_app.logger.info(
                    "Interest sync: rejected %s pending interests (request_id=%s, assigned_volunteer_id=%s)",
                    len(pending_others),
                    req.id,
                    assigned_volunteer_id,
                )

    elif new_status in {"done", "cancelled"}:
        q = VolunteerInterest.query.filter_by(request_id=req.id)
        pending_all = q.filter(VolunteerInterest.status == "pending").all()
        for vi_row in pending_all:
            vi_row.status = "rejected"
            db.session.add(vi_row)

        if pending_all:
            current_app.logger.info(
                "Interest sync: rejected %s pending interests on close (request_id=%s, new_status=%s)",
                len(pending_all),
                req.id,
                new_status,
            )

    db.session.commit()
    audit_admin_action(
        action="STATUS_CHANGE",
        target_type="Request",
        target_id=req.id,
        payload={"old": {"status": old_status}, "new": {"status": new_status}},
    )

    try:
        subject = f"Статусът на вашата заявка #{req.id} е променен на {new_status}"
        recipient = getattr(req, "email", None)
        recipient_name = getattr(req, "name", "Потребител")
        content = (
            f"Статусът на вашата заявка е променен на <b>{new_status}</b>.\n\n"
            f"Описание: {req.description or ''}"
        )
        context = {
            "subject": subject,
            "recipient_name": recipient_name,
            "content": content,
            "request_id": req.id,
            "new_status": new_status,
            "description": req.description,
            "updated_at": req.updated_at,
        }
        if recipient:
            from .admin import _send_status_email_async

            _send_status_email_async(recipient, subject, context)
    except Exception as e:
        import logging

        logging.warning(f"[EMAIL] Async status email scheduling failed: {e}")

    if request.is_json or (
        request.accept_mimetypes and request.accept_mimetypes.best == "application/json"
    ):
        return jsonify({"success": True, "status": new_status or req.status})
    flash(_("Status updated."), "success")
    return redirect(url_for("admin.admin_request_details", req_id=req_id))


@admin_bp.post("/requests/<int:req_id>/status")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_request_set_status(req_id: int):
    return update_status(req_id)


@admin_bp.post("/requests/bulk")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_requests_bulk():
    admin_required_404()

    action = (request.form.get("bulk_action") or "").strip()
    selected_ids_raw = request.form.getlist("selected_ids")
    selected_ids: list[int] = []
    for raw in selected_ids_raw:
        try:
            rid = int(raw)
        except Exception:
            continue
        if rid > 0:
            selected_ids.append(rid)
    selected_ids = sorted(set(selected_ids))

    if not action or not selected_ids:
        flash("No bulk action applied (missing action or selection).", "warning")
        return redirect(url_for("admin.admin_requests"))

    requests = _scope_requests(Request.query).filter(Request.id.in_(selected_ids)).all()
    requests_by_id = {r.id: r for r in requests}
    ordered_reqs = [requests_by_id[rid] for rid in selected_ids if rid in requests_by_id]

    status_map = {
        "set_status_pending": "pending",
        "set_status_in_progress": "in_progress",
        "set_status_done": "done",
        "set_status_rejected": "rejected",
        "status:pending": "pending",
        "status:in_progress": "in_progress",
        "status:done": "done",
        "status:rejected": "rejected",
    }

    changed = 0
    nudged = 0
    skipped = 0
    actor_admin_id = getattr(current_user, "id", None)

    if action in status_map:
        target_status = normalize_request_status(status_map[action])
        for req in ordered_reqs:
            if not can_edit_request(req, current_user):
                skipped += 1
                continue
            old_status = normalize_request_status(getattr(req, "status", None))
            if old_status == target_status:
                continue
            req.status = target_status
            if target_status in {"done", "cancelled"}:
                req.completed_at = utc_now()
            else:
                req.completed_at = None
            try:
                log_request_activity(
                    req,
                    "status_change",
                    old=old_status,
                    new=target_status,
                    actor_admin_id=actor_admin_id,
                )
            except Exception:
                pass
            changed += 1
        db.session.commit()
        flash(f"Bulk status updated: {changed} changed, {skipped} skipped.", "success")
        return redirect(url_for("admin.admin_requests"))

    if action in {"nudge_selected_volunteers", "nudge"}:
        for req in ordered_reqs:
            if not can_edit_request(req, current_user):
                skipped += 1
                continue
            volunteer_id = getattr(req, "assigned_volunteer_id", None)
            if not volunteer_id:
                skipped += 1
                continue
            created = send_nudge_notification(
                request_id=req.id,
                volunteer_id=int(volunteer_id),
                actor_admin_id=actor_admin_id,
            )
            if created:
                nudged += 1
        flash(f"Bulk nudge sent: {nudged} sent, {skipped} skipped.", "success")
        return redirect(url_for("admin.admin_requests"))

    if action in {"open_selected", "open", "copy_ids", "copy_links"}:
        flash("Bulk action is UI-only and has no server-side effect.", "info")
        return redirect(url_for("admin.admin_requests"))

    flash("Unknown bulk action.", "warning")
    return redirect(url_for("admin.admin_requests"))


@admin_bp.post("/requests/<int:req_id>/archive", endpoint="admin_request_archive")
@login_required
@admin_required
@admin_role_required("superadmin")
def admin_request_archive(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)

    old_status = normalize_request_status(getattr(req, "status", None))
    req.status = "cancelled"
    req.completed_at = utc_now()
    req.is_archived = True
    if getattr(req, "archived_at", None) is None:
        req.archived_at = utc_now()

    log_request_activity(
        req,
        "status_change",
        old=old_status,
        new="cancelled",
        actor_admin_id=getattr(current_user, "id", None),
    )
    db.session.commit()
    audit_admin_action(
        action="request.archive",
        target_type="Request",
        target_id=req.id,
        payload={
            "old": {"status": old_status, "archived": False},
            "new": {"status": "cancelled", "archived": True},
        },
    )
    flash("Request archived and closed.", "success")
    return redirect(url_for("admin.admin_request_details", req_id=req.id))


@admin_bp.get("/requests")
@login_required
@admin_required
@admin_role_required("superadmin")
def admin_requests():
    admin_required_404()
    if _admin_role_value() != "superadmin":
        _audit_denied_action(
            required_roles={"superadmin"},
            actor_role=_admin_role_value(),
        )
        abort(403)
    status_labels_bg = {
        "new": "Нови",
        "pending": "Чакащи",
        "approved": "Одобрени",
        "in_progress": "В процес",
        "done": "Приключени",
        "rejected": "Отхвърлени",
    }

    queue = (request.args.get("queue") or "").strip().lower()
    sla_kind = _normalize_sla_kind(request.args.get("sla_kind"))
    sla_days = _normalize_sla_days(request.args.get("sla_days", 30))
    active_sla_queue = bool(queue == "sla" and sla_kind)
    active_sla_filter_label = (
        SLA_QUEUE_KINDS.get(sla_kind, "") if active_sla_queue else ""
    )
    show_deleted = (request.args.get("deleted") or "").strip() == "1"
    query = _scope_requests(Request.query)
    query, status, q, risk, risk_level, no_owner, not_seen_72h, sort = (
        build_requests_query(query, request.args)
    )
    requests = query.all()
    now_aware = utc_now()
    now_naive = datetime.now(UTC).replace(tzinfo=None)
    sla_warn_no_owner_days = 2
    sla_stale_days = 7
    risk_notseen_tier_hours = _notseen_hours_from_risk(risk)
    risk_columns_ready = _table_has_column(
        "requests", "risk_level"
    ) and _table_has_column("requests", "risk_signals")
    critical_count = 0
    attention_count = 0
    no_owner_count = 0
    not_seen_72h_count = 0
    if risk_columns_ready:
        overview_query = _scope_requests(Request.query).filter(
            Request.deleted_at.is_(None)
        )
        critical_count = overview_query.filter(
            func.lower(func.coalesce(Request.risk_level, "")) == "critical"
        ).count()
        attention_count = overview_query.filter(
            func.lower(func.coalesce(Request.risk_level, "")) == "attention"
        ).count()
        no_owner_count = overview_query.filter(
            func.lower(func.coalesce(Request.risk_signals, "")).like("%no_owner%")
        ).count()
        not_seen_72h_count = overview_query.filter(
            func.lower(func.coalesce(Request.risk_signals, "")).like("%not_seen_72h%")
        ).count()
    age_days_by_id = {}
    for r in requests:
        created_at = getattr(r, "created_at", None)
        if created_at is None:
            age_days_by_id[int(r.id)] = 0
            continue
        try:
            if getattr(created_at, "tzinfo", None) is not None:
                created_at = created_at.replace(tzinfo=None)
            age_days_by_id[int(r.id)] = max((now_naive - created_at).days, 0)
        except Exception:
            age_days_by_id[int(r.id)] = 0

    scope_label = "Vue globale"
    if not _is_global_admin():
        try:
            sid = _current_structure_id()
            active_structure = db.session.get(Structure, sid)
            if active_structure and active_structure.name:
                scope_label = f"Structure active : {active_structure.name}"
            else:
                scope_label = f"Structure active : #{sid}"
        except Exception:
            scope_label = "Structure active : —"

    action_counts = {}
    last_signal_by_req = {}
    engagement_by_request = {}
    nudge_ui = {}
    volunteer_actions_supported = _table_exists("volunteer_actions")
    if requests and volunteer_actions_supported:
        req_ids = [r.id for r in requests]
        rows = (
            db.session.query(
                VolunteerAction.request_id,
                VolunteerAction.action,
                func.count(VolunteerAction.id),
            )
            .filter(VolunteerAction.request_id.in_(req_ids))
            .group_by(VolunteerAction.request_id, VolunteerAction.action)
            .all()
        )
        for rid, act, cnt in rows:
            action_counts.setdefault(rid, {}).update({act: cnt})

        last_rows = (
            VolunteerAction.query.filter(VolunteerAction.request_id.in_(req_ids))
            .order_by(
                VolunteerAction.request_id.asc(),
                VolunteerAction.updated_at.desc(),
                VolunteerAction.created_at.desc(),
            )
            .all()
        )
        for action in last_rows:
            if action.request_id not in last_signal_by_req:
                last_signal_by_req[action.request_id] = action

    if requests:
        assigned_volunteer_ids = sorted(
            {
                int(r.assigned_volunteer_id)
                for r in requests
                if getattr(r, "assigned_volunteer_id", None)
            }
        )
        engagement_by_volunteer = {}
        for volunteer_id in assigned_volunteer_ids:
            try:
                engagement_by_volunteer[volunteer_id] = get_volunteer_engagement_score(
                    volunteer_id,
                    now=now_naive,
                )
            except Exception:
                db.session.rollback()
                engagement_by_volunteer[volunteer_id] = {
                    "volunteer_id": int(volunteer_id),
                    "score": 0,
                    "label": "At risk",
                    "seen_within_24h": 0,
                    "not_seen_72h": 0,
                    "can_help": 0,
                    "cant_help": 0,
                }
        engagement_by_request = {
            r.id: engagement_by_volunteer.get(getattr(r, "assigned_volunteer_id", None))
            for r in requests
        }

        def _to_utc_naive(dt):
            if dt is None:
                return None
            if dt.tzinfo is None:
                return dt
            return dt.astimezone(UTC).replace(tzinfo=None)

        now = datetime.now(UTC).replace(tzinfo=None)
        cooldown = timedelta(hours=NUDGE_COOLDOWN_HOURS)
        pairs = {
            (r.id, int(r.assigned_volunteer_id))
            for r in requests
            if getattr(r, "id", None) and getattr(r, "assigned_volunteer_id", None)
        }
        nudge_by_pair: dict[tuple[int, int], datetime] = {}
        if pairs:
            try:
                nudge_rows = (
                    db.session.query(
                        Notification.request_id,
                        Notification.volunteer_id,
                        Notification.created_at,
                    )
                    .filter(Notification.type == "admin_nudge")
                    .filter(
                        tuple_(Notification.request_id, Notification.volunteer_id).in_(
                            pairs
                        )
                    )
                    .all()
                )
            except Exception:
                req_ids = {req_id for req_id, _ in pairs}
                nudge_rows = (
                    db.session.query(
                        Notification.request_id,
                        Notification.volunteer_id,
                        Notification.created_at,
                    )
                    .filter(Notification.type == "admin_nudge")
                    .filter(Notification.request_id.in_(req_ids))
                    .all()
                )

            for req_id, vol_id, created_at in nudge_rows:
                if (req_id, vol_id) not in pairs:
                    continue
                created_naive = _to_utc_naive(created_at)
                if not created_naive:
                    continue
                key = (int(req_id), int(vol_id))
                prev = nudge_by_pair.get(key)
                if prev is None or created_naive > prev:
                    nudge_by_pair[key] = created_naive

        for r in requests:
            rid = getattr(r, "id", None)
            vid = getattr(r, "assigned_volunteer_id", None)
            if not rid:
                continue
            if not vid:
                nudge_ui[rid] = {"disabled": True, "title": "No assigned volunteer"}
                continue

            created_at = nudge_by_pair.get((rid, int(vid)))
            if not created_at:
                nudge_ui[rid] = {
                    "disabled": False,
                    "title": "Send reminder to assigned volunteer",
                }
                continue

            next_at = created_at + cooldown
            if next_at > now:
                remaining = next_at - now
                mins = int(remaining.total_seconds() // 60)
                hrs = mins // 60
                mm = mins % 60
                if hrs > 0:
                    tip = f"Nudge available in {hrs}h {mm}m"
                else:
                    tip = f"Nudge available in {mm}m"
                nudge_ui[rid] = {"disabled": True, "title": tip}
            else:
                nudge_ui[rid] = {
                    "disabled": False,
                    "title": "Send reminder to assigned volunteer",
                }

    return render_template(
        "admin/requests.html",
        STATUS_LABELS_BG=status_labels_bg,
        STATUS_LABELS=STATUS_LABELS,
        requests=requests,
        age_days_by_id=age_days_by_id,
        scope_label=scope_label,
        status=status,
        q=q,
        risk=risk,
        risk_level=risk_level,
        no_owner=no_owner,
        not_seen_72h=not_seen_72h,
        sort=sort,
        show_deleted=show_deleted,
        now_aware=now_aware,
        now_naive=now_naive,
        SLA_WARN_NO_OWNER_DAYS=sla_warn_no_owner_days,
        SLA_STALE_DAYS=sla_stale_days,
        volunteer_action_counts=action_counts,
        last_signal_by_req=last_signal_by_req,
        engagement_by_request=engagement_by_request,
        nudge_ui=nudge_ui,
        risk_notseen_tier_hours=risk_notseen_tier_hours,
        queue=queue,
        sla_kind=sla_kind,
        sla_days=sla_days,
        active_sla_queue=active_sla_queue,
        active_sla_filter_label=active_sla_filter_label,
        critical_count=critical_count,
        attention_count=attention_count,
        no_owner_count=no_owner_count,
        not_seen_72h_count=not_seen_72h_count,
    )


@admin_bp.route("/requests/new", methods=["GET", "POST"])
@admin_required
@admin_role_required("superadmin")
def admin_request_new():
    admin_required_404()
    if _admin_role_value() != "superadmin":
        _audit_denied_action(
            required_roles={"superadmin"},
            actor_role=_admin_role_value(),
        )
        abort(403)

    def _ensure_internal_requester_user() -> User:
        email = "agent.intake@helpchain.local"
        username = "agent_intake"
        existing = User.query.filter(func.lower(User.email) == email).first()
        if existing:
            return existing
        existing = User.query.filter(func.lower(User.username) == username).first()
        if existing:
            return existing
        user = User(
            username=username,
            email=email,
            role="requester",
            is_active=True,
            password_hash="",
        )
        try:
            user.set_password(secrets.token_urlsafe(24))
        except Exception:
            user.password_hash = "!"
        db.session.add(user)
        db.session.flush()
        return user

    categories = [code for code, _label in request_category_choices()]

    structures = Structure.query.order_by(Structure.name.asc(), Structure.id.asc()).all()
    admins = (
        AdminUser.query.filter(AdminUser.is_active.is_(True))
        .order_by(AdminUser.username.asc(), AdminUser.id.asc())
        .all()
    )

    form_data = {
        "title": (request.form.get("title") or "").strip(),
        "description": (request.form.get("description") or "").strip(),
        "person_name": (request.form.get("person_name") or "").strip(),
        "email": (request.form.get("email") or "").strip(),
        "phone": (request.form.get("phone") or "").strip(),
        "address_line": (request.form.get("address_line") or "").strip(),
        "postcode": (request.form.get("postcode") or "").strip(),
        "city": (request.form.get("city") or "").strip(),
        "country": (request.form.get("country") or "France").strip(),
        "category": normalize_request_category(
            (request.form.get("category") or "").strip()
        ),
        "priority": (request.form.get("priority") or "standard").strip().lower(),
        "structure_id": (request.form.get("structure_id") or "").strip(),
        "owner_id": (request.form.get("owner_id") or "").strip(),
        "internal_notes": (request.form.get("internal_notes") or "").strip(),
    }
    form_errors: dict[str, str] = {}

    if request.method == "POST":
        for field_name, max_len in (
            ("address_line", 255),
            ("postcode", 32),
            ("city", 200),
            ("country", 120),
        ):
            if len(form_data[field_name]) > max_len:
                form_errors[field_name] = "Valeur trop longue."
        if not form_data["title"]:
            form_errors["title"] = "Veuillez renseigner le titre."
        if not form_data["description"]:
            form_errors["description"] = "Veuillez renseigner la description."
        if not form_data["person_name"]:
            form_errors["person_name"] = "Veuillez renseigner la personne concernée."
        if not form_data["city"]:
            form_errors["city"] = "Veuillez renseigner la ville ou le territoire."
        if not form_data["category"]:
            form_errors["category"] = "Veuillez sélectionner une catégorie."
        elif form_data["category"] not in set(REQUEST_CATEGORY_CODES):
            form_errors["category"] = "Veuillez sélectionner une catégorie valide."
        if form_data["priority"] not in {"standard", "attention", "urgent"}:
            form_errors["priority"] = "Veuillez sélectionner une priorité valide."
        if form_data["email"] and "@" not in form_data["email"]:
            form_errors["email"] = "Veuillez renseigner une adresse e-mail valide."

        structure_id = None
        if not _is_global_admin():
            try:
                structure_id = _current_structure_id()
            except Exception:
                form_errors["structure_id"] = (
                    "Impossible de déterminer la structure active."
                )
        elif form_data["structure_id"]:
            try:
                structure_id = int(form_data["structure_id"])
            except Exception:
                form_errors["structure_id"] = "Structure invalide."
            else:
                if not db.session.get(Structure, structure_id):
                    form_errors["structure_id"] = "Structure invalide."
        else:
            try:
                structure_id = _current_structure_id()
            except Exception:
                form_errors["structure_id"] = (
                    "Impossible de déterminer la structure active."
                )

        owner_id = None
        if form_data["owner_id"]:
            try:
                owner_id = int(form_data["owner_id"])
            except Exception:
                form_errors["owner_id"] = "Responsable initial invalide."
            else:
                if not db.session.get(AdminUser, owner_id):
                    form_errors["owner_id"] = "Responsable initial invalide."

        if form_errors:
            flash("Veuillez corriger les champs indiqués.", "warning")
        else:
            requester_user = _ensure_internal_requester_user()
            priority_map = {
                "standard": "medium",
                "attention": "high",
                "urgent": "urgent",
            }
            req = Request(
                title=form_data["title"],
                description=form_data["description"],
                name=form_data["person_name"],
                email=form_data["email"] or None,
                phone=form_data["phone"] or None,
                address_line=form_data["address_line"] or None,
                postcode=form_data["postcode"] or None,
                city=form_data["city"],
                country=form_data["country"] or None,
                location_text=request_address_display_text(
                    address_line=form_data["address_line"] or None,
                    postcode=form_data["postcode"] or None,
                    city=form_data["city"] or None,
                    country=form_data["country"] or None,
                ),
                category=form_data["category"],
                priority=priority_map.get(form_data["priority"], "medium"),
                status="pending",
                structure_id=structure_id,
                owner_id=owner_id,
                message=form_data["internal_notes"] or None,
                user_id=requester_user.id,
            )
            db.session.add(req)
            db.session.commit()
            audit_admin_action(
                action="CREATE_REQUEST",
                target_type="Request",
                target_id=req.id,
                payload={
                    "structure_id": req.structure_id,
                    "owner_id": req.owner_id,
                    "status": req.status,
                    "priority": req.priority,
                    "category": req.category,
                },
            )
            flash("Demande créée avec succès.", "success")
            return redirect(
                url_for("admin.admin_request_details", req_id=req.id),
                code=303,
            )

    return render_template(
        "admin/request_new.html",
        form_data=form_data,
        form_errors=form_errors,
        categories=categories,
        structures=structures,
        admins=admins,
        current_structure_id=_current_structure_id(),
    )


def apply_risk_filter(base_query, risk: str, now: datetime):
    closed_statuses = {"done", "cancelled", "rejected"}
    open_filter = or_(
        Request.status.is_(None),
        ~func.lower(Request.status).in_(closed_statuses),
    )
    notseen_hours = _notseen_hours_from_risk(risk)
    if (
        risk
        not in {
            "stale",
            "unassigned",
            "assigned_recent",
            "volunteer_stale",
            "sla_resolve_breach",
            "sla_assign_breach",
        }
        and notseen_hours is None
    ):
        return base_query

    if risk == "stale":
        base_query = base_query.filter(Request.created_at < (now - timedelta(days=8)))
        base_query = base_query.filter(open_filter)
    elif risk == "unassigned":
        base_query = base_query.filter(
            Request.created_at < (now - timedelta(days=3)),
            Request.assigned_volunteer_id.is_(None),
        )
        base_query = base_query.filter(open_filter)
    elif risk == "assigned_recent":
        base_query = base_query.filter(
            Request.created_at >= (now - timedelta(days=7)),
            Request.assigned_volunteer_id.isnot(None),
        )
    elif risk == "volunteer_stale":
        base_query = base_query.filter(
            Request.created_at < (
                now - timedelta(hours=VOLUNTEER_ASSIGN_SLA_HOURS)
            ),
            Request.assigned_volunteer_id.is_(None),
        )
        base_query = base_query.filter(open_filter)
    elif risk == "sla_resolve_breach":
        base_query = base_query.filter(
            Request.created_at < (now - timedelta(days=RESOLVE_SLA_DAYS)),
            Request.completed_at.is_(None),
        )
        base_query = base_query.filter(open_filter)
    elif risk == "sla_assign_breach":
        base_query = base_query.filter(
            Request.created_at < (now - timedelta(hours=ASSIGN_SLA_HOURS)),
            Request.owned_at.is_(None),
        )
        base_query = base_query.filter(open_filter)
    elif notseen_hours is not None:
        notseen_subq, _source = _build_notseen_subquery(now, hours=notseen_hours)
        base_query = base_query.filter(Request.id.in_(select(notseen_subq.c.request_id)))
        base_query = base_query.filter(open_filter)
    return base_query


def build_requests_query(base_query, request_args, legacy: bool = False):
    base_query = _scope_requests(base_query)
    status = (request_args.get("status") or "").strip()
    q = (request_args.get("q") or "").strip()
    category = normalize_request_category((request_args.get("category") or "").strip())
    risk = (request_args.get("risk") or "").strip().lower()
    risk_level = (request_args.get("risk_level") or "").strip().lower()
    no_owner = (request_args.get("no_owner") or "").strip() == "1"
    not_seen_72h = (request_args.get("not_seen_72h") or "").strip() == "1"
    sort = (request_args.get("sort") or "").strip().lower()
    show_deleted = (request_args.get("deleted") or "").strip() == "1"
    queue = (request_args.get("queue") or "").strip().lower()
    sla_kind = _normalize_sla_kind(request_args.get("sla_kind"))
    sla_days = _normalize_sla_days(request_args.get("sla_days", 30))
    now = datetime.now(UTC).replace(tzinfo=None)

    if show_deleted:
        base_query = base_query.filter(Request.deleted_at.isnot(None))
    else:
        base_query = base_query.filter(Request.deleted_at.is_(None))

    if status:
        internal = "pending" if status == "new" else status
        base_query = base_query.filter(Request.status == internal)
    if q:
        like = f"%{q}%"
        base_query = base_query.filter(
            or_(
                Request.title.ilike(like),
                Request.name.ilike(like),
                Request.email.ilike(like),
                Request.phone.ilike(like),
                Request.description.ilike(like),
            )
        )
    if category:
        category_variants = {category}
        for legacy_code in ("general", "social", "medical", "tech", "admin", "other"):
            if normalize_request_category(legacy_code) == category:
                category_variants.add(legacy_code)
        base_query = base_query.filter(
            func.lower(Request.category).in_([c.lower() for c in category_variants])
        )

    base_query = apply_risk_filter(base_query, risk, now)
    if risk_level in {"critical", "attention", "standard"}:
        base_query = base_query.filter(
            func.lower(func.coalesce(Request.risk_level, "standard")) == risk_level
        )
    if no_owner:
        base_query = base_query.filter(
            func.lower(func.coalesce(Request.risk_signals, "")).like("%no_owner%")
        )
    if not_seen_72h:
        base_query = base_query.filter(
            func.lower(func.coalesce(Request.risk_signals, "")).like("%not_seen_72h%")
        )
    if queue == "sla" and sla_kind:
        base_query = _apply_sla_queue_filter(
            base_query,
            sla_kind=sla_kind,
            days=sla_days,
            now=now,
        )
    created_sort_col = func.coalesce(Request.created_at, Request.updated_at)
    if sort == "created_asc":
        base_query = base_query.order_by(created_sort_col.asc(), Request.id.asc())
    elif sort == "created_desc":
        base_query = base_query.order_by(created_sort_col.desc(), Request.id.desc())
    elif sort == "risk_asc":
        base_query = base_query.order_by(
            func.coalesce(Request.risk_score, 0).asc(),
            created_sort_col.desc(),
            Request.id.desc(),
        )
    else:
        base_query = base_query.order_by(
            func.coalesce(Request.risk_score, 0).desc(),
            created_sort_col.desc(),
            Request.id.desc(),
        )
    result = (base_query, status, q, risk, risk_level, no_owner, not_seen_72h, sort)
    if legacy:
        return result[:4]
    return result


@admin_bp.get("/requests/export.csv")
@admin_required
def admin_requests_export_csv():
    admin_required_404()
    query, *_ = build_requests_query(Request.query, request.args)
    rows = query.limit(5000).all()
    log_activity(
        entity_type="export",
        entity_id=0,
        action="requests_export",
        message="Requests export",
        meta={"format": "csv", "anonymized": False},
        persist=True,
    )

    out = StringIO()
    writer = csv.writer(out, delimiter=";")
    writer.writerow(
        [
            "id",
            "created_at",
            "status",
            "priority",
            "category",
            "title",
            "name",
            "email",
            "phone",
            "owner_id",
            "owned_at",
            "completed_at",
        ]
    )

    for r in rows:
        writer.writerow(
            [
                r.id,
                getattr(r, "created_at", "") or "",
                getattr(r, "status", "") or "",
                getattr(r, "priority", "") or "",
                getattr(r, "category", "") or "",
                getattr(r, "title", "") or "",
                getattr(r, "name", "") or "",
                getattr(r, "email", "") or "",
                getattr(r, "phone", "") or "",
                getattr(r, "owner_id", "") or "",
                getattr(r, "owned_at", "") or "",
                getattr(r, "completed_at", "") or "",
            ]
        )

    filename = (
        f"helpchain_requests_"
        f"{datetime.now(UTC).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.csv"
    )
    return Response(
        "\ufeff" + out.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@admin_bp.get("/requests/export.xlsx")
@admin_required
def admin_requests_export_xlsx():
    admin_required_404()
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return Response("openpyxl is not installed", status=500)

    query, *_ = build_requests_query(Request.query, request.args)
    rows = query.limit(5000).all()
    log_activity(
        entity_type="export",
        entity_id=0,
        action="requests_export",
        message="Requests export",
        meta={"format": "xlsx", "anonymized": False},
        persist=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Requests"
    headers = [
        "id",
        "created_at",
        "status",
        "priority",
        "category",
        "title",
        "name",
        "email",
        "phone",
        "owner_id",
        "owned_at",
        "completed_at",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                r.id,
                getattr(r, "created_at", None),
                getattr(r, "status", None),
                getattr(r, "priority", None),
                getattr(r, "category", None),
                getattr(r, "title", None),
                getattr(r, "name", None),
                getattr(r, "email", None),
                getattr(r, "phone", None),
                getattr(r, "owner_id", None),
                getattr(r, "owned_at", None),
                getattr(r, "completed_at", None),
            ]
        )

    phone_col = headers.index("phone") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=phone_col)
        if cell.value is not None:
            cell.value = str(cell.value)
        cell.number_format = "@"

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(10, max_len + 2),
            60,
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = (
        f"helpchain_requests_"
        f"{datetime.now(UTC).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.get("/requests/export_anonymized.csv")
@admin_required
def admin_requests_export_csv_anonymized():
    admin_required_404()
    query, *_ = build_requests_query(Request.query, request.args)
    rows = query.limit(5000).all()
    log_activity(
        entity_type="export",
        entity_id=0,
        action="requests_export",
        message="Requests export (anonymized)",
        meta={"format": "csv", "anonymized": True},
        persist=True,
    )

    out = StringIO()
    writer = csv.writer(out, delimiter=";")
    writer.writerow(
        [
            "id",
            "created_at",
            "status",
            "priority",
            "category",
            "owner_id",
            "owned_at",
            "completed_at",
        ]
    )

    for r in rows:
        writer.writerow(
            [
                r.id,
                getattr(r, "created_at", "") or "",
                getattr(r, "status", "") or "",
                getattr(r, "priority", "") or "",
                getattr(r, "category", "") or "",
                getattr(r, "owner_id", "") or "",
                getattr(r, "owned_at", "") or "",
                getattr(r, "completed_at", "") or "",
            ]
        )

    filename = (
        f"helpchain_requests_ANON_"
        f"{datetime.now(UTC).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.csv"
    )
    return Response(
        "\ufeff" + out.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@admin_bp.get("/requests/export_anonymized.xlsx")
@admin_required
def admin_requests_export_xlsx_anonymized():
    admin_required_404()
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return Response("openpyxl is not installed", status=500)

    query, *_ = build_requests_query(Request.query, request.args)
    rows = query.limit(5000).all()
    log_activity(
        entity_type="export",
        entity_id=0,
        action="requests_export",
        message="Requests export (anonymized)",
        meta={"format": "xlsx", "anonymized": True},
        persist=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Requests (Anon)"
    headers = [
        "id",
        "created_at",
        "status",
        "priority",
        "category",
        "owner_id",
        "owned_at",
        "completed_at",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                r.id,
                getattr(r, "created_at", None),
                getattr(r, "status", None),
                getattr(r, "priority", None),
                getattr(r, "category", None),
                getattr(r, "owner_id", None),
                getattr(r, "owned_at", None),
                getattr(r, "completed_at", None),
            ]
        )

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(10, max_len + 2),
            60,
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = (
        f"helpchain_requests_ANON_"
        f"{datetime.now(UTC).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.get("/requests/<int:req_id>")
@admin_required
@admin_role_required("superadmin")
def admin_request_details(req_id: int):
    admin_required_404()
    activities_supported = _table_has_column("request_activities", "volunteer_id")
    request_logs_supported = _table_exists("request_logs")
    if activities_supported and request_logs_supported:
        req = (
            _scope_requests(Request.query)
            .options(joinedload(Request.logs), joinedload(Request.activities))
            .filter(Request.id == req_id)
            .first_or_404()
        )
    elif activities_supported and not request_logs_supported:
        req = (
            _scope_requests(Request.query)
            .options(joinedload(Request.activities))
            .filter(Request.id == req_id)
            .first_or_404()
        )
    elif (not activities_supported) and request_logs_supported:
        req = (
            _scope_requests(Request.query)
            .options(joinedload(Request.logs))
            .filter(Request.id == req_id)
            .first_or_404()
        )
    else:
        req = (
            _scope_requests(Request.query)
            .filter(Request.id == req_id)
            .first_or_404()
        )
    linked_case = None
    if _cases_enabled():
        linked_case = Case.query.filter(Case.request_id == req.id).first()
    risk_ai_suggestion = _build_risk_ai_suggestion(req)
    operational_blockages = _build_operational_blockages(req, linked_case)
    volunteer_request_states_supported = _table_exists("volunteer_request_states")
    volunteer_interests_supported = _table_exists("volunteer_interests")
    volunteer_actions_supported = _table_exists("volunteer_actions")
    notifications_supported = _table_exists("notifications")
    admin_id = current_user.id
    now = _now_utc()
    latest_actions = []
    audit_logs = []
    if activities_supported:
        latest_actions = (
            RequestActivity.query.filter_by(request_id=req_id)
            .filter(
                RequestActivity.action.in_(
                    ["volunteer_can_help", "volunteer_cant_help"]
                )
            )
            .order_by(RequestActivity.created_at.desc())
            .limit(10)
            .all()
        )
    linked_volunteer_ids = []
    if volunteer_request_states_supported:
        linked_volunteer_ids = [
            int(v_id)
            for (v_id,) in db.session.query(VolunteerRequestState.volunteer_id)
            .filter(VolunteerRequestState.request_id == req_id)
            .distinct()
            .all()
            if v_id is not None
        ]
    volunteer_engagement = []
    if linked_volunteer_ids:
        linked_volunteers = {
            v.id: v
            for v in Volunteer.query.filter(Volunteer.id.in_(linked_volunteer_ids)).all()
        }
        for v_id in linked_volunteer_ids:
            try:
                score_row = get_volunteer_engagement_score(v_id, now=now)
            except Exception:
                db.session.rollback()
                score_row = {
                    "volunteer_id": int(v_id),
                    "score": 0,
                    "label": "At risk",
                    "seen_within_24h": 0,
                    "not_seen_72h": 0,
                    "can_help": 0,
                    "cant_help": 0,
                }
            v = linked_volunteers.get(v_id)
            display = (
                (
                    getattr(v, "name", None)
                    or getattr(v, "email", None)
                    or f"Volunteer #{v_id}"
                )
                if v is not None
                else f"Volunteer #{v_id}"
            )
            score_row["display"] = display
            volunteer_engagement.append(score_row)
        volunteer_engagement.sort(key=lambda x: (-x["score"], x["volunteer_id"]))

    locked_by = None
    if req.owner_id is None:
        req.owner_id = admin_id
        req.owned_at = now
        if activities_supported:
            db.session.add(
                RequestActivity(
                    request_id=req.id,
                    actor_admin_id=admin_id,
                    action="lock",
                    old_value="",
                    new_value=str(admin_id),
                    created_at=now,
                )
            )
        db.session.commit()
    elif req.owner_id == admin_id:
        if _lock_expired(req, now):
            req.owned_at = now
            db.session.commit()
    else:
        if _lock_expired(req, now):
            old_owner = req.owner_id
            req.owner_id = admin_id
            req.owned_at = now
            if activities_supported:
                db.session.add(
                    RequestActivity(
                        request_id=req.id,
                        actor_admin_id=admin_id,
                        action="lock",
                        old_value=str(old_owner),
                        new_value=str(admin_id),
                        created_at=now,
                    )
                )
            db.session.commit()
        else:
            locked_by = req.owner_id
            activities = []
            if activities_supported:
                activities = sorted(
                    (req.activities or []),
                    key=lambda a: a.created_at or datetime.min,
                    reverse=True,
                )[:50]
            interests = []
            if volunteer_interests_supported:
                interests = (
                    VolunteerInterest.query.filter_by(request_id=req_id)
                    .order_by(VolunteerInterest.created_at.desc())
                    .all()
                )
            locked_recommendation = compute_recommendation(req)
            return (
                render_template(
                    "admin/request_details.html",
                    req=req,
                    activities=activities,
                    logs=(req.logs if request_logs_supported else []),
                    STATUS_LABELS_BG=STATUS_LABELS_BG,
                    is_stale=is_stale,
                    interests=interests,
                    latest_actions=latest_actions,
                    volunteer_engagement=volunteer_engagement,
                    audit_logs=audit_logs,
                    case_signals=_compute_case_signals(req, activities, now),
                    recommendation=locked_recommendation,
                    helpchain_recommendation=_build_helpchain_recommendation(
                        req,
                        activities,
                        now,
                    ),
                    case_summary=build_case_summary(req, locked_recommendation),
                    risk_ai_suggestion=risk_ai_suggestion,
                    operational_blockages=operational_blockages,
                    linked_case=linked_case,
                    is_locked=True,
                    locked_by=locked_by,
                ),
                200,
            )
    is_locked = False
    logs = req.logs if request_logs_supported else []
    activities = []
    if activities_supported:
        activities = sorted(
            (req.activities or []),
            key=lambda a: a.created_at or datetime.min,
            reverse=True,
        )[:50]
    interests = []
    if volunteer_interests_supported:
        interests = (
            VolunteerInterest.query.filter_by(request_id=req_id)
            .order_by(VolunteerInterest.created_at.desc())
            .all()
        )

    req_city = (getattr(req, "city", "") or "").strip().lower()

    def _norm_city(val: str) -> str:
        return (val or "").strip().lower()

    vols = Volunteer.query.filter_by(is_active=True).all()
    matched_volunteers = [
        v for v in vols if _norm_city(getattr(v, "location", None)) == req_city
    ]
    matched_volunteers = matched_volunteers[:20]
    matched_volunteer_ids = [v.id for v in matched_volunteers]
    notif_rows = []
    if matched_volunteer_ids and notifications_supported:
        notif_rows = Notification.query.filter(
            Notification.request_id == req.id,
            Notification.type == "new_match",
            Notification.volunteer_id.in_(matched_volunteer_ids),
        ).all()
    notified_count = len(notif_rows)
    seen_count = sum(1 for n in notif_rows if getattr(n, "is_read", False))

    interest_rows = interests
    interested_ids = {i.volunteer_id for i in interest_rows}
    interested_count = len(interested_ids)

    notif_by_vol = {n.volunteer_id: n for n in notif_rows}
    flags_by_vol = {}
    for v in matched_volunteers:
        n = notif_by_vol.get(v.id)
        flags_by_vol[v.id] = {
            "notified": n is not None,
            "seen": bool(getattr(n, "is_read", False)) if n else False,
            "interested": v.id in interested_ids,
        }

    assigned_volunteer = None
    if getattr(req, "assigned_volunteer_id", None):
        assigned_volunteer = db.session.get(Volunteer, req.assigned_volunteer_id)

    volunteer_signals = []
    if volunteer_actions_supported:
        volunteer_signals = (
            VolunteerAction.query.filter_by(request_id=req.id)
            .order_by(VolunteerAction.updated_at.desc())
            .all()
        )
    last_vol_signal = volunteer_signals[0] if volunteer_signals else None
    signal_vol_ids = [va.volunteer_id for va in volunteer_signals]
    volunteers_map = (
        {v.id: v for v in Volunteer.query.filter(Volunteer.id.in_(signal_vol_ids)).all()}
        if signal_vol_ids
        else {}
    )
    can_help_count = sum(1 for va in volunteer_signals if va.action == "CAN_HELP")
    cant_help_count = sum(1 for va in volunteer_signals if va.action == "CANT_HELP")
    case_signals = _compute_case_signals(req, activities, now)
    recommendation = compute_recommendation(req)
    helpchain_recommendation = _build_helpchain_recommendation(req, activities, now)
    case_summary = build_case_summary(req, recommendation)

    return (
        render_template(
            "admin/request_details.html",
            req=req,
            activities=activities,
            logs=logs,
            STATUS_LABELS_BG=STATUS_LABELS_BG,
            is_stale=is_stale,
            interests=interests,
            is_locked=is_locked,
            locked_by=locked_by,
            matched_volunteers=matched_volunteers,
            matched_count=len(matched_volunteers),
            notified_count=notified_count,
            seen_count=seen_count,
            interested_count=interested_count,
            flags_by_vol=flags_by_vol,
            assigned_volunteer=assigned_volunteer,
            volunteer_signals=volunteer_signals,
            last_vol_signal=last_vol_signal,
            volunteers_map=volunteers_map,
            can_help_count=can_help_count,
            cant_help_count=cant_help_count,
            latest_actions=latest_actions,
            volunteer_engagement=volunteer_engagement,
            audit_logs=audit_logs,
            case_signals=case_signals,
            recommendation=recommendation,
            helpchain_recommendation=helpchain_recommendation,
            case_summary=case_summary,
            risk_ai_suggestion=risk_ai_suggestion,
            operational_blockages=operational_blockages,
            linked_case=linked_case,
        ),
        200,
    )


def _priority_with_manual_guard(
    current_priority: str | None,
    derived_priority: str | None,
) -> str:
    current_rank = CASE_PRIORITY_RANK.get((current_priority or "").strip().lower(), -1)
    derived_rank = CASE_PRIORITY_RANK.get((derived_priority or "").strip().lower(), -1)
    if current_rank > derived_rank:
        return (current_priority or "").strip().lower() or "normal"
    return (derived_priority or "").strip().lower() or "normal"


@admin_bp.post("/requests/<int:req_id>/open-case")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_open_case_from_request(req_id: int):
    admin_required_404()
    if not _cases_enabled():
        flash("Case system tables are not available yet. Run migrations first.", "warning")
        return redirect(url_for("admin.admin_request_details", req_id=req_id), code=303)

    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    existing = Case.query.filter(Case.request_id == req.id).first()
    if existing:
        return redirect(url_for("admin.admin_case_detail", case_id=existing.id), code=303)

    now = _now_utc()
    triage = score_request_risk(req)
    derived_priority = _priority_with_manual_guard("normal", triage.get("priority"))
    case_row = Case(
        request_id=req.id,
        structure_id=getattr(req, "structure_id", None),
        owner_user_id=None,
        assigned_professional_lead_id=None,
        latitude=getattr(req, "latitude", None),
        longitude=getattr(req, "longitude", None),
        status="new",
        priority=derived_priority,
        risk_score=int(triage.get("score") or 0),
        opened_at=now,
        assigned_at=None,
        resolved_at=None,
        closed_at=None,
        last_activity_at=now,
        created_at=now,
        updated_at=now,
    )
    db.session.add(case_row)
    db.session.flush()
    _append_case_event(
        case_id=case_row.id,
        actor_user_id=getattr(current_user, "id", None),
        event_type="case_created",
        message=f"Case created from request #{req.id}",
        metadata={"request_id": req.id},
    )
    _append_case_event(
        case_id=case_row.id,
        actor_user_id=getattr(current_user, "id", None),
        event_type="triage_scored",
        message=(
            f"Triage scored at {int(triage.get('score') or 0)}/100 "
            f"({risk_label_from_score(int(triage.get('score') or 0))})"
        ),
        metadata={
            "risk_score": int(triage.get("score") or 0),
            "risk_label": risk_label_from_score(int(triage.get("score") or 0)),
            "derived_priority": derived_priority,
            "matched_rules": triage.get("matched_rules") or [],
            "suggested_category_code": triage.get("suggested_category_code"),
            "suggested_category_label": triage.get("suggested_category_label"),
        },
    )
    update_case_risk(case_row)
    evaluate_case_alerts(case_row)
    db.session.commit()
    flash(f"Case #{case_row.id} opened.", "success")
    return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)


@admin_bp.post("/requests/<int:req_id>/unlock", endpoint="admin_request_unlock")
@admin_required
@admin_role_required("superadmin")
def admin_request_unlock(req_id: int):
    admin_required_404()
    admin_id = _admin_id()
    if not admin_id:
        abort(403)

    req = db.session.get(Request, req_id)
    if not req:
        abort(404)

    old_owner = req.owner_id
    if old_owner is not None:
        req.owner_id = None
        req.owned_at = None
        db.session.add(
            RequestActivity(
                request_id=req.id,
                actor_admin_id=admin_id,
                action="unlock",
                old_value=str(old_owner),
                new_value="",
                created_at=_now_utc(),
            )
        )
        db.session.commit()
        audit_admin_action(
            action="request.unlock",
            target_type="Request",
            target_id=req.id,
            payload={
                "req_id": req.id,
                "old": {"locked": True, "owner_id": old_owner},
                "new": {"locked": False, "owner_id": None},
            },
        )

    flash("Unlocked.", "success")
    return redirect(url_for("admin.admin_request_details", req_id=req_id))


@admin_bp.post(
    "/requests/<int:req_id>/interests/<int:interest_id>/approve",
    endpoint="admin_interest_approve",
)
@admin_required
@admin_role_required("ops", "superadmin")
def admin_interest_approve(req_id: int, interest_id: int):
    current_app.logger.info(
        "ADMIN_APPROVE HIT req_id=%s interest_id=%s",
        req_id,
        interest_id,
    )

    admin_required_404()
    admin = current_user
    admin_id = getattr(admin, "id", None)

    req = db.session.get(Request, req_id)
    if not req:
        abort(404)

    if _locked_by_other(req, admin_id):
        flash(
            "🔒 Заявката е заключена от друг админ. Може да я отключите ръчно.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    vi = db.session.get(VolunteerInterest, interest_id)
    if not vi or vi.request_id != req_id:
        abort(404)

    if _is_request_locked(req):
        flash(
            "🔒 Заявката е заключена (done/cancelled). Смени статуса, за да отключиш действията.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    old_vi = vi.status
    changed_vi = old_vi != "approved"
    if changed_vi:
        vi.status = "approved"
        db.session.add(
            RequestActivity(
                request_id=req_id,
                actor_admin_id=admin_id,
                action="volunteer_interest_approved",
                old_value=old_vi,
                new_value="approved",
            )
        )

    old_rs = req.status
    status_changed = False
    if old_rs == "open":
        req.status = "in_progress"
        _log_status_change_once(req_id, old_rs, req.status, admin_id)
        status_changed = True

    if changed_vi or status_changed:
        current_app.logger.info(
            "BEFORE commit: req.status=%s vi.status=%s",
            req.status,
            vi.status,
        )
        db.session.commit()
        if changed_vi:
            audit_admin_action(
                action="interest.approve",
                target_type="Interest",
                target_id=vi.id,
                payload={
                    "req_id": req.id,
                    "interest_id": vi.id,
                    "old": {"status": old_vi},
                    "new": {"status": vi.status},
                },
            )
        current_app.logger.info(
            "AFTER commit: req.status=%s vi.status=%s",
            req.status,
            vi.status,
        )
        flash("Approved.", "success")
    else:
        flash("No changes.", "info")

    return redirect(url_for("admin.admin_request_details", req_id=req_id))


@admin_bp.post(
    "/requests/<int:req_id>/interests/<int:interest_id>/reject",
    endpoint="admin_interest_reject",
)
@admin_required
@admin_role_required("ops", "superadmin")
def admin_interest_reject(req_id: int, interest_id: int):
    admin_required_404()
    admin = current_user
    admin_id = getattr(admin, "id", None)

    req = db.session.get(Request, req_id)
    if not req:
        abort(404)

    if _locked_by_other(req, admin_id):
        flash(
            "🔒 Заявката е заключена от друг админ. Може да я отключите ръчно.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    interest = db.session.get(VolunteerInterest, interest_id)
    if not interest or interest.request_id != req_id:
        abort(404)

    if _is_request_locked(req):
        flash(
            "🔒 Заявката е заключена (done/cancelled). Смени статуса, за да отключиш действията.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    old_vi = interest.status
    if old_vi == "rejected":
        flash("No changes.", "info")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    current_app.logger.info(
        "ADMIN_REJECT HIT req_id=%s interest_id=%s",
        req_id,
        interest_id,
    )

    reject_reason = (
        request.form.get("reason")
        or request.form.get("reject_reason")
        or request.form.get("note")
        or ""
    ).strip()
    interest.status = "rejected"
    db.session.add(
        RequestActivity(
            request_id=req.id,
            actor_admin_id=admin_id,
            action="volunteer_interest_rejected",
            old_value=old_vi,
            new_value="rejected",
        )
    )

    db.session.commit()
    payload = {
        "req_id": req.id,
        "interest_id": interest.id,
        "old": {"status": old_vi},
        "new": {"status": interest.status},
    }
    if reject_reason:
        payload["reason"] = reject_reason
    audit_admin_action(
        action="interest.reject",
        target_type="Interest",
        target_id=interest.id,
        payload=payload,
    )
    flash("Rejected.", "warning")
    return redirect(url_for("admin.admin_request_details", req_id=req_id))


@admin_bp.post("/requests/<int:req_id>/assign", endpoint="admin_request_assign")
@login_required
@admin_required
@admin_role_required("ops", "superadmin")
def admin_request_assign(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if _locked_by_other(req, getattr(current_user, "id", None)):
        flash(
            "🔒 Заявката е заключена от друг админ. Може да я отключите ръчно.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if _is_request_locked(req):
        flash(
            "This request is locked (done/cancelled). Unlock it by changing status first.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if req.owner_id and req.owner_id != getattr(current_user, "id", None):
        flash("Deja pris en charge.", "warning")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if req.owner_id == getattr(current_user, "id", None):
        flash("Deja assigne a vous.", "info")
        next_url = (request.form.get("next") or "").strip()
        if next_url and is_safe_url(next_url):
            return redirect(next_url)
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    takeover = False
    old_owner = req.owner_id
    req.owner_id = current_user.id
    req.owned_at = utc_now()
    if _table_exists("request_metrics"):
        metric = db.session.query(RequestMetric).filter_by(request_id=req.id).first()
        if metric is None:
            metric = RequestMetric(request_id=req.id)
            db.session.add(metric)
        if metric.time_to_assign is None and req.created_at:
            try:
                metric.time_to_assign = int(
                    (utc_now() - req.created_at).total_seconds()
                )
            except Exception:
                pass
    action_name = "takeover" if takeover else "assign"
    reason = None
    if takeover and req.owned_at:
        try:
            hours = (utc_now() - req.owned_at).total_seconds() / 3600
            reason = f"stale: {hours:.1f}h"
        except Exception:
            reason = "stale"
    new_val = (
        f"{current_user.id}" if reason is None else f"{current_user.id} ({reason})"
    )
    log_request_activity(
        req,
        action_name,
        old=old_owner,
        new=new_val,
        actor_admin_id=current_user.id,
    )
    _audit_request(
        req.id,
        action="assign_owner",
        message="Owner assigned",
        old=str(old_owner) if old_owner is not None else None,
        new=str(current_user.id),
    )
    db.session.commit()
    audit_admin_action(
        action="ASSIGN_OPERATOR",
        target_type="Request",
        target_id=req.id,
        payload={
            "old": {"owner_id": old_owner},
            "new": {"owner_id": current_user.id},
        },
    )
    flash(_("The request has been assigned to you."), "success")
    next_url = (request.form.get("next") or "").strip()
    if next_url and is_safe_url(next_url):
        return redirect(next_url)
    return redirect(url_for("admin.admin_request_details", req_id=req_id))


@admin_bp.post("/requests/<int:req_id>/unassign", endpoint="admin_request_unassign")
@login_required
@admin_required
@admin_role_required("ops", "superadmin")
def admin_request_unassign(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if _locked_by_other(req, getattr(current_user, "id", None)):
        flash(
            "🔒 Заявката е заключена от друг админ. Може да я отключите ръчно.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if _is_request_locked(req):
        flash(
            "This request is locked (done/cancelled). Unlock it by changing status first.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if not can_edit_request(req, current_user):
        abort(403)
    old_owner = req.owner_id
    req.owner_id = None
    req.owned_at = None
    log_request_activity(
        req,
        "unassign",
        old=old_owner,
        new=None,
        actor_admin_id=getattr(current_user, "id", None),
    )
    _audit_request(
        req.id,
        action="unassign_owner",
        message="Owner unassigned",
        old=str(old_owner) if old_owner is not None else None,
        new=None,
    )
    db.session.commit()
    audit_admin_action(
        action="request.unassign_owner",
        target_type="Request",
        target_id=req.id,
        payload={
            "old": {"owner_id": old_owner},
            "new": {"owner_id": None},
        },
    )
    flash(_("Owner removed."), "info")
    return redirect(url_for("admin.admin_request_details", req_id=req_id))


@admin_bp.post(
    "/requests/<int:req_id>/assign_volunteer/<int:volunteer_id>",
    endpoint="admin_assign_volunteer",
)
@login_required
def admin_assign_volunteer(req_id: int, volunteer_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)
    if _is_request_locked(req):
        flash("This request is locked (done/cancelled).", "warning")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    req.assigned_volunteer_id = volunteer_id
    log_request_activity(
        req,
        "assign_volunteer",
        old=getattr(req, "assigned_volunteer_id", None),
        new=volunteer_id,
        actor_admin_id=getattr(current_user, "id", None),
    )
    db.session.commit()
    flash("Assigned to volunteer.", "success")
    return redirect(url_for("admin.admin_request_details", req_id=req.id))


@admin_bp.post(
    "/requests/<int:req_id>/unassign_volunteer",
    endpoint="admin_unassign_volunteer",
)
@login_required
def admin_unassign_volunteer(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)
    old_val = getattr(req, "assigned_volunteer_id", None)
    req.assigned_volunteer_id = None
    log_request_activity(
        req,
        "unassign_volunteer",
        old=old_val,
        new=None,
        actor_admin_id=getattr(current_user, "id", None),
    )
    db.session.commit()
    flash("Volunteer unassigned.", "info")
    return redirect(url_for("admin.admin_request_details", req_id=req.id))


@admin_bp.post("/requests/<int:req_id>/nudge", endpoint="admin_request_nudge")
@login_required
def admin_request_nudge(req_id: int):
    admin_required_404()
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)

    volunteer_id = getattr(req, "assigned_volunteer_id", None)
    if not volunteer_id:
        flash("No assigned volunteer to nudge.", "warning")
        return redirect(request.referrer or url_for("admin.admin_requests"))

    created = send_nudge_notification(
        request_id=req.id,
        volunteer_id=int(volunteer_id),
        actor_admin_id=getattr(current_user, "id", None),
    )
    if created:
        flash("Nudge sent.", "success")
    else:
        flash("Nudge suppressed (recently sent).", "info")
    return redirect(request.referrer or url_for("admin.admin_requests"))


@admin_bp.post("/requests/<int:req_id>/delete", endpoint="admin_request_delete")
@login_required
def admin_request_delete(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)

    if not getattr(req, "is_archived", False):
        flash(
            "Archive the request first. Only archived requests can be deleted.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    if getattr(req, "deleted_at", None) is None:
        req.deleted_at = utc_now()
        req.is_archived = True
        if getattr(req, "archived_at", None) is None:
            req.archived_at = req.deleted_at
        log_request_activity(
            req,
            "delete",
            old=None,
            new=str(req.deleted_at),
            actor_admin_id=getattr(current_user, "id", None),
        )
        db.session.commit()
        flash("Request moved to Deleted.", "success")

    return redirect(url_for("admin.admin_request_details", req_id=req.id))


@admin_bp.post(
    "/requests/<int:req_id>/restore-deleted",
    endpoint="admin_request_restore_deleted",
)
@login_required
def admin_request_restore_deleted(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)

    if getattr(req, "deleted_at", None) is not None:
        old = req.deleted_at
        req.deleted_at = None
        req.is_archived = True
        if getattr(req, "archived_at", None) is None:
            req.archived_at = utc_now()
        log_request_activity(
            req,
            "restore_deleted",
            old=str(old),
            new=None,
            actor_admin_id=getattr(current_user, "id", None),
        )
        db.session.commit()
        flash("Request restored from Deleted (kept archived).", "success")

    return redirect(url_for("admin.admin_request_details", req_id=req.id))


@admin_bp.post("/requests/<int:req_id>/note")
@login_required
def admin_request_add_note(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    note = (request.form.get("note") or "").strip()
    if not note:
        flash("Note is empty.", "warning")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if len(note) > 2000:
        flash("Note is too long (max 2000 chars).", "danger")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    log_request_activity(
        req,
        "note",
        old=None,
        new=note,
        actor_admin_id=getattr(current_user, "id", None),
    )
    _audit_request(
        req.id,
        action="note_add",
        message="Admin note added",
    )
    db.session.commit()
    flash("Note added.", "success")
    return redirect(url_for("admin.admin_request_details", req_id=req.id))


@admin_bp.get("/requests/<int:req_id>/status")
@login_required
def admin_request_status_get_alias(req_id: int):
    return redirect(url_for("admin.admin_request_details", req_id=req_id), code=302)


@admin_bp.get("/requests/<int:req_id>/notes")
@login_required
def admin_request_notes_get_alias(req_id: int):
    return redirect(url_for("admin.admin_request_details", req_id=req_id), code=302)


@admin_bp.post("/requests/<int:req_id>/notes")
@login_required
def admin_request_notes_post_alias(req_id: int):
    return admin_request_add_note(req_id)
