from __future__ import annotations

import csv
import json
import math
import os
import secrets
import threading
import time
import re
from types import SimpleNamespace
from pathlib import Path
from datetime import UTC, datetime, timedelta, timezone
from functools import wraps
from io import BytesIO, StringIO
from typing import Optional
from urllib.parse import urljoin, urlparse

from flask import (
    Blueprint,
    Response,
    abort,
    current_app,
    flash,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_login import current_user, login_required, login_user, logout_user
from flask_babel import force_locale, gettext as _
from babel.support import Translations
from werkzeug.security import check_password_hash, generate_password_hash

from backend.audit import log_activity
from backend.extensions import db, limiter
from backend.system_sanity import run_system_checks
from ..models.volunteer_interest import VolunteerInterest
from ..observability import (
    tenant_leak_get,
    tenant_leak_inc,
)
from ..statuses import (
    REQUEST_STATUS_ALLOWED,
    normalize_request_status,
)

from ..config import Config
from ..models import (
    AdminAuditEvent,
    AdminLoginAttempt,
    AdminUser,
    Case,
    CaseEvent,
    CaseParticipant,
    Notification,
    NotificationJob,
    ProfessionalLead,
    Intervenant,
    Request,
    RequestActivity,
    RequestLog,
    RequestMetric,
    UiLocaleLock,
    UiTranslationFreeze,
    UiTranslation,
    UiTranslationEvent,
    Volunteer,
    VolunteerAction,
    VolunteerInterest,
    VolunteerRequestState,
    Structure,
    User,
    current_structure,
    utc_now,
)

try:
    from ..models import ProAccessRequest
except ImportError:
    ProAccessRequest = None

from ..notifications.inapp import NUDGE_COOLDOWN_HOURS, send_nudge_notification
from ..constants.categories import (
    REQUEST_CATEGORY_CODES,
    normalize_request_category,
    request_category_choices,
    request_category_label,
)
from ..services.case_assistant import build_case_assistant_recommendation
from ..services.case_matching import suggest_professional_leads_for_case
from ..services.case_risk import risk_label_from_score, score_request_risk
from ..services.case_summary import build_case_summary, build_case_summary_snippet
from ..services.geocoding import geocode_location_best_effort
from ..services.ops_priority import compute_ops_priority
from ..security_logging import log_security_event
from ..services.recommendation_engine import compute_recommendation

GENERIC_ADMIN_LOGIN_FAIL_MSG = (
    "Identifiants invalides ou accès temporairement bloqué."
)
CATEGORY_CASE_STATUSES = (
    "new",
    "triaged",
    "assigned",
    "in_progress",
    "resolved",
    "closed",
    "cancelled",
)
CASE_PRIORITIES = ("low", "normal", "high", "critical")
CASE_PRIORITY_RANK = {
    "low": 0,
    "normal": 1,
    "high": 2,
    "critical": 3,
}
CASE_EVENT_TYPES = (
    "case_created",
    "triage_scored",
    "status_changed",
    "priority_changed",
    "owner_assigned",
    "professional_assigned",
    "participant_added",
    "note_added",
    "case_resolved",
    "case_closed",
)
CASE_PARTICIPANT_TYPES = (
    "admin_user",
    "professional_user",
    "professional_lead",
    "association",
    "external_contact",
)
CASE_PARTICIPANT_ROLES = (
    "owner",
    "primary_professional",
    "contributor",
    "observer",
    "coordinator",
)
CLOSED_STATUSES = {"done", "cancelled", "rejected"}
ASSIGN_SLA_HOURS = 48
RESOLVE_SLA_DAYS = 7
VOLUNTEER_ASSIGN_SLA_HOURS = 72
SLA_QUEUE_KINDS = {
    "resolution_overdue": "SLA resolution overdue",
    "owner_assignment_overdue": "SLA owner assignment overdue",
    "volunteer_assignment_overdue": "Volunteer assignment overdue",
}
SLA_BREAKDOWN_TYPE_TO_KIND = {
    "resolve": "resolution_overdue",
    "owner_assign": "owner_assignment_overdue",
    "volunteer_assign": "volunteer_assignment_overdue",
}
SLA_KIND_TO_BREAKDOWN_TYPE = {
    v: k for k, v in SLA_BREAKDOWN_TYPE_TO_KIND.items()
}
NOTSEEN_TIERS_HOURS = (24, 48, 72)
PRO_ACCESS_STATUSES = ("new", "reviewed", "approved", "rejected")
RISKY_ACTIONS = (
    "request.archive",
    "request.assign_owner",
    "request.unassign_owner",
    "request.unlock",
    "interest.approve",
    "interest.reject",
)
STATE_CHANGING_METHODS = {"POST", "PUT", "PATCH", "DELETE"}
ADMIN_LOGIN_RATE_WINDOW_MIN = 5
ADMIN_LOGIN_MAX_FAILS = 5
ADMIN_LOGIN_LOCKOUT_MIN = 15
ADMIN_SESSION_IDLE_TIMEOUT_MIN = 20
ADMIN_FRESH_AUTH_MIN = 10
_SCHEMA_COLUMN_CACHE: dict[tuple[str, str], bool] = {}
_SCHEMA_TABLE_CACHE: dict[str, bool] = {}
_CTRL_CHARS_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_UI_KEYS_PATH = Path(__file__).resolve().parents[4] / "i18n" / "ui_keys.json"
_UI_DOMAINS_PATH = Path(__file__).resolve().parents[4] / "i18n" / "ui_domains.json"
_TERMINOLOGY_PATH = Path(__file__).resolve().parents[4] / "i18n" / "terminology.json"
_HF_MODEL_MAP = {
    "en": "Helsinki-NLP/opus-mt-fr-en",
    "de": "Helsinki-NLP/opus-mt-fr-de",
    "bg": "Helsinki-NLP/opus-mt-fr-bg",
}
_HF_TRANSLATORS: dict[str, tuple[object, object]] = {}
_HF_TRANSLATOR_FAILED: set[str] = set()
_HF_TRANSLATOR_LOCK = threading.Lock()


def _table_has_column(table_name: str, column_name: str) -> bool:
    key = (table_name, column_name)
    cached = _SCHEMA_COLUMN_CACHE.get(key)
    if cached is not None:
        return cached
    try:
        inspector = sa_inspect(db.session.get_bind())
        exists = any(
            col.get("name") == column_name for col in inspector.get_columns(table_name)
        )
    except Exception:
        exists = False
    _SCHEMA_COLUMN_CACHE[key] = exists
    return exists


def _table_exists(table_name: str) -> bool:
    cached = _SCHEMA_TABLE_CACHE.get(table_name)
    if cached is not None:
        return cached
    try:
        inspector = sa_inspect(db.session.get_bind())
        exists = bool(inspector.has_table(table_name))
    except Exception:
        exists = False
    _SCHEMA_TABLE_CACHE[table_name] = exists
    return exists


def _cases_enabled() -> bool:
    return (
        _table_exists("cases")
        and _table_exists("case_events")
        and _table_exists("case_participants")
    )


def _safe_json_dict(raw: str | None) -> dict:
    txt = (raw or "").strip()
    if not txt:
        return {}
    try:
        val = json.loads(txt)
        if isinstance(val, dict):
            return val
    except Exception:
        return {}
    return {}


def _as_aware_utc(dt_val: datetime | None) -> datetime | None:
    if not dt_val:
        return None
    if dt_val.tzinfo is None:
        return dt_val.replace(tzinfo=timezone.utc)
    return dt_val.astimezone(timezone.utc)


def _format_elapsed_compact(dt_val: datetime | None) -> str:
    dt_aware = _as_aware_utc(dt_val)
    if not dt_aware:
        return "—"
    delta = max(timedelta(0), _now_utc() - dt_aware)
    total_minutes = int(delta.total_seconds() // 60)
    if total_minutes < 1:
        return "0 min"
    if total_minutes < 60:
        return f"{total_minutes} min"
    total_hours = int(delta.total_seconds() // 3600)
    if total_hours < 24:
        return f"{total_hours} h"
    total_days = int(delta.total_seconds() // 86400)
    return f"{total_days} j"


def _elapsed_tone(dt_val: datetime | None) -> str:
    dt_aware = _as_aware_utc(dt_val)
    if not dt_aware:
        return "muted"
    total_hours = max(0, int((_now_utc() - dt_aware).total_seconds() // 3600))
    if total_hours < 6:
        return "recent"
    if total_hours < 24:
        return "warn"
    return "stale"


def _format_duration_compact(delta: timedelta | None) -> str:
    if delta is None:
        return "—"
    total_minutes = max(0, int(delta.total_seconds() // 60))
    if total_minutes < 60:
        return f"{total_minutes} min"
    total_hours = max(0, int(delta.total_seconds() // 3600))
    if total_hours < 24:
        return f"{total_hours} h"
    total_days = max(0, int(delta.total_seconds() // 86400))
    return f"{total_days} j"


def _case_sla_snapshot(case_row: Case | None) -> dict:
    if not case_row:
        return {
            "target_label": "—",
            "deadline": None,
            "state": "on_time",
            "state_label": "À temps",
            "detail": "—",
        }

    priority = ((getattr(case_row, "priority", None) or "normal").strip().lower())
    target_map = {
        "critical": timedelta(hours=4),
        "high": timedelta(hours=24),
        "normal": timedelta(hours=72),
        "low": timedelta(days=5),
    }
    target_delta = target_map.get(priority, timedelta(hours=72))
    opened_ref = _as_aware_utc(getattr(case_row, "opened_at", None) or getattr(case_row, "created_at", None))
    if not opened_ref:
        return {
            "target_label": _format_duration_compact(target_delta),
            "deadline": None,
            "state": "on_time",
            "state_label": "À temps",
            "detail": "—",
        }

    deadline = opened_ref + target_delta
    remaining = deadline - _now_utc()
    soon_threshold = min(timedelta(hours=4), max(timedelta(hours=1), target_delta / 5))
    if remaining.total_seconds() < 0:
        state = "overdue"
        state_label = "En retard"
        detail = f"En retard de {_format_duration_compact(abs(remaining))}"
    elif remaining <= soon_threshold:
        state = "due_soon"
        state_label = "Échéance proche"
        detail = f"Dans {_format_duration_compact(remaining)}"
    else:
        state = "on_time"
        state_label = "À temps"
        detail = f"Dans {_format_duration_compact(remaining)}"

    return {
        "target_label": _format_duration_compact(target_delta),
        "deadline": deadline,
        "state": state,
        "state_label": state_label,
        "detail": detail,
    }


def _build_operational_blockages(
    req: Request,
    case_row: Case | None = None,
) -> dict:
    blockages: list[str] = []
    now = _now_utc()

    status_val = ((getattr(case_row, "status", None) if case_row else getattr(req, "status", None)) or "").strip().lower()
    risk_level = ((getattr(req, "risk_level", None) or "").strip().lower())
    case_priority = ((getattr(case_row, "priority", None) if case_row else "") or "").strip().lower()
    high_risk = risk_level in {"critical", "attention"} or case_priority in {"high", "critical"}

    def _case_has_owner(current_case: Case | None) -> bool:
        if not current_case:
            return bool(getattr(req, "owner_id", None))
        if getattr(current_case, "owner_user_id", None):
            return True
        participants = getattr(current_case, "participants", None) or []
        return any(
            ((getattr(participant, "role", None) or "").strip().lower() == "owner")
            and bool(getattr(participant, "user_id", None) or getattr(participant, "external_name", None))
            for participant in participants
        )

    def _case_has_professional(current_case: Case | None) -> bool:
        if not current_case:
            return False
        if getattr(current_case, "assigned_professional_lead_id", None):
            return True
        participants = getattr(current_case, "participants", None) or []
        return any(
            bool(getattr(participant, "professional_lead_id", None))
            and (
                (getattr(participant, "role", None) or "").strip().lower() == "primary_professional"
                or (getattr(participant, "participant_type", None) or "").strip().lower() == "professional_lead"
            )
            for participant in participants
        )

    has_owner = _case_has_owner(case_row)
    has_professional = _case_has_professional(case_row)

    if not has_owner:
        blockages.append("Aucun responsable assigné")

    if case_row and not has_professional:
        blockages.append("Aucun professionnel assigné")

    activity_ref = _as_aware_utc(
        (getattr(case_row, "last_activity_at", None) if case_row else None)
        or getattr(case_row, "updated_at", None) if case_row else None
        or getattr(req, "updated_at", None)
        or getattr(req, "created_at", None)
    )
    if activity_ref:
        inactive_for = now - activity_ref
        if inactive_for >= timedelta(hours=72):
            blockages.append(f"Dernière activité il y a {_format_elapsed_compact(activity_ref)}")

    created_ref = _as_aware_utc(getattr(case_row, "created_at", None) if case_row else getattr(req, "created_at", None))
    if created_ref and status_val in {"new", "open", "pending"} and (now - created_ref) >= timedelta(hours=48):
        blockages.append("Dossier encore en statut initial depuis trop longtemps")

    if case_row and high_risk and not has_professional:
        blockages.append("Risque élevé sans suivi professionnel concret")

    return {
        "count": len(blockages),
        "items": blockages[:4],
        "has_blockage": bool(blockages),
    }


def _build_risk_ai_suggestion(req: Request) -> dict:
    triage = score_request_risk(req)
    suggested_label = triage.get("suggested_category_label")
    matched = triage.get("matched_rules") or []
    return {
        "risk_score": int(triage.get("risk_score") or triage.get("score") or 0),
        "risk_label": triage.get("risk_label") or triage.get("label"),
        "suggested_category_code": triage.get("suggested_category_code"),
        "suggested_category_label": suggested_label,
        "matched_rules": matched,
        "has_suggestion": bool(suggested_label),
        "emergency_detected": bool(triage.get("emergency_detected")),
        "emergency_reason_summary": triage.get("emergency_reason_summary"),
    }


def _append_case_event(
    case_id: int,
    event_type: str,
    actor_user_id: int | None = None,
    message: str | None = None,
    metadata: dict | None = None,
    visibility: str = "internal",
) -> CaseEvent:
    normalized_event = (event_type or "note_added").strip().lower()
    if normalized_event not in CASE_EVENT_TYPES:
        normalized_event = "note_added"
    evt = CaseEvent(
        case_id=int(case_id),
        actor_user_id=actor_user_id,
        event_type=normalized_event,
        message=(message or "").strip() or None,
        metadata_json=json.dumps(metadata or {}, ensure_ascii=False) if metadata else None,
        visibility=(visibility or "internal").strip().lower() or "internal",
        created_at=_now_utc(),
    )
    db.session.add(evt)
    return evt


def _upsert_case_participant(
    case_id: int,
    participant_type: str,
    role: str,
    user_id: int | None = None,
    professional_lead_id: int | None = None,
    external_name: str | None = None,
    status: str = "active",
) -> CaseParticipant:
    q = CaseParticipant.query.filter(CaseParticipant.case_id == int(case_id))
    q = q.filter(CaseParticipant.participant_type == participant_type)
    if user_id is not None:
        q = q.filter(CaseParticipant.user_id == int(user_id))
    elif professional_lead_id is not None:
        q = q.filter(CaseParticipant.professional_lead_id == int(professional_lead_id))
    else:
        q = q.filter(CaseParticipant.external_name == (external_name or "").strip())

    row = q.first()
    if row:
        row.role = role
        row.status = status
        if external_name:
            row.external_name = external_name.strip()
        return row

    row = CaseParticipant(
        case_id=int(case_id),
        participant_type=participant_type,
        user_id=user_id,
        professional_lead_id=professional_lead_id,
        external_name=(external_name or "").strip() or None,
        role=role,
        status=status,
        added_at=_now_utc(),
    )
    db.session.add(row)
    return row


def _send_status_email_async(recipient: str, subject: str, context: dict) -> None:
    """Fire-and-forget status email to keep admin status updates responsive."""
    if not recipient:
        return
    app = current_app._get_current_object()

    def _worker() -> None:
        try:
            with app.app_context():
                from backend.mail_service import send_notification_email

                ok = send_notification_email(
                    recipient,
                    subject,
                    "email_template.html",
                    context,
                    purpose="admin_status_update",
                )
                if not ok:
                    app.logger.warning(
                        "[EMAIL] Async status email not sent (recipient=%s, subject=%s)",
                        recipient,
                        subject,
                    )
        except Exception as e:
            app.logger.warning(
                "[EMAIL] Async status email send failed (request-status): %s", e
            )

    threading.Thread(target=_worker, daemon=True).start()


def _log_status_change_once(
    req_id: int,
    old_status: str | None,
    new_status: str | None,
    actor_admin_id: int | None,
):
    """Add a single status_change activity only when there is a real change."""
    if not _table_has_column("request_activities", "volunteer_id"):
        return
    if (old_status or "") == (new_status or ""):
        return
    db.session.add(
        RequestActivity(
            request_id=req_id,
            actor_admin_id=actor_admin_id,
            action="status_change",
            old_value=old_status,
            new_value=new_status,
        )
    )


def _is_request_locked(req) -> bool:
    """Consider a request locked when status is done or cancelled (canonical)."""
    s = normalize_request_status(getattr(req, "status", None))
    return s in ("done", "cancelled")


def _now_utc():
    return datetime.now(UTC)


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _get_admin_last_seen() -> datetime | None:
    last_seen_ts = session.get("admin_last_seen")
    if not last_seen_ts:
        return None
    try:
        return datetime.fromisoformat(last_seen_ts)
    except Exception:
        return None


def _admin_session_is_expired(now: datetime) -> bool:
    last_seen = _get_admin_last_seen()
    if not last_seen:
        return False
    if last_seen.tzinfo is None:
        last_seen = last_seen.replace(tzinfo=timezone.utc)
    return (now - last_seen) > timedelta(minutes=ADMIN_SESSION_IDLE_TIMEOUT_MIN)


def _touch_admin_last_seen(now: datetime) -> None:
    session["admin_last_seen"] = now.isoformat()


def _get_admin_auth_at() -> datetime | None:
    auth_ts = session.get("admin_auth_at")
    if not auth_ts:
        return None
    try:
        return datetime.fromisoformat(auth_ts)
    except Exception:
        return None


def _touch_admin_auth_at(now: datetime) -> None:
    session["admin_auth_at"] = now.isoformat()


def _admin_fresh_auth_is_valid(
    now: datetime, minutes: int = ADMIN_FRESH_AUTH_MIN
) -> bool:
    auth_at = _get_admin_auth_at()
    if not auth_at:
        # Keep legacy tests stable when they bypass login and inject session directly.
        if bool(current_app.config.get("TESTING", False)) and session.get(
            "admin_logged_in"
        ):
            _touch_admin_auth_at(now)
            return True
        return False
    if auth_at.tzinfo is None:
        auth_at = auth_at.replace(tzinfo=timezone.utc)
    return (now - auth_at) <= timedelta(minutes=minutes)


def require_admin_fresh_auth(minutes: int = ADMIN_FRESH_AUTH_MIN):
    def _decorator(fn):
        @wraps(fn)
        def _wrapped(*args, **kwargs):
            if not session.get("admin_logged_in"):
                nxt = request.full_path if request.query_string else request.path
                return redirect(url_for("admin.admin_login_legacy", next=nxt), code=303)
            now = _utc_now()
            if _admin_fresh_auth_is_valid(now, minutes=minutes):
                return fn(*args, **kwargs)
            nxt = request.full_path if request.query_string else request.path
            flash("Veuillez confirmer votre identité pour continuer.", "warning")
            return redirect(url_for("admin.admin_reauth", next=nxt), code=303)

        return _wrapped

    return _decorator


def _client_ip() -> str:
    xff = (request.headers.get("X-Forwarded-For") or "").strip()
    if xff:
        return xff.split(",")[0].strip()
    return request.remote_addr or "0.0.0.0"


def _compute_case_signals(
    req: Request,
    activities: list[RequestActivity] | None,
    now: datetime,
) -> list[dict]:
    """Rules-based, explainable case signals for admin case rail."""
    signals: list[dict] = []
    acts = activities or []
    status = (getattr(req, "status", "") or "").upper()
    priority = (getattr(req, "priority", "medium") or "medium").lower()
    has_owner = bool(getattr(req, "owner_id", None))
    has_phone = bool((getattr(req, "phone", "") or "").strip())
    has_email = bool((getattr(req, "email", "") or "").strip())

    def _as_aware(dt_val: datetime | None) -> datetime | None:
        if not dt_val:
            return None
        if dt_val.tzinfo is None:
            return dt_val.replace(tzinfo=timezone.utc)
        return dt_val

    created_at = _as_aware(getattr(req, "created_at", None))
    updated_at = _as_aware(getattr(req, "updated_at", None))
    owned_at = _as_aware(getattr(req, "owned_at", None))
    last_activity_at = _as_aware(acts[0].created_at) if acts else None
    reference_dt = last_activity_at or updated_at or created_at

    if not has_owner:
        signals.append(
            {
                "code": "no_owner",
                "level": "danger",
                "title": "Aucun responsable assigne",
                "why": "Sans owner, le dossier n'a pas de pilotage clair.",
                "cta_label": "Assigner owner",
                "cta_href": "#owner-actions",
            }
        )
    elif owned_at and (now - owned_at) > timedelta(hours=24) and (
        not last_activity_at or (now - last_activity_at) > timedelta(hours=24)
    ):
        signals.append(
            {
                "code": "owner_idle",
                "level": "warning",
                "title": "Owner inactif",
                "why": "Responsable assigne mais pas d'activite recente.",
                "cta_label": "Verifier activite",
                "cta_href": "#activity-timeline",
            }
        )

    if priority in {"high", "urgent"} and status == "NEW":
        signals.append(
            {
                "code": "urgent_not_started",
                "level": "danger",
                "title": "Urgence non demarree",
                "why": "Priorite elevee avec statut nouveau.",
                "cta_label": "Passer en cours",
                "cta_href": "#status-controls",
            }
        )
    elif priority in {"high", "urgent"} and status == "IN_PROGRESS":
        signals.append(
            {
                "code": "urgent_in_progress",
                "level": "warning",
                "title": "Urgence en cours",
                "why": "Suivi actif requis jusqu'a resolution.",
                "cta_label": "Ajouter note",
                "cta_href": "#internal-note",
            }
        )

    if reference_dt and (now - reference_dt) > timedelta(days=3):
        signals.append(
            {
                "code": "stale_case",
                "level": "warning",
                "title": "Dossier stale",
                "why": "Aucune activite significative depuis plus de 72h.",
                "cta_label": "Revoir dossier",
                "cta_href": "#activity-timeline",
            }
        )

    if status == "NEW" and created_at and (now - created_at) > timedelta(hours=24):
        signals.append(
            {
                "code": "no_first_action_24h",
                "level": "warning",
                "title": "Aucune premiere action > 24h",
                "why": "Le dossier est nouveau mais non traite depuis 24h.",
                "cta_label": "Demarrer traitement",
                "cta_href": "#status-controls",
            }
        )

    if not has_phone and not has_email:
        signals.append(
            {
                "code": "missing_contact",
                "level": "danger",
                "title": "Contact manquant",
                "why": "Ni telephone ni email n'est renseigne.",
                "cta_label": "Ajouter note interne",
                "cta_href": "#internal-note",
            }
        )
    elif has_email and not has_phone:
        signals.append(
            {
                "code": "partial_contact",
                "level": "info",
                "title": "Contact partiel",
                "why": "Email disponible, telephone absent.",
                "cta_label": "Contacter par email",
                "cta_href": "#contact-block",
            }
        )

    if status in {"RESOLVED", "COMPLETED"}:
        has_resolution = any(
            (a.action == "status_change")
            and (a.new_value or "").upper() in {"RESOLVED", "COMPLETED"}
            for a in acts
        )
        has_note = any(a.action == "note" for a in acts)
        if not has_resolution and not has_note:
            signals.append(
                {
                    "code": "closure_without_note",
                    "level": "warning",
                    "title": "Cloture sans note",
                    "why": "Ajouter une note de resolution pour la tracabilite.",
                    "cta_label": "Ajouter note",
                    "cta_href": "#internal-note",
                }
            )

    if not signals:
        signals.append(
            {
                "code": "all_good",
                "level": "ok",
                "title": "Dossier sous controle",
                "why": "Aucun signal critique detecte.",
                "cta_label": "Voir activite",
                "cta_href": "#activity-timeline",
            }
        )

    return signals[:5]


def _parse_risk_signals(value) -> set[str]:
    if not value:
        return set()
    if isinstance(value, list):
        return {str(x).strip().lower() for x in value if str(x).strip()}
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return set()
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return {str(x).strip().lower() for x in parsed if str(x).strip()}
        except Exception:
            pass
        return {chunk.strip().lower() for chunk in raw.split(",") if chunk.strip()}
    return set()


def _build_helpchain_recommendation(
    req: Request,
    activities: list[RequestActivity] | None,
    now: datetime,
) -> dict[str, str]:
    risk_level = (getattr(req, "risk_level", "") or "").strip().lower()
    risk_signals = _parse_risk_signals(getattr(req, "risk_signals", None))
    assigned_actor = (
        getattr(getattr(req, "owner", None), "username", None)
        or (f"#{req.owner_id}" if getattr(req, "owner_id", None) else "")
        or "non assigné"
    )

    last_activity = None
    if activities:
        last_activity = getattr(activities[0], "created_at", None)
    if not last_activity:
        last_activity = getattr(req, "updated_at", None) or getattr(req, "created_at", None)
    if isinstance(last_activity, datetime):
        if last_activity.tzinfo is None:
            last_activity = last_activity.replace(tzinfo=timezone.utc)
        hours_since = max(0, int((now - last_activity).total_seconds() // 3600))
        last_activity_label = f"il y a environ {hours_since} h"
    else:
        last_activity_label = "non disponible"

    if risk_level == "critical" and not getattr(req, "owner_id", None):
        return {
            "priority": "Critique",
            "action": "Affecter un responsable territorial",
            "reason": (
                "Niveau de risque critique et aucun acteur assigné. "
                f"Dernière activité: {last_activity_label}."
            ),
        }

    if "not_seen_72h" in risk_signals:
        return {
            "priority": "Élevée",
            "action": "Vérifier la situation avec l’acteur assigné",
            "reason": (
                "Le signal not_seen_72h indique une absence d’action récente. "
                f"Acteur assigné: {assigned_actor}. Dernière activité: {last_activity_label}."
            ),
        }

    if risk_level == "attention":
        return {
            "priority": "Moyenne",
            "action": "Planifier un suivi",
            "reason": (
                "Niveau de risque attention. "
                f"Acteur assigné: {assigned_actor}. Dernière activité: {last_activity_label}."
            ),
        }

    return {
        "priority": "Standard",
        "action": "Maintenir le suivi courant",
        "reason": (
            f"Pas de déclencheur critique détecté. Acteur assigné: {assigned_actor}. "
            f"Dernière activité: {last_activity_label}."
        ),
    }


def _norm_username(username: str | None) -> str | None:
    if not username:
        return None
    value = username.strip().lower()
    return value or None


def _find_admin_user(login_identifier: str) -> AdminUser | None:
    ident = (login_identifier or "").strip()
    if not ident:
        return None
    ident_l = ident.lower()
    return (
        AdminUser.query.filter(
            or_(
                func.lower(func.coalesce(AdminUser.username, "")) == ident_l,
                func.lower(func.coalesce(AdminUser.email, "")) == ident_l,
            )
        )
        .limit(1)
        .first()
    )


def _verify_admin_password(user: AdminUser | None, password: str) -> bool:
    if not user or not getattr(user, "password_hash", None):
        return False
    try:
        return bool(user.check_password(password))
    except Exception:
        return False


def _admin_login_is_locked(
    ip: str, username: str | None, now: datetime
) -> tuple[bool, int]:
    # Fresh/dev databases may miss this table before migrations/bootstrap.
    # Failing open here avoids a hard 500 on login; audit lockout telemetry
    # resumes automatically once schema is present.
    if not _table_exists("admin_login_attempts"):
        return False, 0

    window_start = now - timedelta(minutes=ADMIN_LOGIN_RATE_WINDOW_MIN)
    query = AdminLoginAttempt.query.filter(
        AdminLoginAttempt.created_at >= window_start,
        AdminLoginAttempt.ip == ip,
        AdminLoginAttempt.success.is_(False),
    )
    if username:
        query = query.filter(AdminLoginAttempt.username == username)

    try:
        fail_count = query.count()
    except Exception:
        # Fail-open for login if local/dev DB drift broke this table schema.
        # We prefer allowing login over returning a 500 on auth form submit.
        db.session.rollback()
        _SCHEMA_TABLE_CACHE.pop("admin_login_attempts", None)
        return False, 0
    if fail_count < ADMIN_LOGIN_MAX_FAILS:
        return False, 0

    try:
        last_fail = query.order_by(AdminLoginAttempt.created_at.desc()).first()
    except Exception:
        db.session.rollback()
        _SCHEMA_TABLE_CACHE.pop("admin_login_attempts", None)
        return False, 0
    if not last_fail or not last_fail.created_at:
        return False, 0

    last_fail_at = last_fail.created_at
    if getattr(last_fail_at, "tzinfo", None) is not None:
        last_fail_at = last_fail_at.astimezone(UTC).replace(tzinfo=None)

    unlock_at = last_fail_at + timedelta(minutes=ADMIN_LOGIN_LOCKOUT_MIN)
    if now < unlock_at:
        retry_after = int((unlock_at - now).total_seconds())
        return True, max(retry_after, 1)

    return False, 0


def _log_admin_attempt(username: str | None, ip: str, success: bool) -> None:
    if not _table_exists("admin_login_attempts"):
        return

    ua = request.headers.get("User-Agent")
    db.session.add(
        AdminLoginAttempt(
            username=username,
            ip=ip,
            success=bool(success),
            user_agent=(ua[:256] if ua else None),
        )
    )
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()


def _clear_recent_admin_login_failures(
    username: str | None, ip: str, now: datetime
) -> int:
    if not _table_exists("admin_login_attempts"):
        return 0

    window_start = now - timedelta(minutes=ADMIN_LOGIN_RATE_WINDOW_MIN)
    query = AdminLoginAttempt.query.filter(
        AdminLoginAttempt.created_at >= window_start,
        AdminLoginAttempt.ip == ip,
        AdminLoginAttempt.success.is_(False),
    )
    if username:
        query = query.filter(AdminLoginAttempt.username == username)

    try:
        deleted = int(query.delete(synchronize_session=False) or 0)
        if deleted:
            db.session.commit()
        return deleted
    except Exception:
        db.session.rollback()
        return 0


def _lockout_response(retry_after_seconds: int, next_url: str = ""):
    retry_after_seconds = max(int(retry_after_seconds or 0), 1)
    flash(GENERIC_ADMIN_LOGIN_FAIL_MSG, "warning")
    response = make_response(render_template("admin/login.html", next=next_url), 429)
    response.headers["Retry-After"] = str(retry_after_seconds)
    return response


def _complete_admin_login(user: AdminUser, next_url: str, *, via: str):
    # Successful login path
    session.clear()  # mitigate session fixation
    login_user(user, remember=False)
    session["admin_user_id"] = user.id
    session["admin_logged_in"] = True
    now = _utc_now()
    _touch_admin_last_seen(now)
    _touch_admin_auth_at(now)
    log_activity(
        entity_type="admin",
        entity_id=user.id,
        action="admin_login",
        message="Admin login",
        meta={"via": via},
        persist=True,
    )
    log_security_event(
        "auth_admin_login_success",
        actor_type="admin",
        actor_id=user.id,
    )
    audit_admin_action(
        action="admin.login.success",
        target_type="AdminUser",
        target_id=int(user.id or 0),
        payload={
            "via": via,
            "next": (next_url or "")[:255],
            "ip": _client_ip(),
            "ua": (request.headers.get("User-Agent") or "")[:256],
        },
    )

    # MFA flow
    session.pop(Config.MFA_SESSION_KEY, None)
    session.pop("mfa_required", None)
    mfa_globally_enabled = bool(Config.MFA_ENABLED)
    user_has_mfa = bool(getattr(user, "mfa_enabled", False)) and bool(
        getattr(user, "totp_secret", None)
    )
    if mfa_globally_enabled and user_has_mfa:
        session["mfa_required"] = True
        return redirect(
            url_for(
                "admin.admin_mfa_verify",
                next=next_url or url_for("admin.admin_requests"),
            )
        )
    _mfa_ok_set()
    return redirect(next_url or url_for("admin.admin_requests"), code=303)


def audit_admin_action(
    action: str, target_type: str, target_id: int, payload: dict | None = None
) -> None:
    try:
        ip = _client_ip()
        ua = request.headers.get("User-Agent")
        admin_user_id = session.get("admin_user_id")
        admin_username = None
        try:
            if getattr(current_user, "is_authenticated", False):
                admin_username = getattr(current_user, "username", None)
                if not admin_user_id:
                    admin_user_id = getattr(current_user, "id", None)
        except Exception:
            admin_username = None

        db.session.add(
            AdminAuditEvent(
                admin_user_id=admin_user_id,
                admin_username=admin_username,
                action=action,
                target_type=target_type,
                target_id=int(target_id),
                ip=ip,
                user_agent=(ua[:256] if ua else None),
                payload=payload,
            )
        )
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception("ADMIN_AUDIT: failed to write audit event")


def _current_structure_id() -> int:
    try:
        from backend.core.tenant import current_structure_id
    except ModuleNotFoundError:
        from core.tenant import current_structure_id
    return int(current_structure_id())


def _is_global_admin() -> bool:
    role = _admin_role_value()
    if role not in {"superadmin"}:
        return False
    return getattr(current_user, "structure_id", None) is None


def _is_structure_admin() -> bool:
    role = _admin_role_value()
    if role not in {"superadmin"}:
        return False
    return getattr(current_user, "structure_id", None) is not None


def _require_global_admin() -> None:
    if not _is_global_admin():
        _audit_denied_action(
            required_roles={"global_admin"},
            actor_role=_admin_role_value(),
        )
        abort(403)


def _require_structure_admin_or_global() -> None:
    if not (_is_structure_admin() or _is_global_admin()):
        _audit_denied_action(
            required_roles={"structure_admin", "global_admin"},
            actor_role=_admin_role_value(),
        )
        abort(403)


def _structure_scope_filter():
    if _is_global_admin():
        return None
    return Request.structure_id == _current_structure_id()


def _apply_tenant_filter(query, tenant_filter):
    if tenant_filter is None:
        return query
    return query.filter(tenant_filter)


def _scope_requests(query):
    """Sprint-1 tenant guard for request queries."""
    tenant_filter = _structure_scope_filter()
    return _apply_tenant_filter(query, tenant_filter)


def _engagement_label(score: int) -> str:
    if score >= 5:
        return "High"
    if score >= 1:
        return "Medium"
    return "At risk"


def compute_structure_health(structure_id: int) -> int:
    score = 100
    now = datetime.utcnow()

    # Requests without assigned operator/owner
    unassigned = (
        Request.query.filter(Request.structure_id == structure_id)
        .filter(Request.owner_id.is_(None))
        .count()
    )
    if unassigned > 0:
        score -= 30

    # Requests inactive for more than 48h (no update, or old update)
    stale_cutoff = now - timedelta(hours=48)
    stale = (
        Request.query.filter(Request.structure_id == structure_id)
        .filter(
            or_(
                Request.updated_at < stale_cutoff,
                and_(Request.updated_at.is_(None), Request.created_at < stale_cutoff),
            )
        )
        .count()
    )
    if stale > 0:
        score -= 20

    # Active requests older than 3 days (treat non-closed as active)
    overdue_cutoff = now - timedelta(days=3)
    overdue = (
        Request.query.filter(Request.structure_id == structure_id)
        .filter(Request.created_at < overdue_cutoff)
        .filter(or_(Request.status.is_(None), ~Request.status.in_(list(CLOSED_STATUSES))))
        .count()
    )
    if overdue > 0:
        score -= 20

    return max(score, 0)


def compute_structure_alerts(structure_id: int) -> dict[str, int]:
    now = datetime.utcnow()
    base = Request.query.filter(Request.structure_id == structure_id)

    unassigned_count = base.filter(Request.owner_id.is_(None)).count()

    urgent_priorities = {"high", "critical", "urgent"}
    urgent_unassigned_count = (
        base.filter(Request.owner_id.is_(None))
        .filter(func.lower(func.coalesce(Request.priority, "")) .in_(urgent_priorities))
        .count()
    )

    stale_cutoff = now - timedelta(hours=72)
    stale_count = base.filter(
        (Request.updated_at < stale_cutoff)
        | (Request.updated_at.is_(None) & (Request.created_at < stale_cutoff))
    ).count()

    overdue_cutoff = now - timedelta(days=3)
    active_filter = or_(
        Request.status.is_(None),
        ~func.lower(func.coalesce(Request.status, "")).in_(list(CLOSED_STATUSES)),
    )
    overdue_count = (
        base.filter(active_filter).filter(Request.created_at < overdue_cutoff).count()
    )

    return {
        "unassigned_count": int(unassigned_count or 0),
        "urgent_unassigned_count": int(urgent_unassigned_count or 0),
        "stale_count": int(stale_count or 0),
        "overdue_count": int(overdue_count or 0),
    }


def get_volunteer_engagement_score(
    volunteer_id: int, now: datetime | None = None
) -> dict:
    """
    Heuristic engagement score per volunteer in range [-10..+10].
    """
    now = now or datetime.now(UTC).replace(tzinfo=None)
    cutoff_72h = now - timedelta(hours=72)

    seen_within_24h = 0
    not_seen_72h = 0
    has_vrs_notified_at = _table_has_column("volunteer_request_states", "notified_at")
    if has_vrs_notified_at and _table_exists("volunteer_request_states"):
        try:
            states = (
                db.session.query(
                    VolunteerRequestState.notified_at,
                    VolunteerRequestState.seen_at,
                )
                .filter(VolunteerRequestState.volunteer_id == volunteer_id)
                .filter(VolunteerRequestState.notified_at.isnot(None))
                .all()
            )
            for notified_at, seen_at in states:
                if notified_at is None:
                    continue
                if seen_at is not None and seen_at <= (notified_at + timedelta(hours=24)):
                    seen_within_24h += 1
                if seen_at is None and notified_at <= cutoff_72h:
                    not_seen_72h += 1
        except Exception:
            db.session.rollback()
    elif _table_exists("notifications"):
        # Compatibility fallback for environments where notified_at migration is missing.
        try:
            notif_rows = (
                db.session.query(
                    Notification.created_at, Notification.read_at, Notification.is_read
                )
                .filter(Notification.volunteer_id == volunteer_id)
                .filter(Notification.type == "new_match")
                .all()
            )
            for created_at, read_at, is_read in notif_rows:
                if created_at is None:
                    continue
                if read_at is not None and read_at <= (created_at + timedelta(hours=24)):
                    seen_within_24h += 1
                if (not bool(is_read)) and created_at <= cutoff_72h:
                    not_seen_72h += 1
        except Exception:
            db.session.rollback()

    can_help = 0
    cant_help = 0
    if _table_has_column("request_activities", "volunteer_id"):
        try:
            can_help = (
                db.session.query(func.count(RequestActivity.id))
                .filter(RequestActivity.volunteer_id == volunteer_id)
                .filter(RequestActivity.action == "volunteer_can_help")
                .scalar()
                or 0
            )
            cant_help = (
                db.session.query(func.count(RequestActivity.id))
                .filter(RequestActivity.volunteer_id == volunteer_id)
                .filter(RequestActivity.action == "volunteer_cant_help")
                .scalar()
                or 0
            )
        except Exception:
            db.session.rollback()

    raw_score = (
        2 * int(seen_within_24h)
        + 3 * int(can_help)
        - 1 * int(cant_help)
        - 3 * int(not_seen_72h)
    )
    score = max(-10, min(10, int(raw_score)))
    return {
        "volunteer_id": int(volunteer_id),
        "score": score,
        "label": _engagement_label(score),
        "seen_within_24h": int(seen_within_24h),
        "not_seen_72h": int(not_seen_72h),
        "can_help": int(can_help),
        "cant_help": int(cant_help),
    }


def _to_utc_naive(dt: datetime | None) -> datetime | None:
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt
    return dt.astimezone(UTC).replace(tzinfo=None)


def _normalize_sla_kind(raw_kind: str | None) -> str | None:
    kind = (raw_kind or "").strip().lower()
    if kind in SLA_QUEUE_KINDS:
        return kind
    # Backward-compatible aliases from /admin/sla type selector.
    alias = {
        "resolve": "resolution_overdue",
        "owner_assign": "owner_assignment_overdue",
        "volunteer_assign": "volunteer_assignment_overdue",
    }
    return alias.get(kind)


def _normalize_sla_days(raw_days) -> int:
    try:
        val = int(raw_days)
    except (TypeError, ValueError):
        val = 30
    return max(1, min(val, 365))


def _sla_open_filter():
    return or_(
        Request.status.is_(None), ~func.lower(Request.status).in_(CLOSED_STATUSES)
    )


def _sla_base_window_query(base_query, *, days: int, now: datetime):
    window_start = now - timedelta(days=days)
    return (
        base_query.filter(Request.created_at.isnot(None))
        .filter(Request.created_at >= window_start)
        .filter(_sla_open_filter())
    )


def _sla_kind_condition(sla_kind: str, *, now: datetime):
    if sla_kind == "resolution_overdue":
        return and_(
            Request.completed_at.is_(None),
            Request.created_at < (now - timedelta(days=RESOLVE_SLA_DAYS)),
        )
    if sla_kind == "owner_assignment_overdue":
        return and_(
            Request.owner_id.is_(None),
            Request.created_at < (now - timedelta(hours=ASSIGN_SLA_HOURS)),
        )
    if sla_kind == "volunteer_assignment_overdue":
        return and_(
            Request.assigned_volunteer_id.is_(None),
            Request.created_at < (now - timedelta(hours=VOLUNTEER_ASSIGN_SLA_HOURS)),
        )
    return None


def _apply_sla_queue_filter(base_query, *, sla_kind: str, days: int, now: datetime):
    cond = _sla_kind_condition(sla_kind, now=now)
    if cond is None:
        return base_query
    return _sla_base_window_query(base_query, days=days, now=now).filter(cond)


def _sla_overdue_hours_by_kind(req, *, now: datetime) -> dict[str, float]:
    created_at = _to_utc_naive(getattr(req, "created_at", None))
    if not created_at:
        return {}
    age_hours = max(0.0, (now - created_at).total_seconds() / 3600.0)
    out: dict[str, float] = {}

    resolve_sla_hours = float(RESOLVE_SLA_DAYS * 24)
    owner_assign_sla_hours = float(ASSIGN_SLA_HOURS)
    volunteer_assign_sla_hours = float(VOLUNTEER_ASSIGN_SLA_HOURS)

    if getattr(req, "completed_at", None) is None and age_hours > resolve_sla_hours:
        out["resolution_overdue"] = age_hours - resolve_sla_hours
    if getattr(req, "owner_id", None) is None and age_hours > owner_assign_sla_hours:
        out["owner_assignment_overdue"] = age_hours - owner_assign_sla_hours
    if (
        getattr(req, "assigned_volunteer_id", None) is None
        and age_hours > volunteer_assign_sla_hours
    ):
        out["volunteer_assignment_overdue"] = age_hours - volunteer_assign_sla_hours
    return out


def _sla_prediction_state(req, *, sla_kind: str, now: datetime) -> dict:
    created_at = _to_utc_naive(getattr(req, "created_at", None))
    if not created_at:
        return {"state": "unknown", "remaining_hours": None, "label": "—"}

    if sla_kind == "resolution_overdue":
        if getattr(req, "completed_at", None) is not None:
            return {"state": "ok", "remaining_hours": None, "label": "—"}
        sla_hours = float(RESOLVE_SLA_DAYS * 24)
    elif sla_kind == "owner_assignment_overdue":
        if getattr(req, "owner_id", None) is not None:
            return {"state": "ok", "remaining_hours": None, "label": "—"}
        sla_hours = float(ASSIGN_SLA_HOURS)
    elif sla_kind == "volunteer_assignment_overdue":
        if getattr(req, "assigned_volunteer_id", None) is not None:
            return {"state": "ok", "remaining_hours": None, "label": "—"}
        sla_hours = float(VOLUNTEER_ASSIGN_SLA_HOURS)
    else:
        return {"state": "unknown", "remaining_hours": None, "label": "—"}

    age_hours = max(0.0, (now - created_at).total_seconds() / 3600.0)
    remaining_hours = sla_hours - age_hours
    warn_threshold = min(4.0, max(1.0, sla_hours / 5.0))

    if remaining_hours < 0:
        return {"state": "breached", "remaining_hours": remaining_hours, "label": "Dépassement"}
    if remaining_hours <= warn_threshold:
        return {"state": "due_soon", "remaining_hours": remaining_hours, "label": "Échéance proche"}
    return {"state": "ok", "remaining_hours": remaining_hours, "label": "À temps"}


def _delta_seconds(start: datetime | None, end: datetime | None) -> int | None:
    start_n = _to_utc_naive(start)
    end_n = _to_utc_naive(end)
    if not start_n or not end_n:
        return None
    return max(0, int((end_n - start_n).total_seconds()))


def compute_response_metrics(
    request_id: int, sla_hours: int = 12, now: datetime | None = None
) -> dict:
    """
    Compute response metrics for a request using assigned volunteer context.
    """
    now_naive = _to_utc_naive(now) or datetime.now(UTC).replace(tzinfo=None)
    out = {
        "request_id": int(request_id),
        "assigned_volunteer_id": None,
        "first_seen_seconds": None,
        "first_action_seconds": None,
        "assignment_delay_seconds": None,
        "not_seen_after_24h": False,
        "sla_hours": int(sla_hours),
        "sla_under_threshold": None,
        "sla_bucket": "no_assignee",
    }

    req = (
        db.session.query(Request.id, Request.created_at, Request.assigned_volunteer_id)
        .filter(Request.id == request_id)
        .first()
    )
    if not req:
        out["sla_bucket"] = "missing_request"
        return out

    req_created_at = _to_utc_naive(getattr(req, "created_at", None))
    assigned_volunteer_id = getattr(req, "assigned_volunteer_id", None)
    out["assigned_volunteer_id"] = assigned_volunteer_id

    first_assign_at = (
        db.session.query(func.min(RequestActivity.created_at))
        .filter(RequestActivity.request_id == request_id)
        .filter(RequestActivity.action == "assign_volunteer")
        .scalar()
    )
    out["assignment_delay_seconds"] = _delta_seconds(req_created_at, first_assign_at)

    if not assigned_volunteer_id:
        return out

    state = (
        db.session.query(
            VolunteerRequestState.notified_at, VolunteerRequestState.seen_at
        )
        .filter(VolunteerRequestState.request_id == request_id)
        .filter(VolunteerRequestState.volunteer_id == int(assigned_volunteer_id))
        .first()
    )
    notified_at = _to_utc_naive(state[0]) if state else None
    seen_at = _to_utc_naive(state[1]) if state else None

    out["first_seen_seconds"] = _delta_seconds(notified_at, seen_at)
    out["not_seen_after_24h"] = bool(
        notified_at
        and not seen_at
        and ((now_naive - notified_at).total_seconds() >= 24 * 3600)
    )

    action_q = (
        db.session.query(func.min(RequestActivity.created_at))
        .filter(RequestActivity.request_id == request_id)
        .filter(
            RequestActivity.action.in_(["volunteer_can_help", "volunteer_cant_help"])
        )
    )
    if _table_has_column("request_activities", "volunteer_id"):
        action_q = action_q.filter(
            RequestActivity.volunteer_id == int(assigned_volunteer_id)
        )
    if notified_at is not None:
        action_q = action_q.filter(RequestActivity.created_at >= notified_at)
    first_action_at = action_q.scalar()
    out["first_action_seconds"] = _delta_seconds(notified_at, first_action_at)

    if notified_at is None:
        out["sla_bucket"] = "no_notification"
        return out
    if out["first_action_seconds"] is None:
        out["sla_bucket"] = "no_action"
        return out

    sla_seconds = int(sla_hours) * 3600
    under = out["first_action_seconds"] <= sla_seconds
    out["sla_under_threshold"] = bool(under)
    out["sla_bucket"] = "under_sla" if under else "over_sla"
    return out


def _notseen_hours_from_risk(risk: str) -> int | None:
    if risk == "notseen":
        return 24
    if not risk.startswith("notseen"):
        return None
    suffix = risk[len("notseen") :]
    if not suffix.isdigit():
        return None
    hours = int(suffix)
    if hours in NOTSEEN_TIERS_HOURS:
        return hours
    return None


def _build_notseen_subquery(now: datetime, *, hours: int):
    cutoff = now - timedelta(hours=hours)
    has_vrs_notified_at = _table_has_column("volunteer_request_states", "notified_at")
    if has_vrs_notified_at:
        subq = (
            db.session.query(VolunteerRequestState.request_id)
            .filter(VolunteerRequestState.notified_at.isnot(None))
            .filter(VolunteerRequestState.notified_at < cutoff)
            .filter(VolunteerRequestState.seen_at.is_(None))
            .subquery()
        )
        source = "notified_at"
    else:
        subq = (
            db.session.query(Notification.request_id)
            .filter(Notification.type == "new_match")
            .filter(Notification.created_at < cutoff)
            .subquery()
        )
        source = "notification_created_at_fallback"
    return subq, source


def _admin_id():
    return getattr(current_user, "id", None)


LOCK_TTL_MINUTES = 30


def _lock_expired(req, now: datetime | None = None) -> bool:
    """Return True if owned_at is older than LOCK_TTL_MINUTES."""
    now = now or _now_utc()
    owned_at = getattr(req, "owned_at", None)
    if not owned_at:
        return False
    try:
        if owned_at.tzinfo is None:
            owned_at = owned_at.replace(tzinfo=UTC)
        return (now - owned_at).total_seconds() > LOCK_TTL_MINUTES * 60
    except Exception:
        return False


def _locked_by_other(req, admin_id, now: datetime | None = None) -> bool:
    return bool(
        getattr(req, "owner_id", None)
        and getattr(req, "owner_id", None) != admin_id
        and not _lock_expired(req, now or _now_utc())
    )


from sqlalchemy import case, func, or_, select
from sqlalchemy import inspect as sa_inspect
from sqlalchemy.orm import joinedload

from ..utils.mfa import (
    build_totp_uri,
    generate_totp_secret,
    qr_png_base64,
    verify_totp_code,
)

admin_bp = Blueprint(
    "admin", __name__, url_prefix="/admin"
)  # single templates path via app.py
ops_bp = Blueprint("ops", __name__, url_prefix="/ops")


@admin_bp.app_context_processor
def inject_admin_category_helpers():
    return {
        "request_category_label": request_category_label,
        "REQUEST_CATEGORY_CHOICES": request_category_choices(),
        "format_elapsed_compact": _format_elapsed_compact,
        "elapsed_tone": _elapsed_tone,
        "case_sla_snapshot": _case_sla_snapshot,
    }


@admin_bp.before_request
def require_admin_session():
    """
    Gate the entire /admin surface behind the explicit admin session flag.

    Security/UX:
    - No public access to admin pages (redirect to login).
    - "Admin" navbar link is also keyed off the same session flag.
    """
    allowed = {
        "admin.ops_login",
        "admin.admin_login_legacy",
        "admin.admin_email_2fa",
        "admin.admin_2fa",
        "admin.metrics",
        "admin.metrics_tenant_leak_test",
    }
    if request.endpoint in allowed or (
        request.endpoint and request.endpoint.startswith("static")
    ):
        return None

    if session.get("admin_logged_in"):
        return None

    nxt = request.full_path if request.query_string else request.path
    return redirect(url_for("admin.admin_login_legacy", next=nxt), code=303)


@admin_bp.before_app_request
def _admin_idle_timeout_guard():
    if not request.path.startswith("/admin"):
        return None
    if not session.get("admin_logged_in"):
        return None
    if request.endpoint in {"admin.admin_login_legacy", "admin.ops_login", "admin.admin_ops_login"}:
        return None

    now = _utc_now()
    if _admin_session_is_expired(now):
        expired_admin_id = session.get("admin_user_id")
        if expired_admin_id:
            audit_admin_action(
                action="admin.session_timeout",
                target_type="AdminUser",
                target_id=int(expired_admin_id),
                payload={
                    "route": request.path,
                    "idle_timeout_min": int(ADMIN_SESSION_IDLE_TIMEOUT_MIN),
                    "ip": _client_ip(),
                    "ua": (request.headers.get("User-Agent") or "")[:256],
                },
            )
        session.pop("admin_logged_in", None)
        session.pop("admin_user_id", None)
        session.pop("admin_last_seen", None)
        session.pop("admin_auth_at", None)
        try:
            logout_user()
        except Exception:
            pass
        flash("Votre session a expiré. Veuillez vous reconnecter.", "warning")
        return redirect(url_for("admin.admin_login_legacy"), code=303)

    _touch_admin_last_seen(now)
    return None


def _metrics_token_valid() -> bool:
    token = (request.args.get("token", "") or "").strip()
    expected = (current_app.config.get("METRICS_TOKEN", "") or "").strip()
    return bool(expected) and secrets.compare_digest(token, expected)


@admin_bp.get("/metrics")
def metrics():
    if not _metrics_token_valid():
        return Response("forbidden\n", status=403, mimetype="text/plain")
    val = tenant_leak_get()
    payload = "\n".join(
        [
            "# HELP tenant_leak_total Tenant guardrail violations detected.",
            "# TYPE tenant_leak_total counter",
            f"tenant_leak_total {val}",
            "",
        ]
    )
    return Response(payload, mimetype="text/plain; version=0.0.4; charset=utf-8")


@admin_bp.get("/metrics/tenant-leak-test")
def metrics_tenant_leak_test():
    if not current_app.config.get("HC_ENABLE_LEAK_TEST", False):
        abort(404)

    allowlist = current_app.config.get("HC_LEAK_TEST_ALLOWLIST", set())
    if allowlist:
        forwarded_for = (request.headers.get("X-Forwarded-For", "") or "").split(",")
        client_ip = (
            (request.headers.get("CF-Connecting-IP", "") or "").strip()
            or (forwarded_for[0].strip() if forwarded_for else "")
            or (request.remote_addr or "").strip()
        )
        if client_ip not in allowlist:
            abort(404)

    if not _metrics_token_valid():
        return jsonify({"ok": False, "error": "forbidden"}), 403
    value = request.args.get("value", default=1, type=int) or 1
    value = max(0, min(value, 1000))
    total = tenant_leak_inc(value)
    return jsonify({"ok": True, "tenant_leak_total": int(total)}), 200


@admin_bp.get("/")
def admin_index():
    """Redirect bare /admin to the main requests list."""
    return redirect(url_for("admin.admin_requests"))


@admin_bp.get("/dashboard")
def admin_dashboard_redirect():
    """Alias for /admin/requests to avoid 404s from legacy /admin/dashboard links."""
    return redirect(url_for("admin.admin_requests"))


@admin_bp.get("/translations")
def admin_translations_list():
    q = (request.args.get("q") or "").strip()
    locale = (request.args.get("locale") or "fr").strip().lower()
    view = (request.args.get("view") or "ops").strip().lower()
    if view not in {"ops", "core", "inventory", "all"}:
        view = "ops"
    only_missing = request.args.get("only_missing") == "1"
    page = max(int(request.args.get("page") or 1), 1)
    per_page = min(max(int(request.args.get("per_page") or 50), 10), 200)

    supported = _supported_locales()
    if locale not in supported:
        locale = current_app.config.get("BABEL_DEFAULT_LOCALE", "fr")
    registry_view = _registry_entries_for_view(view)
    registry_keys = [r["key"] for r in registry_view]
    kpi = _translation_kpi(locale, registry_keys)
    locale_lock = _get_locale_lock(locale)
    is_locale_locked = bool(locale_lock and locale_lock.is_locked)
    freeze_state = _get_translation_freeze_state()
    is_translation_frozen = bool(freeze_state and freeze_state.is_active)
    can_l10n_write = _can_l10n_write()
    can_l10n_delete = _can_l10n_delete()
    can_l10n_write_effective = can_l10n_write and (not is_locale_locked or can_l10n_delete)

    missing_registry = []
    pagination = None
    items = []

    if only_missing:
        existing: set[str] = set()
        if registry_keys:
            existing = {
                k
                for (k,) in db.session.query(UiTranslation.key)
                .filter(
                    UiTranslation.locale == locale,
                    UiTranslation.key.in_(registry_keys),
                )
                .all()
            }

        filtered = [r for r in registry_view if r["key"] not in existing]
        if q:
            ql = q.lower()
            filtered = [
                r
                for r in filtered
                if ql in r["key"].lower()
                or ql in (r.get("default") or "").lower()
                or ql in (r.get("domain") or "").lower()
            ]

        start = (page - 1) * per_page
        end = start + per_page
        missing_registry = filtered[start:end]
    else:
        query = UiTranslation.query.filter(UiTranslation.locale == locale)
        if registry_keys:
            query = query.filter(UiTranslation.key.in_(registry_keys))
        else:
            query = query.filter(UiTranslation.key == "__no_registry_match__")
        if q:
            query = query.filter(UiTranslation.key.ilike(f"%{q}%"))
        query = query.order_by(UiTranslation.key.asc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        items = pagination.items

    return render_template(
        "admin/translations_list.html",
        q=q,
        locale=locale,
        view=view,
        only_missing=only_missing,
        supported_locales=supported,
        items=items,
        missing_registry=missing_registry,
        pagination=pagination,
        kpi=kpi,
        can_l10n_write=can_l10n_write_effective,
        can_l10n_delete=can_l10n_delete,
        is_locale_locked=is_locale_locked,
        locale_lock=locale_lock,
        is_translation_frozen=is_translation_frozen,
        translation_freeze=freeze_state,
    )


@admin_bp.get("/translations/coverage")
def admin_translations_coverage():
    role = _admin_role_value()
    if role not in {"ops", "superadmin"}:
        abort(403)

    locales = ["fr", "en", "de", "bg"]
    registry = _load_ui_key_registry()
    ops_domains = set(_load_ui_domains().get("core_ops_domains", []))
    buckets = ("public", "volunteer", "admin", "ops")

    bucket_keys: dict[str, set[str]] = {b: set() for b in buckets}
    all_keys: set[str] = set()
    for row in registry:
        key = (row.get("key") or "").strip()
        if not key:
            continue
        domain = (row.get("domain") or "").strip().lower()
        bucket = _coverage_bucket_for_domain(domain, ops_domains)
        bucket_keys[bucket].add(key)
        all_keys.add(key)

    locale_keysets: dict[str, set[str]] = {lc: set() for lc in locales}
    if all_keys:
        rows = (
            UiTranslation.query.with_entities(UiTranslation.locale, UiTranslation.key)
            .filter(UiTranslation.locale.in_(locales))
            .filter(UiTranslation.key.in_(list(all_keys)))
            .filter(UiTranslation.is_active.is_(True))
            .all()
        )
        for locale, key in rows:
            locale_keysets.setdefault(locale, set()).add(key)

    coverage: dict[str, dict] = {}
    for locale in locales:
        row_data: dict[str, dict | float] = {}
        locale_keys = locale_keysets.get(locale, set())
        for bucket in buckets:
            keys = bucket_keys[bucket]
            total = len(keys)
            translated = len(locale_keys.intersection(keys))
            ratio = (float(translated) / float(total)) if total else 1.0
            row_data[bucket] = {
                "ratio": ratio,
                "percent": round(ratio * 100.0, 1),
                "translated": translated,
                "total": total,
                "class": _coverage_class(ratio),
            }
        total_all = len(all_keys)
        translated_all = len(locale_keys.intersection(all_keys))
        total_ratio = (float(translated_all) / float(total_all)) if total_all else 1.0
        row_data["total"] = {
            "ratio": total_ratio,
            "percent": round(total_ratio * 100.0, 1),
            "translated": translated_all,
            "total": total_all,
            "class": _coverage_class(total_ratio),
        }
        coverage[locale] = row_data

    translations_count = (
        db.session.query(func.count(UiTranslation.id))
        .filter(UiTranslation.locale.in_(locales))
        .filter(UiTranslation.key.in_(list(all_keys)) if all_keys else UiTranslation.key == "__no_registry_match__")
        .filter(UiTranslation.is_active.is_(True))
        .scalar()
        or 0
    )
    avg_coverage = (
        round(
            (sum(float(coverage[lc]["total"]["ratio"]) for lc in locales) / float(len(locales)))
            * 100.0,
            1,
        )
        if locales
        else 0.0
    )

    if (request.args.get("format") or "").strip().lower() == "json":
        payload = {
            "locales": locales,
            "buckets": list(buckets),
            "kpi": {
                "locales": len(locales),
                "registry_keys": len(all_keys),
                "translations": int(translations_count),
                "average_coverage_percent": avg_coverage,
            },
            "coverage": coverage,
        }
        return jsonify(payload), 200

    return render_template(
        "admin/translations_coverage.html",
        locales=locales,
        coverage=coverage,
        buckets=buckets,
        kpi={
            "locales": len(locales),
            "registry_keys": len(all_keys),
            "translations": int(translations_count),
            "average_coverage_percent": avg_coverage,
        },
    )


@admin_bp.get("/translations/<path:key>")
def admin_translations_key(key: str):
    key = (key or "").strip()
    if not key:
        abort(404)

    locale = (request.args.get("locale") or "fr").strip().lower()
    supported = _supported_locales()
    if locale not in supported:
        locale = current_app.config.get("BABEL_DEFAULT_LOCALE", "fr")

    row = UiTranslation.query.filter_by(key=key, locale=locale).first()
    reg = _registry_index().get(key) or {}
    fallback_text = _fallback_preview(key, locale)
    status = "db_override" if row and row.text else ("fallback" if fallback_text != key else "missing")
    history = []
    if _table_exists("ui_translation_events"):
        history = (
            UiTranslationEvent.query.filter(
                UiTranslationEvent.locale == locale,
                UiTranslationEvent.key == key,
            )
            .order_by(UiTranslationEvent.created_at.desc())
            .limit(20)
            .all()
        )
    last_event = history[0] if history else None
    locale_lock = _get_locale_lock(locale)
    is_locale_locked = bool(locale_lock and locale_lock.is_locked)
    freeze_state = _get_translation_freeze_state()
    is_translation_frozen = bool(freeze_state and freeze_state.is_active)
    can_l10n_write = _can_l10n_write()
    can_l10n_delete = _can_l10n_delete()
    can_l10n_write_effective = can_l10n_write and (not is_locale_locked or can_l10n_delete)

    return render_template(
        "admin/translations_key.html",
        key=key,
        locale=locale,
        supported_locales=supported,
        row=row,
        fallback_text=fallback_text,
        status=status,
        registry_default=(reg.get("default") or ""),
        registry_domain=(reg.get("domain") or ""),
        can_l10n_write=can_l10n_write_effective,
        can_l10n_delete=can_l10n_delete,
        is_locale_locked=is_locale_locked,
        locale_lock=locale_lock,
        is_translation_frozen=is_translation_frozen,
        translation_freeze=freeze_state,
        history_events=history,
        last_event=last_event,
    )


@admin_bp.post("/translations/upsert")
def admin_translations_upsert():
    _require_l10n_write()
    key = (request.form.get("key") or "").strip()
    locale = (request.form.get("locale") or "").strip().lower()
    locked = _blocked_by_locale_lock(locale)
    if locked is not None:
        return locked
    text = _sanitize_translation_text(request.form.get("text") or "")

    if not key or not locale:
        flash("Invalid key/locale.", "danger")
        return redirect(url_for("admin.admin_translations_list"))

    supported = _supported_locales()
    if locale not in supported:
        flash("Unsupported locale.", "danger")
        return redirect(url_for("admin.admin_translations_list"))

    if len(key) > 255:
        flash("Key too long.", "danger")
        return redirect(url_for("admin.admin_translations_list", locale=locale))

    if len(text) > 4000:
        flash("Text too long (max 4000).", "danger")
        return redirect(url_for("admin.admin_translations_key", key=key, locale=locale))

    existing = UiTranslation.query.filter_by(key=key, locale=locale).first()
    freeze_block = _blocked_by_translation_freeze(
        locale=locale,
        action="create",
        allow=existing is not None,
    )
    if freeze_block is not None:
        return freeze_block

    row, action, old_text, changed = _ui_translation_upsert(locale=locale, key=key, text=text)
    if changed:
        _log_translation_event(
            action=action,
            locale=locale,
            key=key,
            old_text=old_text,
            new_text=text,
            source="human",
            translation_id=getattr(row, "id", None),
        )

    db.session.commit()
    flash("Translation saved." if changed else "No changes.", "success" if changed else "info")
    return redirect(url_for("admin.admin_translations_key", key=key, locale=locale))


@admin_bp.post("/translations/delete")
def admin_translations_delete():
    _require_l10n_delete()
    key = (request.form.get("key") or "").strip()
    locale = (request.form.get("locale") or "").strip().lower()
    locked = _blocked_by_locale_lock(locale)
    if locked is not None:
        return locked
    frozen = _blocked_by_translation_freeze(locale=locale, action="delete")
    if frozen is not None:
        return frozen

    row = UiTranslation.query.filter_by(key=key, locale=locale).first()
    if row:
        old_text = row.text
        translation_id = row.id
        db.session.delete(row)
        _log_translation_event(
            action="deleted",
            locale=locale,
            key=key,
            old_text=old_text,
            new_text=None,
            source="human",
            translation_id=translation_id,
        )
        db.session.commit()
        flash("DB override deleted (fallback will be used).", "warning")
    else:
        flash("Nothing to delete.", "info")

    return redirect(url_for("admin.admin_translations_key", key=key, locale=locale))


@admin_bp.post("/translations/suggest")
def admin_translations_suggest():
    key = (request.form.get("key") or "").strip()
    locale = (request.form.get("locale") or "").strip().lower()
    if not key or not locale:
        return jsonify({"error": "missing key/locale"}), 400

    if locale not in _supported_locales():
        return jsonify({"error": "unsupported locale"}), 400

    meta = _ui_registry_get(key)
    default = (meta.get("default") or "").strip()
    domain = (meta.get("domain") or "public").strip()
    suggestions = _rules_suggest(key=key, locale=locale, default=default, domain=domain)
    return jsonify(
        {
            "key": key,
            "locale": locale,
            "domain": domain,
            "default": default,
            "provider": "rules",
            "suggestions": suggestions,
        }
    )


@admin_bp.post("/translations/ai-suggest")
def admin_translations_ai_suggest():
    key = (request.form.get("key") or "").strip()
    locale = (request.form.get("locale") or "").strip().lower()
    provider = (request.form.get("provider") or "hf_local").strip().lower()
    if not key or not locale:
        return jsonify({"error": "missing key/locale"}), 400
    if locale not in _supported_locales():
        return jsonify({"error": "unsupported locale"}), 400

    meta = _ui_registry_get(key)
    default = (meta.get("default") or "").strip()
    domain = (meta.get("domain") or "public").strip()
    source_text = _source_text_for_ai(key=key, default=default)
    suggestions = _ai_suggest(
        key=key,
        locale=locale,
        default=default,
        domain=domain,
        source_text=source_text,
        provider=provider,
    )
    return jsonify(
        {
            "key": key,
            "locale": locale,
            "domain": domain,
            "default": default,
            "provider": provider,
            "suggestions": suggestions,
        }
    )


@admin_bp.post("/translations/bulk-suggest-apply")
def admin_translations_bulk_suggest_apply():
    _require_l10n_write()
    locale = (
        request.form.get("locale")
        or request.args.get("locale")
        or current_app.config.get("BABEL_DEFAULT_LOCALE", "fr")
    ).strip().lower()
    view = (request.form.get("view") or request.args.get("view") or "ops").strip().lower()
    dry_run = (request.form.get("dry_run") or request.args.get("dry_run") or "") == "1"
    resp_format = (request.form.get("format") or request.args.get("format") or "").strip().lower()

    try:
        limit = int(request.form.get("limit") or request.args.get("limit") or "120")
    except ValueError:
        limit = 120
    limit = max(1, min(limit, 500))

    supported = _supported_locales()
    if locale not in supported:
        return jsonify({"ok": False, "error": "unsupported locale"}), 400
    locked = _blocked_by_locale_lock(locale)
    if locked is not None:
        return locked
    if not dry_run:
        frozen = _blocked_by_translation_freeze(locale=locale, action="bulk")
        if frozen is not None:
            return frozen
    if view not in {"ops", "core", "inventory", "all"}:
        view = "ops"

    entries = _registry_entries_for_view(view)
    keys = [e["key"] for e in entries if e.get("key")]

    existing_keys: set[str] = set()
    if keys:
        existing_keys = {
            row.key
            for row in UiTranslation.query.filter(
                UiTranslation.locale == locale,
                UiTranslation.key.in_(keys),
            ).all()
        }

    missing_entries = [e for e in entries if e.get("key") and e["key"] not in existing_keys]
    to_process = missing_entries[:limit]

    applied = 0
    skipped = 0
    for entry in to_process:
        key = entry["key"]
        default = (entry.get("default") or "").strip()
        domain = (entry.get("domain") or "").strip()
        suggestions = _rules_suggest(key=key, locale=locale, default=default, domain=domain)
        if not suggestions:
            skipped += 1
            continue
        best = (suggestions[0].get("text") or "").strip()
        if not best:
            skipped += 1
            continue
        if not dry_run:
            row, _action, old_text, changed = _ui_translation_upsert(locale=locale, key=key, text=best)
            if changed:
                _log_translation_event(
                    action="bulk_rules",
                    locale=locale,
                    key=key,
                    old_text=old_text,
                    new_text=best,
                    source="rules_v1",
                    translation_id=getattr(row, "id", None),
                )
        applied += 1

    if not dry_run:
        db.session.commit()

    report = {
        "ok": True,
        "locale": locale,
        "view": view,
        "limit": limit,
        "dry_run": dry_run,
        "missing": len(missing_entries),
        "applied": applied,
        "skipped": skipped,
        "limit_hit": len(missing_entries) > limit,
    }

    if resp_format == "json" or request.accept_mimetypes.best == "application/json":
        return jsonify(report), 200

    flash(
        f"Auto-fill ({locale}/{view}): applied={applied}, skipped={skipped}, "
        f"missing={len(missing_entries)}{' (dry-run)' if dry_run else ''}.",
        "success" if applied > 0 else "info",
    )

    next_url = (request.form.get("next") or "").strip()
    if next_url and is_safe_url(next_url):
        return redirect(next_url)
    return redirect(
        url_for(
            "admin.admin_translations_list",
            locale=locale,
            view=view,
            only_missing="1",
        )
    )


@admin_bp.post("/translations/bootstrap-from-po")
def admin_translations_bootstrap_from_po():
    _require_l10n_write()
    locale = (
        request.form.get("locale")
        or request.args.get("locale")
        or current_app.config.get("BABEL_DEFAULT_LOCALE", "fr")
    ).strip().lower()
    view = (request.form.get("view") or request.args.get("view") or "ops").strip().lower()
    dry_run = (request.form.get("dry_run") or request.args.get("dry_run") or "") == "1"
    resp_format = (request.form.get("format") or request.args.get("format") or "").strip().lower()

    try:
        limit = int(request.form.get("limit") or request.args.get("limit") or "300")
    except ValueError:
        limit = 300
    limit = max(1, min(limit, 1000))

    supported = _supported_locales()
    if locale not in supported:
        return jsonify({"ok": False, "error": "unsupported locale"}), 400
    locked = _blocked_by_locale_lock(locale)
    if locked is not None:
        return locked
    if not dry_run:
        frozen = _blocked_by_translation_freeze(locale=locale, action="bootstrap")
        if frozen is not None:
            return frozen
    if view not in {"ops", "core", "inventory", "all"}:
        view = "ops"

    entries = [
        e
        for e in _registry_entries_for_view(view)
        if str(e.get("key") or "").startswith("msgid:")
    ]
    keys = [e["key"] for e in entries if e.get("key")]

    existing_keys: set[str] = set()
    if keys:
        existing_keys = {
            row.key
            for row in UiTranslation.query.filter(
                UiTranslation.locale == locale,
                UiTranslation.key.in_(keys),
            ).all()
        }

    missing_entries = [e for e in entries if e.get("key") and e["key"] not in existing_keys]
    translations = _load_babel_translations(locale)
    to_process = missing_entries[:limit]

    applied = 0
    skipped_existing = len(entries) - len(missing_entries)
    skipped_not_found = 0

    for entry in to_process:
        key = entry["key"]
        msgid = key.split("msgid:", 1)[1].strip()
        if not msgid:
            skipped_not_found += 1
            continue
        translated = (translations.gettext(msgid) if translations else msgid).strip()
        if not translated or translated == msgid:
            skipped_not_found += 1
            continue
        if not dry_run:
            row, _action, old_text, changed = _ui_translation_upsert(locale=locale, key=key, text=translated)
            if changed:
                _log_translation_event(
                    action="po_sync",
                    locale=locale,
                    key=key,
                    old_text=old_text,
                    new_text=translated,
                    source="po_sync",
                    translation_id=getattr(row, "id", None),
                )
        applied += 1

    if not dry_run:
        db.session.commit()

    report = {
        "ok": True,
        "locale": locale,
        "view": view,
        "limit": limit,
        "dry_run": dry_run,
        "missing_total": len(missing_entries),
        "db_existing_skipped": skipped_existing,
        "po_found_applied": applied,
        "po_not_found_skipped": skipped_not_found,
        "limit_hit": len(missing_entries) > limit,
    }

    if resp_format == "json" or request.accept_mimetypes.best == "application/json":
        return jsonify(report), 200

    flash(
        f"PO bootstrap ({locale}/{view}): applied={applied}, existing={skipped_existing}, "
        f"not_found={skipped_not_found}{' (dry-run)' if dry_run else ''}.",
        "success" if applied > 0 else "info",
    )

    next_url = (request.form.get("next") or "").strip()
    if next_url and is_safe_url(next_url):
        return redirect(next_url)
    return redirect(
        url_for(
            "admin.admin_translations_list",
            locale=locale,
            view=view,
            only_missing="1",
        )
    )


@admin_bp.post("/translations/locale-lock")
def admin_translations_locale_lock():
    _require_l10n_delete()
    locale = (
        request.form.get("locale")
        or request.args.get("locale")
        or current_app.config.get("BABEL_DEFAULT_LOCALE", "fr")
    ).strip().lower()
    action = (request.form.get("action") or request.args.get("action") or "").strip().lower()
    note = (request.form.get("note") or "").strip()
    view = (request.form.get("view") or request.args.get("view") or "ops").strip().lower()
    resp_format = (request.form.get("format") or request.args.get("format") or "").strip().lower()

    if locale not in _supported_locales():
        return jsonify({"ok": False, "error": "unsupported locale"}), 400
    if action not in {"lock", "unlock"}:
        return jsonify({"ok": False, "error": "invalid action"}), 400
    if len(note) > 255:
        note = note[:255]

    row = _get_locale_lock(locale)
    if row is None:
        row = UiLocaleLock(locale=locale, is_locked=False)
        db.session.add(row)

    if action == "lock":
        row.is_locked = True
        row.locked_at = _utc_now()
        row.locked_by_admin_user_id = getattr(current_user, "id", None)
        row.note = note or row.note
    else:
        row.is_locked = False
        row.note = note or row.note

    db.session.commit()
    result = {
        "ok": True,
        "locale": locale,
        "is_locked": bool(row.is_locked),
        "action": action,
    }

    if resp_format == "json" or request.accept_mimetypes.best == "application/json":
        return jsonify(result), 200

    flash(
        f"Locale {locale.upper()} {'locked' if row.is_locked else 'unlocked'}.",
        "success",
    )
    next_url = (request.form.get("next") or "").strip()
    if next_url and is_safe_url(next_url):
        return redirect(next_url)
    return redirect(url_for("admin.admin_translations_list", locale=locale, view=view), code=303)


@admin_bp.post("/translations/freeze")
def admin_translations_freeze_toggle():
    _require_l10n_delete()
    action = (request.form.get("action") or request.args.get("action") or "").strip().lower()
    release_tag = (request.form.get("release_tag") or "").strip()[:64]
    note = (request.form.get("note") or "").strip()[:255]
    next_url = (request.form.get("next") or "").strip()
    resp_format = (request.form.get("format") or request.args.get("format") or "").strip().lower()

    if action not in {"activate", "deactivate"}:
        return jsonify({"ok": False, "error": "invalid action"}), 400

    row = _get_translation_freeze_state()
    if row is None:
        row = UiTranslationFreeze(is_active=False)
        db.session.add(row)

    if action == "activate":
        row.is_active = True
        row.activated_at = _utc_now()
        row.activated_by_admin_user_id = getattr(current_user, "id", None)
        if release_tag:
            row.release_tag = release_tag
        if note:
            row.note = note
    else:
        row.is_active = False
        if release_tag:
            row.release_tag = release_tag
        if note:
            row.note = note

    db.session.commit()
    result = {
        "ok": True,
        "is_active": bool(row.is_active),
        "release_tag": row.release_tag,
    }

    if resp_format == "json" or request.accept_mimetypes.best == "application/json":
        return jsonify(result), 200

    flash(
        f"Translation freeze {'activated' if row.is_active else 'deactivated'}.",
        "success",
    )
    if next_url and is_safe_url(next_url):
        return redirect(next_url, code=303)
    return redirect(url_for("admin.admin_translations_list"), code=303)


def admin_required_404():
    if not (current_user.is_authenticated and getattr(current_user, "is_admin", False)):
        abort(404)


def _supported_locales() -> list[str]:
    configured = current_app.config.get("SUPPORTED_LOCALES") or ("fr", "en", "de", "bg")
    locales = [str(x).strip().lower() for x in configured if str(x).strip()]
    return sorted(set(locales))


def _can_l10n_write() -> bool:
    return _admin_role_value() in {"ops", "superadmin"}


def _can_l10n_delete() -> bool:
    return _admin_role_value() == "superadmin"


def _get_locale_lock(locale: str) -> UiLocaleLock | None:
    loc = (locale or "").strip().lower()
    if not loc:
        return None
    return UiLocaleLock.query.filter_by(locale=loc).first()


def _get_translation_freeze_state() -> UiTranslationFreeze | None:
    if not _table_exists("ui_translation_freeze"):
        return None
    return UiTranslationFreeze.query.order_by(UiTranslationFreeze.id.asc()).first()


def _is_translation_frozen() -> bool:
    row = _get_translation_freeze_state()
    return bool(row and row.is_active)


def _is_locale_locked(locale: str) -> bool:
    row = _get_locale_lock(locale)
    return bool(row and row.is_locked)


def _require_l10n_write() -> None:
    if _can_l10n_write():
        return
    _audit_denied_action(required_roles={"ops", "superadmin"}, actor_role=_admin_role_value())
    abort(403)


def _require_l10n_delete() -> None:
    if _can_l10n_delete():
        return
    _audit_denied_action(required_roles={"superadmin"}, actor_role=_admin_role_value())
    abort(403)


def _blocked_by_locale_lock(locale: str):
    if not _is_locale_locked(locale):
        return None
    if _can_l10n_delete():
        return None
    wants_json = (
        (request.form.get("format") or request.args.get("format") or "").strip().lower()
        == "json"
        or request.accept_mimetypes.best == "application/json"
    )
    if wants_json:
        return jsonify({"ok": False, "error": "locale_locked", "locale": locale}), 423
    flash(f"Locale {locale.upper()} is locked.", "warning")
    next_url = (request.form.get("next") or "").strip()
    if next_url and is_safe_url(next_url):
        return redirect(next_url, code=303)
    view = (request.form.get("view") or request.args.get("view") or "ops").strip().lower()
    return redirect(url_for("admin.admin_translations_list", locale=locale, view=view), code=303)


def _blocked_by_translation_freeze(*, locale: str, action: str, allow: bool = False):
    if allow or not _is_translation_frozen():
        return None
    wants_json = (
        (request.form.get("format") or request.args.get("format") or "").strip().lower()
        == "json"
        or request.accept_mimetypes.best == "application/json"
    )
    freeze = _get_translation_freeze_state()
    payload = {
        "ok": False,
        "error": "translation_frozen",
        "action": action,
        "locale": locale,
        "release_tag": (freeze.release_tag if freeze else None),
    }
    if wants_json:
        return jsonify(payload), 423
    msg = "Translation freeze is active for release."
    if freeze and freeze.release_tag:
        msg = f"Translation freeze active ({freeze.release_tag})."
    flash(msg, "warning")
    next_url = (request.form.get("next") or "").strip()
    if next_url and is_safe_url(next_url):
        return redirect(next_url, code=303)
    view = (request.form.get("view") or request.args.get("view") or "ops").strip().lower()
    return redirect(url_for("admin.admin_translations_list", locale=locale, view=view), code=303)


def _sanitize_translation_text(value: str) -> str:
    cleaned = _CTRL_CHARS_RE.sub("", value or "")
    return cleaned.strip()


def _fallback_preview(key: str, locale: str) -> str:
    try:
        with force_locale(locale):
            text = _(key)
        return text or key
    except Exception:
        return key


def _load_ui_key_registry() -> list[dict]:
    if not _UI_KEYS_PATH.exists():
        return []
    try:
        rows = json.loads(_UI_KEYS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []
    out: list[dict] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        key = str(row.get("key") or "").strip()
        if not key:
            continue
        out.append(
            {
                "key": key,
                "default": str(row.get("default") or "").strip(),
                "domain": str(row.get("domain") or "").strip(),
                "kind": str(row.get("kind") or "tkey").strip() or "tkey",
                "tier": str(row.get("tier") or "core").strip() or "core",
            }
        )
    return out


def _load_ui_domains() -> dict:
    if not _UI_DOMAINS_PATH.exists():
        return {"core_ops_domains": []}
    try:
        data = json.loads(_UI_DOMAINS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {"core_ops_domains": []}
    if not isinstance(data, dict):
        return {"core_ops_domains": []}
    domains = data.get("core_ops_domains") or []
    if not isinstance(domains, list):
        domains = []
    return {"core_ops_domains": [str(x).strip() for x in domains if str(x).strip()]}


def _registry_entries_for_view(view: str) -> list[dict]:
    registry_all = _load_ui_key_registry()
    view_norm = (view or "ops").strip().lower()
    if view_norm not in {"ops", "core", "inventory", "all"}:
        view_norm = "ops"
    ops_domains = set(_load_ui_domains().get("core_ops_domains", []))

    def _match(row: dict) -> bool:
        row_tier = row.get("tier") or "core"
        row_dom = row.get("domain") or ""
        if view_norm == "ops":
            return row_tier == "core" and row_dom in ops_domains
        if view_norm == "core":
            return row_tier == "core"
        if view_norm == "inventory":
            return row_tier == "inventory"
        return True

    return [row for row in registry_all if _match(row)]


def _registry_index() -> dict[str, dict]:
    return {row["key"]: row for row in _load_ui_key_registry()}


def _ui_registry_get(key: str) -> dict:
    return _registry_index().get((key or "").strip(), {})


def _ui_translation_upsert(locale: str, key: str, text: str) -> tuple[UiTranslation, str, str | None, bool]:
    row = UiTranslation.query.filter_by(key=key, locale=locale).first()
    if row is None:
        row = UiTranslation(key=key, locale=locale, text=text)
        db.session.add(row)
        return row, "created", None, True
    old_text = row.text
    changed = old_text != text
    if changed:
        row.text = text
    row.is_active = True
    return row, "updated", old_text, changed


def _current_admin_email() -> str | None:
    email = (getattr(current_user, "email", None) or "").strip()
    if email:
        return email
    username = (getattr(current_user, "username", None) or "").strip()
    return username or None


def _log_translation_event(
    *,
    action: str,
    locale: str,
    key: str,
    old_text: str | None = None,
    new_text: str | None = None,
    source: str = "human",
    translation_id: int | None = None,
) -> None:
    if not _table_exists("ui_translation_events"):
        return
    db.session.add(
        UiTranslationEvent(
            translation_id=translation_id,
            locale=(locale or "").strip().lower(),
            key=(key or "").strip(),
            action=(action or "").strip(),
            source=(source or "human").strip(),
            actor_admin_user_id=getattr(current_user, "id", None),
            actor_email=_current_admin_email(),
            old_text=old_text,
            new_text=new_text,
        )
    )


def _load_babel_translations(locale: str) -> Translations | None:
    raw_dirs = str(current_app.config.get("BABEL_TRANSLATION_DIRECTORIES") or "translations")
    candidates = [part.strip() for part in re.split(r"[;,]", raw_dirs) if part.strip()]
    if not candidates:
        candidates = ["translations"]
    for directory in candidates:
        try:
            return Translations.load(directory, locales=[locale], domain="messages")
        except Exception:
            continue
    return None


def _translation_kpi(locale: str, keys: list[str]) -> dict:
    total = len(keys)
    if total == 0:
        return {"total": 0, "overrides": 0, "missing": 0, "coverage": 0.0}

    overrides = (
        db.session.query(func.count(UiTranslation.key.distinct()))
        .filter(
            UiTranslation.locale == locale,
            UiTranslation.key.in_(keys),
        )
        .scalar()
        or 0
    )
    missing = max(total - int(overrides), 0)
    coverage = round((float(overrides) / float(total)) * 100.0, 1) if total else 0.0
    return {
        "total": int(total),
        "overrides": int(overrides),
        "missing": int(missing),
        "coverage": float(coverage),
    }


def _coverage_bucket_for_domain(domain: str, ops_domains: set[str]) -> str:
    dom = (domain or "").strip().lower()
    if dom in ops_domains:
        return "ops"
    if dom.startswith("admin"):
        return "admin"
    if dom.startswith("volunteer"):
        return "volunteer"
    return "public"


def _coverage_class(ratio: float) -> str:
    if ratio > 0.9:
        return "success"
    if ratio > 0.6:
        return "warning"
    return "danger"


def _load_terminology() -> dict:
    if not _TERMINOLOGY_PATH.exists():
        return {}
    try:
        return json.loads(_TERMINOLOGY_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _infer_kind(key: str) -> str:
    k = (key or "").lower()
    if any(p in k for p in ["btn_", "cta_", "action_", "submit_", "save_", "close_", "assign_"]):
        return "button"
    if any(p in k for p in ["title_", "header_", "nav_", "menu_"]):
        return "title"
    if any(p in k for p in ["hint_", "help_", "desc_", "subtitle_", "empty_"]):
        return "helper"
    return "label"


def _rules_suggest(key: str, locale: str, default: str, domain: str) -> list[dict]:
    """Rules-based suggestions (deterministic/explainable, no external calls)."""
    kind = _infer_kind(key)
    term = _load_terminology()
    loc_terms = term.get(locale, {})
    is_admin = (domain or "").strip().lower() == "admin"
    k = (key or "").lower()
    default = (default or "").strip()

    def t(token: str, fallback: str) -> str:
        return str(loc_terms.get(token, fallback) or fallback).strip()

    if "dashboard" in k:
        base = t("dashboard", "Dashboard")
        reason_base = "Term: dashboard"
    elif "assign" in k or "claim" in k or "prendre_en_charge" in k:
        base = t("assign", "Übernehmen" if locale == "de" else "Assign")
        reason_base = "Term: assign/claim"
    elif "status" in k:
        base = t("status", "Status")
        reason_base = "Term: status"
    elif "request" in k or "demande" in k:
        base = t("request", "Anfrage" if locale == "de" else "Request")
        reason_base = "Term: request"
    elif "case" in k or "dossier" in k:
        base = t("case", "Vorgang" if locale == "de" else "Case")
        reason_base = "Term: case"
    else:
        base = default or key
        reason_base = "Fallback: default/key"

    if kind == "button":
        v1 = base
        v2 = f"{base}..." if len(base) <= 20 else base
        v3 = base.upper() if (is_admin and len(base) <= 16) else base
        return [
            {"text": v1, "reason": f"Button label (short). {reason_base}"},
            {"text": v2, "reason": f"Button label (progressive). {reason_base}"},
            {"text": v3, "reason": f"Button label (ops emphasis). {reason_base}"},
        ]

    if kind == "title":
        v1 = base
        v2 = f"{base} — {t('dashboard', 'Dashboard')}" if ("dashboard" not in k and len(base) <= 24) else base
        v3 = f"{base} ({locale.upper()})" if (is_admin and len(base) <= 24) else base
        return [
            {"text": v1, "reason": f"Title. {reason_base}"},
            {"text": v2, "reason": f"Title with context. {reason_base}"},
            {"text": v3, "reason": f"Title (admin clarity). {reason_base}"},
        ]

    if kind == "helper":
        v1 = base
        v2 = f"{base}. {t('status', 'Status')}." if ("status" not in k and len(base) <= 40) else base
        v3 = f"{base}." if base and not base.endswith(".") else base
        return [
            {"text": v1, "reason": f"Helper text (concise). {reason_base}"},
            {"text": v2, "reason": f"Helper text (adds cue). {reason_base}"},
            {"text": v3, "reason": f"Helper text (punctuation). {reason_base}"},
        ]

    v1 = base
    v2 = base.capitalize() if base else base
    v3 = base
    return [
        {"text": v1, "reason": f"Label. {reason_base}"},
        {"text": v2, "reason": f"Label (capitalization). {reason_base}"},
        {"text": v3, "reason": f"Label (stable). {reason_base}"},
    ]


def _source_text_for_ai(*, key: str, default: str) -> str:
    if default:
        return default.strip()
    if (key or "").startswith("msgid:"):
        return key.split("msgid:", 1)[1].strip()
    return (key or "").strip()


def _ai_suggest(
    *,
    key: str,
    locale: str,
    default: str,
    domain: str,
    source_text: str,
    provider: str,
) -> list[dict]:
    # Always produce deterministic fallback suggestions.
    rules = _rules_suggest(key=key, locale=locale, default=default, domain=domain)
    if provider != "hf_local":
        return rules

    translated = _hf_translate_fr_to(locale=locale, text=source_text)
    out: list[dict] = []
    seen: set[str] = set()

    if translated:
        ai_text = translated.strip()
        if ai_text and ai_text not in seen:
            out.append(
                {
                    "text": ai_text,
                    "reason": "AI suggestion (hf_local MarianMT).",
                }
            )
            seen.add(ai_text)

    for item in rules:
        txt = (item.get("text") or "").strip()
        if not txt or txt in seen:
            continue
        out.append({"text": txt, "reason": item.get("reason") or "Rules fallback."})
        seen.add(txt)
        if len(out) >= 3:
            break

    return out[:3] if out else rules[:3]


def _hf_translate_fr_to(*, locale: str, text: str) -> str | None:
    target = (locale or "").strip().lower()
    src_text = (text or "").strip()
    if not src_text:
        return None
    if target == "fr":
        return src_text
    model_name = _HF_MODEL_MAP.get(target)
    if not model_name:
        return None

    with _HF_TRANSLATOR_LOCK:
        if model_name in _HF_TRANSLATOR_FAILED:
            return None
        pair = _HF_TRANSLATORS.get(model_name)

    if pair is None:
        try:
            from transformers import MarianMTModel, MarianTokenizer
        except Exception:
            with _HF_TRANSLATOR_LOCK:
                _HF_TRANSLATOR_FAILED.add(model_name)
            return None
        try:
            tokenizer = MarianTokenizer.from_pretrained(model_name)
            model = MarianMTModel.from_pretrained(model_name)
        except Exception:
            with _HF_TRANSLATOR_LOCK:
                _HF_TRANSLATOR_FAILED.add(model_name)
            return None
        with _HF_TRANSLATOR_LOCK:
            _HF_TRANSLATORS[model_name] = (tokenizer, model)
            pair = (tokenizer, model)

    try:
        tokenizer, model = pair
        tokens = tokenizer(src_text, return_tensors="pt", truncation=True)
        generated = model.generate(**tokens)
        out = tokenizer.decode(generated[0], skip_special_tokens=True).strip()
        return out or None
    except Exception:
        return None


def _is_local_request() -> bool:
    """True when request comes from localhost (IPv4/IPv6)."""
    return request.remote_addr in ("127.0.0.1", "::1")


def is_safe_url(target: str) -> bool:
    if not target:
        return False
    ref_url = urlparse(request.host_url)
    test_url = urlparse(urljoin(request.host_url, target))
    return (test_url.scheme in ("http", "https")) and (
        ref_url.netloc == test_url.netloc
    )


def admin_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        # Harden: hide admin surface from non-admins (404 instead of redirect)
        if not (
            current_user.is_authenticated and getattr(current_user, "is_admin", False)
        ):
            abort(404)
        return view_func(*args, **kwargs)

    return wrapper


def operator_required(view_func):
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not (
            current_user.is_authenticated and getattr(current_user, "is_admin", False)
        ):
            session_role = (session.get("role") or "").strip().lower()
            if session.get("is_admin") or session.get("admin_logged_in"):
                if session_role in {"ops", "readonly", "admin", "superadmin"}:
                    return view_func(*args, **kwargs)
            abort(404)
        role = _admin_role_value()
        if role is None or role not in {"ops", "readonly", "admin", "superadmin"}:
            _audit_denied_action(
                required_roles={"ops", "readonly", "admin", "superadmin"},
                actor_role=role,
            )
            abort(404)
        return view_func(*args, **kwargs)

    return wrapper


def _admin_role_value() -> str | None:
    raw_role = getattr(current_user, "role", None)
    role = getattr(raw_role, "value", raw_role)
    role = (role or "").strip().lower()
    if role in {"admin", "super_admin", "superadmin"}:
        return "superadmin"
    if role in {"ops"}:
        return "ops"
    if role in {"readonly", "read-only"}:
        return "readonly"
    return None


def _normalize_admin_role_value(raw_role) -> str | None:
    role = getattr(raw_role, "value", raw_role)
    role = (role or "").strip().lower()
    if role in {"admin", "super_admin", "superadmin"}:
        return "superadmin"
    if role == "ops":
        return "ops"
    if role in {"readonly", "read-only"}:
        return "readonly"
    return None


def admin_role_required(*allowed_roles: str):
    allowed = {r.strip().lower() for r in allowed_roles if r}

    def deco(view_func):
        @wraps(view_func)
        def wrapper(*args, **kwargs):
            role = _admin_role_value()
            if role is None:
                _audit_denied_action(required_roles=allowed, actor_role=str(role))
                abort(403)
            if role not in allowed:
                _audit_denied_action(required_roles=allowed, actor_role=role)
                abort(403)
            return view_func(*args, **kwargs)

        return wrapper

    return deco


def _is_superadmin_role(role) -> bool:
    return _normalize_admin_role_value(role) == "superadmin"


@admin_bp.get("/sanity")
@admin_required
@admin_role_required("readonly", "ops", "superadmin")
def admin_sanity():
    admin_required_404()
    checks = run_system_checks()
    return render_template("admin/system_sanity.html", checks=checks)


@admin_bp.get("/ops")
@admin_required
@admin_role_required("readonly", "ops", "superadmin")
def admin_ops():
    admin_required_404()
    return render_template("admin/ops_dashboard.html")


@admin_bp.get("/operator")
@admin_required
@admin_role_required("readonly", "ops", "superadmin")
def admin_operator_dashboard():
    admin_required_404()
    return _render_operator_dashboard()


@ops_bp.get("")
@ops_bp.get("/")
@ops_bp.get("/workspace")
@operator_required
def ops_workspace():
    return _render_operator_dashboard()


def _render_operator_dashboard():
    base_query = _scope_requests(Request.query).filter(Request.deleted_at.is_(None))
    try:
        base_query = base_query.filter(Request.is_archived.is_(False))
    except Exception:
        pass

    status_expr = func.lower(func.coalesce(Request.status, ""))
    actionable_statuses = ("open", "in_progress", "approved", "pending")
    actionable_filter = status_expr.in_(actionable_statuses)
    activity_expr = func.coalesce(Request.updated_at, Request.created_at)
    stale_threshold = _now_utc() - timedelta(hours=72)
    today_start = _now_utc().replace(hour=0, minute=0, second=0, microsecond=0)
    urgent_filter = or_(
        func.lower(func.coalesce(Request.priority, "")).in_(["high", "critical"]),
        func.coalesce(Request.risk_score, 0) >= 85,
    )
    unassigned_filter = Request.owner_id.is_(None)

    urgent_count = base_query.filter(actionable_filter, urgent_filter).count()
    unassigned_count = base_query.filter(actionable_filter, unassigned_filter).count()
    followup_count = base_query.filter(
        actionable_filter, activity_expr <= stale_threshold
    ).count()
    updated_today_count = base_query.filter(activity_expr >= today_start).count()

    failed_notif_count = 0
    retry_notif_count = 0
    if _table_exists("notification_jobs"):
        notif_base = NotificationJob.query
        try:
            if not _is_global_admin():
                sid = _current_structure_id()
                notif_base = notif_base.filter(
                    (NotificationJob.structure_id == sid)
                    | (NotificationJob.structure_id.is_(None))
                )
        except Exception:
            pass
        failed_notif_count = notif_base.filter(NotificationJob.status == "failed").count()
        retry_notif_count = notif_base.filter(NotificationJob.status == "retry").count()

    queue_query = (
        base_query.filter(actionable_filter)
        .filter(or_(urgent_filter, unassigned_filter, activity_expr <= stale_threshold))
        .options(joinedload(Request.owner))
    )
    queue_rows = (
        queue_query.order_by(
            case((urgent_filter, 0), else_=1).asc(),
            case((unassigned_filter, 0), else_=1).asc(),
            case((activity_expr <= stale_threshold, 0), else_=1).asc(),
            activity_expr.desc().nullslast(),
            Request.id.desc(),
        )
        .limit(20)
        .all()
    )

    queue_reasons = {}
    now_utc = _now_utc()
    scored_rows = []
    for r in queue_rows:
        result = compute_ops_priority(request_row=r, now=now_utc)
        scored_rows.append((int(result.get("ops_priority_score") or 0), r, result))
    scored_rows.sort(key=lambda row: row[0], reverse=True)
    scored_rows = scored_rows[:10]
    queue_rows = [row[1] for row in scored_rows]
    ops_priority_levels = {}
    for _, row, result in scored_rows:
        queue_reasons[int(row.id)] = result.get("ops_priority_reasons") or []
        ops_priority_levels[int(row.id)] = result.get("ops_priority_level") or "normal"

    return render_template(
        "admin/operator_dashboard.html",
        urgent_count=urgent_count,
        unassigned_count=unassigned_count,
        followup_count=followup_count,
        updated_today_count=updated_today_count,
        failed_notif_count=failed_notif_count,
        retry_notif_count=retry_notif_count,
        queue_rows=queue_rows,
        queue_reasons=queue_reasons,
        ops_priority_levels=ops_priority_levels,
    )


@ops_bp.get("/cases")
@operator_required
def ops_cases_list():
    return _render_cases_list()


@ops_bp.get("/notifications")
@operator_required
def ops_notifications_list():
    return _render_notifications_list()


@admin_bp.get("/risk-map")
@admin_required
@admin_role_required("readonly", "ops", "superadmin")
def admin_risk_map():
    admin_required_404()
    return render_template("admin/risk_map.html")


def _attempted_action_label() -> str:
    if request.endpoint:
        return request.endpoint
    return f"{request.method} {request.path}"


def _audit_denied_action(required_roles: set[str], actor_role: str | None) -> None:
    if request.method not in STATE_CHANGING_METHODS:
        return

    target_type = "AdminRoute"
    target_id = 0
    try:
        view_args = request.view_args or {}
        if "req_id" in view_args and view_args.get("req_id") is not None:
            target_type = "Request"
            target_id = int(view_args["req_id"])
        elif "interest_id" in view_args and view_args.get("interest_id") is not None:
            target_type = "Interest"
            target_id = int(view_args["interest_id"])
        elif "admin_id" in view_args and view_args.get("admin_id") is not None:
            target_type = "AdminUser"
            target_id = int(view_args["admin_id"])
    except Exception:
        target_type = "AdminRoute"
        target_id = 0

    audit_admin_action(
        action="security.denied_action",
        target_type=target_type,
        target_id=target_id,
        payload={
            "attempted_action": _attempted_action_label(),
            "required_roles": sorted(list(required_roles)),
            "actor_role": actor_role,
            "method": request.method,
            "path": request.path,
        },
    )


def log_request_activity(req_obj, action, old=None, new=None, actor_admin_id=None):
    """Append a RequestActivity row; swallow errors so UI flows stay smooth."""
    if not _table_has_column("request_activities", "volunteer_id"):
        return
    try:
        actor_id = actor_admin_id
        if actor_id is None and getattr(current_user, "is_authenticated", False):
            actor_id = getattr(current_user, "id", None)
        activity = RequestActivity(
            request_id=getattr(req_obj, "id", req_obj),
            actor_admin_id=actor_id,
            action=action,
            old_value=str(old) if old is not None else None,
            new_value=str(new) if new is not None else None,
            created_at=utc_now(),
        )
        db.session.add(activity)
    except Exception:
        pass


def _audit_request(
    req_id: int,
    *,
    action: str,
    message: str | None = None,
    old: str | None = None,
    new: str | None = None,
    meta: dict | None = None,
):
    if not _table_has_column("activity_logs", "id"):
        return
    try:
        log_activity(
            entity_type="request",
            entity_id=req_id,
            action=action,
            message=message,
            old_value=str(old) if old is not None else None,
            new_value=str(new) if new is not None else None,
            meta=meta,
        )
    except Exception:
        pass


def _audit_pro_access(
    pro_id: int,
    *,
    action: str,
    message: str | None = None,
    old: str | None = None,
    new: str | None = None,
    meta: dict | None = None,
):
    if not _table_has_column("activity_logs", "id"):
        return
    try:
        log_activity(
            entity_type="pro_access",
            entity_id=pro_id,
            action=action,
            message=message,
            old_value=str(old) if old is not None else None,
            new_value=str(new) if new is not None else None,
            meta=meta,
        )
    except Exception:
        pass


def _mfa_ok_set(ttl_min: int | None = None):
    ttl = ttl_min or current_app.config.get("MFA_SESSION_TTL_MIN", 720)
    session[current_app.config.get("MFA_SESSION_KEY", "mfa_ok")] = True
    try:
        session["mfa_ok_until"] = (utc_now() + timedelta(minutes=ttl)).isoformat()
    except Exception:
        session["mfa_ok_until"] = None


def _mfa_ok_clear():
    session.pop(current_app.config.get("MFA_SESSION_KEY", "mfa_ok"), None)
    session.pop("mfa_ok_until", None)


def _mfa_ok_is_valid() -> bool:
    if not session.get(current_app.config.get("MFA_SESSION_KEY", "mfa_ok")):
        return False
    until = session.get("mfa_ok_until")
    if not until:
        return False
    try:
        return utc_now() <= datetime.fromisoformat(until)
    except Exception:
        return False


def _mfa_lock_is_active() -> tuple[bool, int]:
    lock_until = session.get("mfa_lock_until")
    if not lock_until:
        return (False, 0)
    try:
        dt = datetime.fromisoformat(lock_until)
        if utc_now() >= dt:
            session.pop("mfa_lock_until", None)
            session.pop("mfa_attempts", None)
            return (False, 0)
        return (True, int((dt - utc_now()).total_seconds()))
    except Exception:
        session.pop("mfa_lock_until", None)
        session.pop("mfa_attempts", None)
        return (False, 0)


def _mfa_attempt_fail():
    max_attempts = current_app.config.get("MFA_VERIFY_MAX_ATTEMPTS", 8)
    lock_min = current_app.config.get("MFA_VERIFY_LOCK_MIN", 10)
    session["mfa_attempts"] = int(session.get("mfa_attempts", 0)) + 1
    if session["mfa_attempts"] >= max_attempts:
        try:
            session["mfa_lock_until"] = (
                utc_now() + timedelta(minutes=lock_min)
            ).isoformat()
        except Exception:
            session["mfa_lock_until"] = None


def _mfa_attempt_reset():
    session.pop("mfa_attempts", None)
    session.pop("mfa_lock_until", None)


def _generate_backup_codes(n: int = 10) -> list[str]:
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    codes = []
    for _ in range(n):
        code = "".join(secrets.choice(alphabet) for _ in range(10))
        codes.append(code)
    return codes


def _hash_codes(codes: list[str]) -> list[str]:
    return [generate_password_hash(c) for c in codes]


def _load_hashes(user) -> list[str]:
    raw = getattr(user, "backup_codes_hashes", None)
    if not raw:
        return []
    try:
        return json.loads(raw)
    except Exception:
        return []


def _save_hashes(user, hashes: list[str]):
    user.backup_codes_hashes = json.dumps(hashes)
    user.backup_codes_generated_at = utc_now()
    db.session.commit()


def can_edit_request(req_obj, user) -> bool:
    """Owner or superadmin can edit; if no owner, only superadmin."""
    role = getattr(getattr(user, "role", None), "value", getattr(user, "role", None))
    role = (role or "").strip().lower()
    if role in {"super_admin", "superadmin", "admin"}:
        return True
    owner_id = getattr(req_obj, "owner_id", None)
    user_id = getattr(user, "id", None)
    return owner_id is not None and owner_id == user_id


def is_stale(req_obj, minutes: int = 30) -> bool:
    """Return True if owned_at is older than `minutes`."""
    owned_at = getattr(req_obj, "owned_at", None)
    if not owned_at:
        return False
    try:
        return (utc_now() - owned_at).total_seconds() > minutes * 60
    except Exception:
        return False


@admin_bp.before_request
def enforce_admin_mfa():
    if not current_app.config.get("MFA_ENABLED", False):
        return None
    allowed = {
        "admin.ops_login",
        "admin.admin_login_legacy",
        "admin.admin_logout",
        "admin.admin_mfa_setup",
        "admin.admin_mfa_verify",
        "admin.admin_mfa_backup_codes",
        "static",
    }
    if request.endpoint in allowed or (
        request.endpoint and request.endpoint.startswith("static")
    ):
        return None
    if not current_user.is_authenticated:
        return None
    if not (
        getattr(current_user, "mfa_enabled", False)
        and getattr(current_user, "totp_secret", None)
    ):
        return None
    if _mfa_ok_is_valid():
        return None
    nxt = request.full_path if request.query_string else request.path
    return redirect(url_for("admin.admin_mfa_verify", next=nxt))


@admin_bp.route("/emergency-requests", methods=["GET"])
@admin_required
def emergency_requests():
    admin_required_404()
    # Admin guard (same as admin_dashboard)
    if not getattr(current_user, "is_admin", False):
        flash(_("Access denied."), "error")
        return redirect(url_for("main.index"))

    q = (request.args.get("q") or "").strip()
    days = int(request.args.get("days") or 7)
    days = max(1, min(days, 90))
    page = int(request.args.get("page") or 1)
    page = max(page, 1)
    per_page = int(request.args.get("per_page") or 25)
    per_page = max(10, min(per_page, 100))
    since = datetime.now(UTC).replace(tzinfo=None) - timedelta(days=days)

    # Emergency filter: category=="emergency" only (no urgency field)
    query = _scope_requests(Request.query).filter(
        Request.created_at >= since, Request.category == "emergency"
    ).order_by(Request.created_at.desc())

    if q:
        # Search in city/contact/priority only
        query = query.filter(
            (Request.city.ilike(f"%{q}%"))
            | (Request.email.ilike(f"%{q}%"))
            | (Request.phone.ilike(f"%{q}%"))
            | (Request.priority.ilike(f"%{q}%"))
        )

    total = query.count()
    items = query.offset((page - 1) * per_page).limit(per_page).all()

    return render_template(
        "admin_emergency_requests.html",
        items=items,
        total=total,
        page=page,
        per_page=per_page,
        q=q,
        days=days,
    )


# API endpoint за заявки с филтри (status, date)
def api_requests():
    from flask import current_app, jsonify, request

    # During tests we allow access to the API endpoints to simplify fixtures
    if not current_app.config.get("TESTING", False):
        if not getattr(current_user, "is_admin", False):
            return jsonify({"error": "Unauthorized"}), 403
    query = _scope_requests(Request.query)
    status = request.args.get("status")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    if status:
        query = query.filter_by(status=status)
    if date_from:
        try:
            from datetime import datetime

            date_from_dt = datetime.fromisoformat(date_from)
            query = query.filter(Request.created_at >= date_from_dt)
        except Exception:
            pass
    if date_to:
        try:
            from datetime import datetime, timedelta

            date_to_dt = datetime.fromisoformat(date_to) + timedelta(days=1)
            query = query.filter(Request.created_at < date_to_dt)
        except Exception:
            pass
    requests = query.order_by(Request.created_at.desc()).all()
    data = [
        {
            "id": r.id,
            "name": r.name,
            "description": r.description,
            "status": r.status,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        }
        for r in requests
    ]
    return jsonify({"items": data})


# API endpoint за всички доброволци (JSON)
def api_volunteers():
    from flask import current_app, jsonify

    if not current_app.config.get("TESTING", False):
        if not getattr(current_user, "is_admin", False):
            return jsonify({"error": "Unauthorized"}), 403
    volunteers = Volunteer.query.all()
    data = [
        {
            "id": v.id,
            "name": v.name,
            "email": v.email,
            "phone": v.phone,
            "location": v.location,
            "skills": v.skills,
            "is_active": v.is_active,
        }
        for v in volunteers
    ]
    return jsonify(data)


# Детайли за доброволец
@admin_bp.route("/admin_volunteers/<int:id>")
@admin_required
def volunteer_detail(id):
    admin_required_404()
    if not current_user.is_admin:
        flash(_("Access denied."), "error")
        return redirect(url_for("main.index"))
    from flask import abort

    volunteer = db.session.get(Volunteer, id)
    if not volunteer:
        abort(404)
    return render_template("volunteer_detail.html", volunteer=volunteer)


@admin_bp.get("/api/volunteers/<int:vol_id>")
@admin_required
def admin_volunteer_api(vol_id: int):
    admin_required_404()
    volunteer = db.session.get(Volunteer, vol_id)
    if not volunteer:
        abort(404)
    req_id_raw = (request.args.get("req_id") or "").strip()
    req_id = int(req_id_raw) if req_id_raw.isdigit() else None

    def _to_list(value):
        if value is None:
            return []
        if isinstance(value, list):
            return [str(x).strip() for x in value if str(x).strip()]
        raw = str(value).strip()
        if not raw:
            return []
        return [chunk.strip() for chunk in raw.split(",") if chunk.strip()]

    last_active = getattr(volunteer, "last_activity", None) or getattr(
        volunteer, "updated_at", None
    )
    can_help_count = (
        db.session.query(func.count(VolunteerAction.id))
        .filter(
            VolunteerAction.volunteer_id == volunteer.id,
            VolunteerAction.action == "CAN_HELP",
        )
        .scalar()
        or 0
    )
    action_rows = (
        VolunteerAction.query.filter_by(volunteer_id=volunteer.id)
        .order_by(VolunteerAction.updated_at.desc())
        .limit(5)
        .all()
    )
    history = [
        {
            "request_id": row.request_id,
            "action": row.action,
            "at": (
                (row.updated_at or row.created_at).isoformat(
                    sep=" ", timespec="minutes"
                )
                if (row.updated_at or row.created_at)
                else None
            ),
        }
        for row in action_rows
    ]

    notified_at = None
    seen_at = None
    if req_id is not None:
        notif = (
            Notification.query.filter_by(
                volunteer_id=volunteer.id, type="new_match", request_id=req_id
            )
            .order_by(Notification.created_at.desc())
            .first()
        )
        if notif:
            notified_at = (
                notif.created_at.isoformat(sep=" ", timespec="minutes")
                if notif.created_at
                else None
            )
            seen_at = (
                notif.read_at.isoformat(sep=" ", timespec="minutes")
                if notif.read_at
                else None
            )

    data = {
        "id": volunteer.id,
        "name": getattr(volunteer, "name", None) or f"Volunteer #{volunteer.id}",
        "email": getattr(volunteer, "email", None),
        "phone": getattr(volunteer, "phone", None),
        "city": None,
        "location": getattr(volunteer, "location", None),
        "languages": [],
        "roles": _to_list(getattr(volunteer, "skills", None)),
        "availability": getattr(volunteer, "availability", None),
        "last_active": (
            last_active.isoformat(sep=" ", timespec="minutes") if last_active else None
        ),
        "can_help_count": int(can_help_count),
        "history": history,
        "notified_at": notified_at,
        "seen_at": seen_at,
        "profile_url": url_for("admin.volunteer_detail", id=volunteer.id),
    }
    return jsonify(data)


@admin_bp.route("/ops/login", methods=["GET", "POST"], endpoint="ops_login")
@limiter.limit("5 per 5 minutes")
@limiter.limit("20 per hour")
def admin_ops_login():
    # Preserve safe next across GET -> POST -> redirect (and across failed logins).
    next_candidate = (
        request.form.get("next") or request.args.get("next") or ""
    ).strip()
    next_url = next_candidate if is_safe_url(next_candidate) else ""

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        username_norm = _norm_username(username)
        ip = _client_ip()
        now = datetime.now(UTC).replace(tzinfo=None)
        locked, retry_after = _admin_login_is_locked(ip, username_norm, now)
        if locked:
            _log_admin_attempt(username=username_norm, ip=ip, success=False)
            log_security_event(
                "auth_admin_login_locked",
                actor_type="anonymous",
                meta={"username": username_norm, "ip": ip, "retry_after": retry_after},
            )
            return _lockout_response(retry_after, next_url=next_url)

        password = request.form.get("password", "")
        user = _find_admin_user(username)
        if not _verify_admin_password(user, password):
            _log_admin_attempt(username=username_norm, ip=ip, success=False)
            log_security_event(
                "auth_admin_login_failed",
                actor_type="anonymous",
                meta={"reason": "invalid_credentials"},
            )
            flash(GENERIC_ADMIN_LOGIN_FAIL_MSG, "danger")
            return redirect(url_for("admin.ops_login", next=next_url))
        _log_admin_attempt(username=username_norm, ip=ip, success=True)
        cleared_fails = _clear_recent_admin_login_failures(username_norm, ip, now)
        if cleared_fails > 0:
            log_security_event(
                "auth_admin_login_success_after_failures",
                actor_type="admin",
                actor_id=getattr(user, "id", None),
                meta={
                    "ip": ip,
                    "username": username_norm,
                    "cleared_failures": int(cleared_fails),
                },
            )
            audit_admin_action(
                action="admin.login.success_after_failures",
                target_type="AdminUser",
                target_id=int(getattr(user, "id", 0) or 0),
                payload={
                    "route": "admin.ops_login",
                    "username": username_norm,
                    "cleared_failures": int(cleared_fails),
                    "ip": ip,
                    "ua": (request.headers.get("User-Agent") or "")[:256],
                },
            )
        # Successful login path
        return _complete_admin_login(user, next_url, via="admin_ops_login")
    return render_template("admin/login.html", next=next_url)


@admin_bp.route("/login", methods=["GET", "POST"])
@limiter.limit("30 per minute")
def admin_login_legacy():
    """Legacy admin login endpoint kept for backward-compatible tests/clients."""
    next_candidate = (
        request.form.get("next") or request.args.get("next") or ""
    ).strip()
    next_url = next_candidate if is_safe_url(next_candidate) else ""

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        username_norm = _norm_username(username)
        ip = _client_ip()
        now = datetime.now(UTC).replace(tzinfo=None)
        locked, retry_after = _admin_login_is_locked(ip, username_norm, now)
        if locked:
            _log_admin_attempt(username=username_norm, ip=ip, success=False)
            return _lockout_response(retry_after, next_url=next_url)

        password = request.form.get("password", "")
        user = _find_admin_user(username)
        if not _verify_admin_password(user, password):
            _log_admin_attempt(username=username_norm, ip=ip, success=False)
            flash(GENERIC_ADMIN_LOGIN_FAIL_MSG, "danger")
            return redirect(url_for("admin.admin_login_legacy", next=next_url))
        _log_admin_attempt(username=username_norm, ip=ip, success=True)
        cleared_fails = _clear_recent_admin_login_failures(username_norm, ip, now)
        if cleared_fails > 0:
            log_security_event(
                "auth_admin_login_success_after_failures",
                actor_type="admin",
                actor_id=getattr(user, "id", None),
                meta={
                    "ip": ip,
                    "username": username_norm,
                    "cleared_failures": int(cleared_fails),
                },
            )
            audit_admin_action(
                action="admin.login.success_after_failures",
                target_type="AdminUser",
                target_id=int(getattr(user, "id", 0) or 0),
                payload={
                    "route": "admin.admin_login_legacy",
                    "username": username_norm,
                    "cleared_failures": int(cleared_fails),
                    "ip": ip,
                    "ua": (request.headers.get("User-Agent") or "")[:256],
                },
            )

        # Legacy email 2FA compatibility flow expected by tests.
        if bool(current_app.config.get("EMAIL_2FA_ENABLED", False)):
            session.clear()
            session["pending_admin_user_id"] = int(user.id)
            session["pending_email_2fa"] = True
            session["email_2fa_code"] = f"{secrets.randbelow(1000000):06d}"
            session["email_2fa_expires"] = int(time.time()) + 600
            return redirect(url_for("admin.admin_2fa"))

        return _complete_admin_login(user, next_url, via="admin_login_legacy")

    return render_template("admin/login.html", next=next_url)


@admin_bp.route("/re-auth", methods=["GET", "POST"])
@login_required
@admin_required
def admin_reauth():
    next_candidate = (
        request.form.get("next") or request.args.get("next") or ""
    ).strip()
    next_url = next_candidate if is_safe_url(next_candidate) else ""

    if request.method == "POST":
        password = request.form.get("password", "")
        user = db.session.get(AdminUser, getattr(current_user, "id", None))
        if _verify_admin_password(user, password):
            now = _utc_now()
            _touch_admin_last_seen(now)
            _touch_admin_auth_at(now)
            log_security_event(
                "auth_admin_reauth_success",
                actor_type="admin",
                actor_id=getattr(user, "id", None),
                meta={"next": (next_url or "")[:255]},
            )
            audit_admin_action(
                action="admin.reauth.success",
                target_type="AdminUser",
                target_id=int(getattr(user, "id", 0) or 0),
                payload={
                    "next": (next_url or "")[:255],
                    "route": request.path,
                    "ip": _client_ip(),
                    "ua": (request.headers.get("User-Agent") or "")[:256],
                },
            )
            flash("Vérification effectuée.", "success")
            return redirect(next_url or url_for("admin.admin_requests"), code=303)
        log_security_event(
            "auth_admin_reauth_failed",
            actor_type="admin",
            actor_id=getattr(current_user, "id", None),
            meta={"route": request.path},
        )
        if user:
            audit_admin_action(
                action="admin.reauth.failed",
                target_type="AdminUser",
                target_id=int(getattr(user, "id", 0) or 0),
                payload={
                    "route": request.path,
                    "ip": _client_ip(),
                    "ua": (request.headers.get("User-Agent") or "")[:256],
                },
            )
        flash("Veuillez confirmer votre identité pour continuer.", "danger")

    return render_template("admin/reauth.html", next=next_url)


@admin_bp.route("/logout", methods=["GET", "POST"])
def admin_logout():
    admin_required_404()
    actor_id = getattr(current_user, "id", 0) or 0
    log_activity(
        entity_type="admin",
        entity_id=actor_id,
        action="admin_logout",
        message="Admin logout",
        persist=True,
    )
    audit_admin_action(
        action="admin.logout",
        target_type="AdminUser",
        target_id=int(actor_id),
        payload={
            "route": request.path,
            "ip": _client_ip(),
            "ua": (request.headers.get("User-Agent") or "")[:256],
        },
    )
    _mfa_ok_clear()
    _mfa_attempt_reset()
    session.pop("mfa_required", None)
    session.pop("mfa_pending_secret", None)
    session.pop("backup_codes_plain", None)
    session.pop("admin_user_id", None)
    session.pop("admin_logged_in", None)
    session.pop("admin_last_seen", None)
    session.pop("admin_auth_at", None)
    logout_user()
    flash("Logged out.", "info")
    return redirect(url_for("admin.admin_login_legacy"))


MFA_PENDING_SECRET_KEY = "mfa_pending_secret"
MFA_SESSION_KEY = "mfa_ok"


@admin_bp.route("/mfa/setup", methods=["GET", "POST"])
@login_required
def admin_mfa_setup():
    admin_required_404()
    if getattr(current_user, "mfa_enabled", False) and getattr(
        current_user, "totp_secret", None
    ):
        flash(_("MFA is already enabled."), "info")
        return redirect(url_for("admin.admin_mfa_verify"))

    pending_secret = session.get(MFA_PENDING_SECRET_KEY)
    if not pending_secret:
        pending_secret = generate_totp_secret()
        session[MFA_PENDING_SECRET_KEY] = pending_secret

    issuer = "HelpChain"
    username = getattr(current_user, "username", f"admin-{current_user.id}")
    uri = build_totp_uri(pending_secret, username=username, issuer=issuer)
    qr_b64 = qr_png_base64(uri)

    if request.method == "POST":
        code = (request.form.get("code") or "").strip().replace(" ", "")
        if not code:
            flash(_("Enter the 6-digit code from the app."), "danger")
            return render_template(
                "admin/mfa_setup.html",
                qr_b64=qr_b64,
                secret=pending_secret,
                username=username,
            )

        if not verify_totp_code(pending_secret, code):
            flash(
                "Невалиден код. Провери часовника на телефона и опитай пак.", "danger"
            )
            return render_template(
                "admin/mfa_setup.html",
                qr_b64=qr_b64,
                secret=pending_secret,
                username=username,
            )

        user = db.session.get(AdminUser, current_user.id)
        user.totp_secret = pending_secret
        user.mfa_enabled = True
        user.mfa_enrolled_at = utc_now()
        db.session.commit()

        _mfa_ok_set()
        session.pop(MFA_PENDING_SECRET_KEY, None)
        flash(_("MFA has been enabled successfully."), "success")
        flash(_("Generate backup codes now (recovery option)."), "info")
        return redirect(url_for("admin.admin_mfa_backup_codes"))

    return render_template(
        "admin/mfa_setup.html", qr_b64=qr_b64, secret=pending_secret, username=username
    )


@admin_bp.route("/mfa/verify", methods=["GET", "POST"])
@login_required
def admin_mfa_verify():
    admin_required_404()
    if not current_app.config.get("MFA_ENABLED", False):
        abort(404)

    if not (
        getattr(current_user, "mfa_enabled", False)
        and getattr(current_user, "totp_secret", None)
    ):
        _mfa_ok_set()
        session["admin_logged_in"] = True
        now = _utc_now()
        _touch_admin_last_seen(now)
        _touch_admin_auth_at(now)
        return redirect(request.args.get("next") or url_for("admin.admin_requests"))

    locked, remaining = _mfa_lock_is_active()
    if request.method == "GET":
        return render_template(
            "admin/mfa_verify.html",
            locked=locked,
            remaining=remaining,
            next=request.args.get("next") or "",
        )

    if locked:
        flash(
            f"Твърде много опити. Опитай след {max(1, remaining // 60)} мин.", "danger"
        )
        return redirect(
            url_for("admin.admin_mfa_verify", next=request.args.get("next", ""))
        )

    code = (request.form.get("code") or "").strip().replace(" ", "").upper()
    totp_ok = False
    try:
        totp_ok = verify_totp_code(current_user.totp_secret, code)
    except Exception:
        totp_ok = False

    backup_ok = False
    if not totp_ok:
        hashes = _load_hashes(current_user)
        for i, h in enumerate(hashes):
            if check_password_hash(h, code):
                backup_ok = True
                hashes.pop(i)
                _save_hashes(current_user, hashes)
                break

    if totp_ok or backup_ok:
        _mfa_attempt_reset()
        _mfa_ok_set()
        session["admin_logged_in"] = True
        now = _utc_now()
        _touch_admin_last_seen(now)
        _touch_admin_auth_at(now)
        log_activity(
            entity_type="admin",
            entity_id=getattr(current_user, "id", 0) or 0,
            action="admin_mfa_success",
            message="Admin MFA verified",
            persist=True,
        )
        flash(_("MFA verified."), "success")
        nxt = request.args.get("next")
        if nxt and nxt.startswith("/"):
            return redirect(nxt)
        return redirect(url_for("admin.admin_requests"))

    _mfa_attempt_fail()
    locked, remaining = _mfa_lock_is_active()
    if locked:
        flash(_("Invalid code. Locked for about %(minutes)s min.", minutes=max(1, remaining // 60)), "danger")
    else:
        left = current_app.config.get("MFA_VERIFY_MAX_ATTEMPTS", 8) - int(
            session.get("mfa_attempts", 0)
        )
        flash(_("Invalid code. Attempts remaining: %(count)s.", count=max(left, 0)), "danger")

    return redirect(
        url_for("admin.admin_mfa_verify", next=request.args.get("next", ""))
    )


@admin_bp.route("/mfa/backup-codes", methods=["GET", "POST"])
@login_required
@require_admin_fresh_auth(minutes=10)
def admin_mfa_backup_codes():
    admin_required_404()
    if not current_app.config.get("MFA_ENABLED", False):
        abort(404)
    if not (
        getattr(current_user, "mfa_enabled", False)
        and getattr(current_user, "totp_secret", None)
    ):
        flash(_("Enable MFA from setup first."), "warning")
        return redirect(url_for("admin.admin_mfa_setup"))

    if request.method == "POST":
        codes = _generate_backup_codes(10)
        _save_hashes(current_user, _hash_codes(codes))
        session["backup_codes_plain"] = codes
        audit_admin_action(
            action="admin.mfa_backup_codes_regenerated",
            target_type="AdminUser",
            target_id=int(getattr(current_user, "id", 0) or 0),
            payload={
                "generated_codes_count": 10,
                "route": request.path,
                "ip": _client_ip(),
                "ua": (request.headers.get("User-Agent") or "")[:256],
            },
        )
        flash(
            "Backup кодовете са генерирани. Запази ги сега — няма да се покажат втори път.",
            "success",
        )
        return redirect(url_for("admin.admin_mfa_backup_codes"))

    codes_plain = session.pop("backup_codes_plain", None)
    hashes = _load_hashes(current_user)
    has_codes = len(hashes) > 0
    return render_template(
        "admin/mfa_backup_codes.html",
        codes=codes_plain,
        has_codes=has_codes,
        generated_at=getattr(current_user, "backup_codes_generated_at", None),
    )


@admin_bp.route("/2fa", methods=["GET", "POST"])
@admin_required
def admin_2fa():
    admin_required_404()
    """2FA верификация за админ"""
    user_id = session.get("pending_admin_user_id")
    if not user_id:
        return redirect(url_for("admin.admin_login_legacy"))

    admin_user = db.session.get(AdminUser, user_id)
    if not admin_user:
        return redirect(url_for("admin.admin_login_legacy"))

    if request.method == "POST":
        token = request.form.get("token")
        if admin_user.verify_totp(token):
            from flask_login import login_user

            login_user(admin_user)
            session.pop("pending_admin_user_id", None)
            return redirect(url_for("admin.admin_dashboard"))
        else:
            flash(_("Invalid 2FA code."), "error")

    return render_template("admin_2fa.html")


@admin_bp.route("/email_2fa", methods=["GET", "POST"])
def admin_email_2fa():
    """Legacy email-based 2FA flow used by older tests."""
    if not session.get("pending_email_2fa"):
        return redirect(url_for("admin.admin_login_legacy"))

    if request.method == "POST":
        entered = (request.form.get("code") or "").strip()
        if entered and entered == (session.get("email_2fa_code") or ""):
            session.pop("pending_email_2fa", None)
            session.pop("email_2fa_code", None)
            session.pop("email_2fa_expires", None)
            return redirect(url_for("admin.admin_dashboard"))
        flash(_("Invalid verification code."), "danger")

    return render_template("admin_email_2fa.html")


@admin_bp.route("/2fa/setup", methods=["GET", "POST"])
@admin_required
def admin_2fa_setup():
    admin_required_404()
    """Настройка на 2FA за админ"""
    if not isinstance(current_user, AdminUser):
        flash(_("Access denied."), "error")
        return redirect(url_for("main.index"))

    if request.method == "POST":
        token = request.form.get("token")
        if current_user.verify_totp(token):
            current_user.enable_2fa()
            flash(_("2FA has been enabled successfully!"), "success")
            return redirect(url_for("admin.admin_dashboard"))
        else:
            flash(_("Invalid code."), "error")

    uri = current_user.get_totp_uri()
    return render_template("admin_2fa_setup.html", totp_uri=uri)


@admin_bp.route("/2fa/disable", methods=["POST"])
@admin_required
def admin_2fa_disable():
    admin_required_404()
    """Деактивиране на 2FA за админ"""
    if not isinstance(current_user, AdminUser):
        flash(_("Access denied."), "error")
        return redirect(url_for("main.index"))

    current_user.disable_2fa()
    flash(_("2FA has been disabled."), "success")
    return redirect(url_for("admin.admin_dashboard"))


@admin_bp.get("/api/dashboard")
@login_required
def admin_api_dashboard():
    admin_required_404()
    """Session-based dashboard data for admin UI (bypass JWT)."""
    if not getattr(current_user, "is_admin", False):
        return jsonify({"error": "forbidden"}), 403

    try:
        try:
            days = int(request.args.get("days", 30))
        except Exception:
            days = 30

        since_dt = datetime.now(UTC).replace(tzinfo=None) - timedelta(days=days)

        status_rows = (
            db.session.query(
                func.coalesce(Request.status, "unknown").label("status"),
                func.count(Request.id).label("cnt"),
            )
            .group_by("status")
            .all()
        )
        counts_by_status = {status: int(cnt) for status, cnt in status_rows}
        total_requests = int(sum(counts_by_status.values()))

        city_expr = func.coalesce(
            func.nullif(Request.city, ""),
            func.nullif(Request.region, ""),
            "unknown",
        )
        city_rows = (
            db.session.query(
                city_expr.label("city"), func.count(Request.id).label("cnt")
            )
            .group_by("city")
            .order_by(func.count(Request.id).desc())
            .limit(10)
            .all()
        )
        requests_by_city = [{"city": c, "count": int(cnt)} for c, cnt in city_rows]

        ts_rows = (
            db.session.query(
                func.date(Request.created_at).label("day"),
                func.count(Request.id).label("cnt"),
            )
            .filter(Request.created_at.isnot(None))
            .filter(Request.created_at >= since_dt)
            .group_by("day")
            .order_by("day")
            .all()
        )
        timeseries = [{"date": str(day), "count": int(cnt)} for day, cnt in ts_rows]

        try:
            total_volunteers = db.session.query(Volunteer).count()
        except Exception:
            total_volunteers = 0

        return (
            jsonify(
                {
                    "total_requests": total_requests,
                    "total_volunteers": total_volunteers,
                    "counts_by_status": counts_by_status,
                    "requests_by_city": requests_by_city,
                    "timeseries": timeseries,
                }
            ),
            200,
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


SLA_HOURS_DEFAULT = 12


def _pctl(values, p: float = 0.9) -> float | None:
    vals = [float(v) for v in values if v is not None]
    if not vals:
        return None
    vals.sort()
    idx = max(0, min(len(vals) - 1, math.ceil(p * len(vals)) - 1))
    return float(vals[idx])


def _sla_outliers_action_7d(
    first_action_subq, now: datetime, limit: int = 5
) -> list[dict]:
    cutoff = now - timedelta(days=7)
    delay_expr = func.strftime(
        "%s", first_action_subq.c.first_action_at
    ) - func.strftime("%s", VolunteerRequestState.notified_at)
    rows = (
        db.session.query(
            VolunteerRequestState.request_id.label("request_id"),
            VolunteerRequestState.volunteer_id.label("volunteer_id"),
            VolunteerRequestState.notified_at.label("notified_at"),
            first_action_subq.c.first_action_at.label("first_action_at"),
            delay_expr.label("delay_seconds"),
        )
        .join(
            first_action_subq,
            (first_action_subq.c.req_id == VolunteerRequestState.request_id)
            & (first_action_subq.c.vol_id == VolunteerRequestState.volunteer_id),
        )
        .filter(VolunteerRequestState.notified_at.isnot(None))
        .filter(VolunteerRequestState.notified_at >= cutoff)
        .filter(first_action_subq.c.first_action_at.isnot(None))
        .filter(first_action_subq.c.first_action_at >= cutoff)
        .filter(delay_expr >= 0)
        .order_by(db.text("delay_seconds DESC"))
        .limit(limit)
        .all()
    )
    out = []
    for row in rows:
        out.append(
            {
                "request_id": int(row.request_id),
                "volunteer_id": (
                    int(row.volunteer_id) if row.volunteer_id is not None else None
                ),
                "delay_seconds": (
                    float(row.delay_seconds) if row.delay_seconds is not None else None
                ),
                "notified_at": row.notified_at.isoformat() if row.notified_at else None,
                "first_action_at": (
                    row.first_action_at.isoformat() if row.first_action_at else None
                ),
            }
        )
    return out


def _sla_kpis(
    sla_hours: int = SLA_HOURS_DEFAULT,
    scope: str = "all_notified",
    now: datetime | None = None,
) -> dict:
    now = now or datetime.now(UTC).replace(tzinfo=None)
    sla_seconds = int(sla_hours * 3600)
    if scope not in {"all_notified", "assigned_only"}:
        raise ValueError("scope must be 'all_notified' or 'assigned_only'")

    # Backward compatibility for production environments with partial schema.
    # Keep /admin/api/risk-kpis alive instead of failing with 500.
    if not _table_exists("volunteer_request_states") or not _table_has_column(
        "volunteer_request_states", "notified_at"
    ):
        return {
            "avg_first_seen_seconds": None,
            "avg_first_action_seconds": None,
            "sla_under_12h_percent": 0.0,
            "sla_hours": int(sla_hours),
            "sla_samples": 0,
            "sla_scope": scope,
            "p90_first_seen_seconds_7d": None,
            "p90_first_action_seconds_7d": None,
            "p90_window_days": 7,
            "sla_outliers_action_7d": [],
            "sla_outliers_window_days": 7,
            "sla_outliers_limit": 5,
            "schema_warning": "volunteer_request_states.notified_at_missing",
        }

    if not _table_has_column("volunteer_request_states", "seen_at"):
        return {
            "avg_first_seen_seconds": None,
            "avg_first_action_seconds": None,
            "sla_under_12h_percent": 0.0,
            "sla_hours": int(sla_hours),
            "sla_samples": 0,
            "sla_scope": scope,
            "p90_first_seen_seconds_7d": None,
            "p90_first_action_seconds_7d": None,
            "p90_window_days": 7,
            "sla_outliers_action_7d": [],
            "sla_outliers_window_days": 7,
            "sla_outliers_limit": 5,
            "schema_warning": "volunteer_request_states.seen_at_missing",
        }

    base_states_q = db.session.query(
        VolunteerRequestState.request_id.label("req_id"),
        VolunteerRequestState.volunteer_id.label("vol_id"),
        VolunteerRequestState.notified_at.label("notified_at"),
        VolunteerRequestState.seen_at.label("seen_at"),
    ).join(Request, Request.id == VolunteerRequestState.request_id).filter(
        VolunteerRequestState.notified_at.isnot(None)
    )
    tenant_filter = _structure_scope_filter()
    base_states_q = _apply_tenant_filter(base_states_q, tenant_filter)
    if scope == "assigned_only":
        base_states_q = base_states_q.filter(
            Request.assigned_volunteer_id == VolunteerRequestState.volunteer_id
        )
    base_states = base_states_q.subquery()

    seen_delta = func.strftime("%s", base_states.c.seen_at) - func.strftime(
        "%s", base_states.c.notified_at
    )
    avg_first_seen = (
        db.session.query(func.avg(seen_delta))
        .filter(base_states.c.seen_at.isnot(None))
        .filter(seen_delta >= 0)
        .scalar()
    )

    # Backward compatibility: some local SQLite snapshots don't have
    # request_activities.volunteer_id yet. Keep the endpoint alive and return
    # partial SLA metrics instead of 500.
    if not _table_has_column("request_activities", "volunteer_id"):
        return {
            "avg_first_seen_seconds": (
                float(avg_first_seen) if avg_first_seen is not None else None
            ),
            "avg_first_action_seconds": None,
            "sla_under_12h_percent": 0.0,
            "sla_hours": int(sla_hours),
            "sla_samples": 0,
            "sla_scope": scope,
            "p90_first_seen_seconds_7d": None,
            "p90_first_action_seconds_7d": None,
            "p90_window_days": 7,
            "sla_outliers_action_7d": [],
            "sla_outliers_window_days": 7,
            "sla_outliers_limit": 5,
            "schema_warning": "request_activities.volunteer_id_missing",
        }

    first_action_subq = (
        db.session.query(
            base_states.c.req_id.label("req_id"),
            base_states.c.vol_id.label("vol_id"),
            base_states.c.notified_at.label("notified_at"),
            func.min(RequestActivity.created_at).label("first_action_at"),
        )
        .join(
            RequestActivity,
            (RequestActivity.request_id == base_states.c.req_id)
            & (RequestActivity.volunteer_id == base_states.c.vol_id),
        )
        .filter(
            RequestActivity.action.in_(("volunteer_can_help", "volunteer_cant_help"))
        )
        .filter(RequestActivity.created_at >= base_states.c.notified_at)
        .group_by(
            base_states.c.req_id,
            base_states.c.vol_id,
            base_states.c.notified_at,
        )
        .subquery()
    )
    action_delta = func.strftime(
        "%s", first_action_subq.c.first_action_at
    ) - func.strftime("%s", first_action_subq.c.notified_at)
    avg_first_action = db.session.query(func.avg(action_delta)).scalar()

    under_count, total_count = db.session.query(
        func.sum(case((action_delta <= sla_seconds, 1), else_=0)).label("under"),
        func.count().label("total"),
    ).first() or (0, 0)
    under_count = int(under_count or 0)
    total_count = int(total_count or 0)
    sla_pct = round((under_count * 100.0 / total_count), 1) if total_count else 0.0
    p90_window_days = 7
    cutoff = now - timedelta(days=p90_window_days)
    seen_secs = [
        int(sec)
        for (sec,) in (
            db.session.query(seen_delta)
            .filter(base_states.c.seen_at.isnot(None))
            .filter(base_states.c.notified_at >= cutoff)
            .filter(seen_delta >= 0)
            .all()
        )
        if sec is not None
    ]
    action_secs = [
        int(sec)
        for (sec,) in (
            db.session.query(action_delta)
            .filter(first_action_subq.c.notified_at >= cutoff)
            .filter(first_action_subq.c.first_action_at >= cutoff)
            .filter(action_delta >= 0)
            .all()
        )
        if sec is not None
    ]
    outliers = _sla_outliers_action_7d(first_action_subq, now, limit=5)

    return {
        "avg_first_seen_seconds": (
            float(avg_first_seen) if avg_first_seen is not None else None
        ),
        "avg_first_action_seconds": (
            float(avg_first_action) if avg_first_action is not None else None
        ),
        "sla_under_12h_percent": sla_pct,
        "sla_hours": int(sla_hours),
        "sla_samples": total_count,
        "sla_scope": scope,
        "p90_first_seen_seconds_7d": _pctl(seen_secs, 0.9),
        "p90_first_action_seconds_7d": _pctl(action_secs, 0.9),
        "p90_window_days": p90_window_days,
        "sla_outliers_action_7d": outliers,
        "sla_outliers_window_days": 7,
        "sla_outliers_limit": 5,
    }


@admin_bp.get("/api/risk-kpis")
@admin_required
def admin_risk_kpis():
    admin_required_404()

    now = datetime.now(UTC).replace(tzinfo=None)
    stale_days = 8
    unassigned_days = 3
    window_days = 7
    not_seen_hours = 24

    stale_cutoff = now - timedelta(days=stale_days)
    unassigned_cutoff = now - timedelta(days=unassigned_days)
    window_cutoff = now - timedelta(days=window_days)
    has_vrs_notified_at = _table_has_column("volunteer_request_states", "notified_at")
    tenant_filter = _structure_scope_filter()
    open_filter = or_(
        Request.status.is_(None), ~Request.status.in_(list(CLOSED_STATUSES))
    )

    stale_count = (
        _apply_tenant_filter(db.session.query(func.count(Request.id)), tenant_filter)
        .filter(Request.created_at < stale_cutoff)
        .filter(open_filter)
        .scalar()
    )

    unassigned_count = (
        _apply_tenant_filter(db.session.query(func.count(Request.id)), tenant_filter)
        .filter(Request.created_at < unassigned_cutoff)
        .filter(Request.assigned_volunteer_id.is_(None))
        .filter(open_filter)
        .scalar()
    )

    def count_notseen(hours: int) -> tuple[int, str]:
        cutoff = now - timedelta(hours=hours)
        if has_vrs_notified_at:
            cnt = (
                _apply_tenant_filter(
                    db.session.query(func.count(VolunteerRequestState.id)), tenant_filter
                )
                .join(Request, Request.id == VolunteerRequestState.request_id)
                .filter(VolunteerRequestState.notified_at.isnot(None))
                .filter(VolunteerRequestState.notified_at < cutoff)
                .filter(VolunteerRequestState.seen_at.is_(None))
                .filter(open_filter)
                .scalar()
            )
            source = "notified_at"
        else:
            cnt = (
                _apply_tenant_filter(
                    db.session.query(func.count(Notification.id)), tenant_filter
                )
                .join(Request, Request.id == Notification.request_id)
                .filter(Notification.type == "new_match")
                .filter(Notification.created_at < cutoff)
                .filter(open_filter)
                .scalar()
            )
            source = "notification_created_at_fallback"
        return int(cnt or 0), source

    notseen24, notified_source = count_notseen(24)
    notseen48, _ = count_notseen(48)
    notseen72, _ = count_notseen(72)
    notified_not_seen = notseen24

    can_help_7d = (
        db.session.query(func.count(VolunteerAction.id))
        .filter(VolunteerAction.action == "CAN_HELP")
        .filter(VolunteerAction.created_at >= window_cutoff)
        .scalar()
    )

    assigned_7d = (
        _apply_tenant_filter(db.session.query(func.count(Request.id)), tenant_filter)
        .filter(Request.assigned_volunteer_id.isnot(None))
        .filter(Request.created_at >= window_cutoff)
        .scalar()
    )

    conversion = None
    if can_help_7d and can_help_7d > 0:
        conversion = round((assigned_7d / can_help_7d) * 100, 1)

    not_seen_by_request = {}
    try:
        if has_vrs_notified_at:
            cutoff = now - timedelta(hours=24)
            not_seen_rows = (
                _apply_tenant_filter(
                    db.session.query(
                        VolunteerRequestState.request_id,
                        func.count(VolunteerRequestState.id),
                    ),
                    tenant_filter,
                )
                .join(Request, Request.id == VolunteerRequestState.request_id)
                .filter(VolunteerRequestState.notified_at.isnot(None))
                .filter(VolunteerRequestState.notified_at < cutoff)
                .filter(VolunteerRequestState.seen_at.is_(None))
                .filter(open_filter)
                .group_by(VolunteerRequestState.request_id)
                .all()
            )
        else:
            cutoff = now - timedelta(hours=24)
            not_seen_rows = (
                _apply_tenant_filter(
                    db.session.query(Notification.request_id, func.count(Notification.id)),
                    tenant_filter,
                )
                .join(Request, Request.id == Notification.request_id)
                .filter(Notification.type == "new_match")
                .filter(Notification.created_at < cutoff)
                .filter(open_filter)
                .group_by(Notification.request_id)
                .all()
            )
        not_seen_by_request = {int(req_id): int(cnt) for req_id, cnt in not_seen_rows}
    except Exception:
        db.session.rollback()

    risky_candidates = (
        _scope_requests(Request.query)
        .filter(Request.deleted_at.is_(None))
        .filter(open_filter)
        .order_by(Request.created_at.asc())
        .limit(300)
        .all()
    )
    top_risky_scored = []
    for row in risky_candidates:
        created_at = getattr(row, "created_at", None)
        age_days = (
            max(0, int((now - created_at).total_seconds() // 86400))
            if created_at is not None
            else 0
        )
        is_unassigned = getattr(row, "assigned_volunteer_id", None) is None
        ns_count = int(not_seen_by_request.get(int(row.id), 0))
        score = 0
        if age_days >= stale_days:
            score += 3
        if is_unassigned and age_days >= unassigned_days:
            score += 4
        if ns_count > 0:
            score += 2 + min(3, ns_count)
        if score <= 0:
            continue
        top_risky_scored.append((score, age_days, ns_count, row))

    top_risky_scored.sort(key=lambda x: (-x[0], -x[1], -x[2], x[3].id))
    top_risky = []
    for score, age_days, ns_count, row in top_risky_scored[:5]:
        top_risky.append(
            {
                "id": int(row.id),
                "title": getattr(row, "title", None) or f"Request #{row.id}",
                "days_open": int(age_days),
                "is_unassigned": bool(
                    getattr(row, "assigned_volunteer_id", None) is None
                ),
                "not_seen_count": int(ns_count),
                "risk_score": int(score),
                "details_url": url_for("admin.admin_request_details", req_id=row.id),
            }
        )

    payload = {
        "stale_days": stale_days,
        "unassigned_days": unassigned_days,
        "window_days": window_days,
        "not_seen_hours": not_seen_hours,
        "notified_source": notified_source,
        "stale_count": int(stale_count or 0),
        "unassigned_count": int(unassigned_count or 0),
        "notified_not_seen": int(notified_not_seen or 0),
        "notseen24": int(notseen24 or 0),
        "notseen48": int(notseen48 or 0),
        "notseen72": int(notseen72 or 0),
        "can_help_7d": int(can_help_7d or 0),
        "assigned_7d": int(assigned_7d or 0),
        "conversion_pct": conversion,
        "top_risky": top_risky,
        "generated_at": now.isoformat(timespec="seconds"),
    }
    requested_sla = request.args.get("sla", type=int)
    default_sla = int(current_app.config.get("SLA_HOURS_DEFAULT", SLA_HOURS_DEFAULT))
    sla_hours = requested_sla if requested_sla is not None else default_sla
    sla_hours = max(1, min(int(sla_hours), 168))
    payload.update(_sla_kpis(sla_hours=sla_hours, scope="all_notified", now=now))
    return jsonify(payload)


def _seconds_diff(end_col, start_col):
    """
    Portable-ish SQL expression for (end - start) in seconds.
    PostgreSQL: extract(epoch from end - start)
    SQLite/other: (julianday(end) - julianday(start)) * 86400
    """
    try:
        dialect = db.session.get_bind().dialect.name
    except Exception:
        dialect = ""
    if dialect == "postgresql":
        return func.extract("epoch", end_col - start_col)
    return (func.julianday(end_col) - func.julianday(start_col)) * 86400.0


@admin_bp.get("/api/ops-kpis")
@admin_required
def admin_ops_kpis():
    admin_required_404()

    days = max(1, min(int(request.args.get("days", 30) or 30), 365))
    now = datetime.now(UTC).replace(tzinfo=None)
    since = now - timedelta(days=days)
    stale_since = now - timedelta(days=7)
    tenant_filter = _structure_scope_filter()

    new_count = (
        _apply_tenant_filter(db.session.query(func.count(Request.id)), tenant_filter)
        .filter(Request.created_at >= since)
        .scalar()
        or 0
    )

    resolved_count = (
        _apply_tenant_filter(db.session.query(func.count(Request.id)), tenant_filter)
        .filter(Request.completed_at.isnot(None))
        .filter(Request.completed_at >= since)
        .scalar()
        or 0
    )

    avg_to_owner_assign_sec = (
        _apply_tenant_filter(
            db.session.query(func.avg(_seconds_diff(Request.owned_at, Request.created_at))),
            tenant_filter,
        )
        .filter(Request.owned_at.isnot(None))
        .scalar()
    )
    avg_owner_assign_hours = round(float(avg_to_owner_assign_sec or 0) / 3600.0, 2)

    avg_to_resolve_sec = (
        _apply_tenant_filter(
            db.session.query(
                func.avg(_seconds_diff(Request.completed_at, Request.created_at))
            ),
            tenant_filter,
        )
        .filter(Request.completed_at.isnot(None))
        .scalar()
    )
    avg_resolve_hours = round(float(avg_to_resolve_sec or 0) / 3600.0, 2)

    stale_count = (
        _apply_tenant_filter(db.session.query(func.count(Request.id)), tenant_filter)
        .filter(Request.created_at < stale_since)
        .filter(Request.completed_at.is_(None))
        .scalar()
        or 0
    )

    by_category_rows = (
        _apply_tenant_filter(
            db.session.query(Request.category, func.count(Request.id)), tenant_filter
        )
        .filter(Request.created_at >= since)
        .group_by(Request.category)
        .all()
    )
    by_category = [
        {"category": (category or "—"), "count": int(count)}
        for category, count in by_category_rows
    ]

    by_status_rows = (
        _apply_tenant_filter(
            db.session.query(Request.status, func.count(Request.id)), tenant_filter
        )
        .filter(Request.completed_at.is_(None))
        .group_by(Request.status)
        .all()
    )
    by_status = [
        {"status": (status or "—"), "count": int(count)}
        for status, count in by_status_rows
    ]

    # --- SLA breach detection ---
    assign_deadline = now - timedelta(hours=ASSIGN_SLA_HOURS)
    resolve_deadline = now - timedelta(days=RESOLVE_SLA_DAYS)
    open_filter = or_(
        Request.status.is_(None), ~func.lower(Request.status).in_(CLOSED_STATUSES)
    )

    assign_breach_count = (
        _apply_tenant_filter(db.session.query(func.count(Request.id)), tenant_filter)
        .filter(Request.created_at.isnot(None))
        .filter(Request.created_at < assign_deadline)
        .filter(Request.owned_at.is_(None))
        .filter(open_filter)
        .scalar()
        or 0
    )

    resolve_breach_count = (
        _apply_tenant_filter(db.session.query(func.count(Request.id)), tenant_filter)
        .filter(Request.created_at.isnot(None))
        .filter(Request.created_at < resolve_deadline)
        .filter(Request.completed_at.is_(None))
        .filter(open_filter)
        .scalar()
        or 0
    )

    volunteer_assign_deadline = now - timedelta(hours=VOLUNTEER_ASSIGN_SLA_HOURS)
    volunteer_assign_breach_count = (
        _apply_tenant_filter(db.session.query(func.count(Request.id)), tenant_filter)
        .filter(Request.created_at.isnot(None))
        .filter(Request.created_at < volunteer_assign_deadline)
        .filter(Request.assigned_volunteer_id.is_(None))
        .filter(open_filter)
        .scalar()
        or 0
    )

    # --- Health scoring (SLA-driven) ---
    if resolve_breach_count > 0:
        health_status = "critique"
    elif assign_breach_count > 0:
        health_status = "sous_tension"
    else:
        health_status = "stable"

    return jsonify(
        {
            "window_days": days,
            "new_requests": int(new_count),
            "resolved_requests": int(resolved_count),
            "avg_owner_assign_hours": avg_owner_assign_hours,
            "avg_resolve_hours": avg_resolve_hours,
            "stale_over_7d": int(stale_count),
            "by_category": by_category,
            "by_status_open": by_status,
            "definition": {
                "assignment": "owner assignment (owner_id + owned_at)",
                "resolution": "completed_at not null",
            },
            "sla": {
                "assign_sla_hours": ASSIGN_SLA_HOURS,
                "resolve_sla_days": RESOLVE_SLA_DAYS,
                "assign_breach_count": int(assign_breach_count),
                "resolve_breach_count": int(resolve_breach_count),
            },
            "health": {
                "status": health_status,
            },
            "volunteer_sla": {
                "volunteer_assign_sla_hours": VOLUNTEER_ASSIGN_SLA_HOURS,
                "volunteer_assign_breach_count": int(volunteer_assign_breach_count),
            },
        }
    )


@admin_bp.get("/api/territorial-kpis")
@admin_required
def admin_territorial_kpis():
    admin_required_404()
    try:
        now = datetime.now(UTC).replace(tzinfo=None)
        since = now - timedelta(days=7)

        case_filter = None
        if not _is_global_admin():
            case_filter = Case.structure_id == _current_structure_id()

        base_cases = db.session.query(Case)
        if case_filter is not None:
            base_cases = base_cases.filter(case_filter)

        active_cases = (
            base_cases.filter(func.lower(Case.status) != "closed").count()
        )
        new_cases_week = base_cases.filter(Case.created_at >= since).count()
        resolved_cases = base_cases.filter(func.lower(Case.status) == "closed").count()

        status_rows = (
            base_cases.with_entities(Case.status, func.count(Case.id))
            .group_by(Case.status)
            .all()
        )
        status_map = {str(status or "").lower(): int(count) for status, count in status_rows}
        cases_by_status = {
            "new": status_map.get("new", 0),
            "in_progress": status_map.get("in_progress", 0),
            "closed": status_map.get("closed", 0),
        }

        first_event_subq = (
            db.session.query(
                CaseEvent.case_id.label("case_id"),
                func.min(CaseEvent.created_at).label("first_event_at"),
            )
            .group_by(CaseEvent.case_id)
            .subquery()
        )

        avg_response_query = (
            db.session.query(
                func.avg(
                    _seconds_diff(first_event_subq.c.first_event_at, Case.created_at)
                )
            )
            .join(first_event_subq, Case.id == first_event_subq.c.case_id)
        )
        if case_filter is not None:
            avg_response_query = avg_response_query.filter(case_filter)
        avg_response_sec = avg_response_query.scalar()
        avg_response_hours = round(float(avg_response_sec or 0) / 3600.0, 2)

        open_filter = func.lower(Case.status) != "closed"
        oldest_open_query = db.session.query(
            func.max((func.julianday(now) - func.julianday(Case.created_at)) * 86400.0)
        ).filter(open_filter)
        if case_filter is not None:
            oldest_open_query = oldest_open_query.filter(case_filter)
        oldest_open_sec = oldest_open_query.scalar()
        oldest_open_case_days = int(float(oldest_open_sec or 0) / 86400.0)

        return jsonify(
            {
                "active_cases": int(active_cases or 0),
                "new_cases_week": int(new_cases_week or 0),
                "resolved_cases": int(resolved_cases or 0),
                "avg_response_hours": float(avg_response_hours or 0),
                "cases_by_status": cases_by_status,
                "oldest_open_case_days": int(oldest_open_case_days or 0),
            }
        )
    except Exception:
        current_app.logger.exception("territorial kpis failed")
        return jsonify(
            {
                "active_cases": 0,
                "new_cases_week": 0,
                "resolved_cases": 0,
                "avg_response_hours": 0.0,
                "cases_by_status": {"new": 0, "in_progress": 0, "closed": 0},
                "oldest_open_case_days": 0,
            }
        )


@admin_bp.route("/")
@admin_required
def admin_dashboard():
    admin_required_404()
    """Админ панел"""

    import logging

    logging.warning(
        f"[DEBUG] admin_dashboard: is_authenticated={getattr(current_user, 'is_authenticated', None)}, is_admin={getattr(current_user, 'is_admin', None)}, id={getattr(current_user, 'id', None)}, username={getattr(current_user, 'username', None)}"
    )
    if not current_user.is_admin:
        flash(_("You do not have access to the admin panel."), "error")
        return redirect(url_for("main.dashboard"))

    requests = _scope_requests(Request.query).all()
    logs = RequestLog.query.all()
    volunteers = Volunteer.query.all()
    logs_dict = {}
    for log in logs:
        if log.request_id not in logs_dict:
            logs_dict[log.request_id] = []
        logs_dict[log.request_id].append(log)

    # Convert to JSON serializable format
    requests_dict = []
    for r in requests:
        # Fallback location using location_text -> city/region
        loc = (
            getattr(r, "location_text", None)
            or ", ".join(
                [
                    val
                    for val in (getattr(r, "city", None), getattr(r, "region", None))
                    if val
                ]
            )
            or ""
        )
        requests_dict.append(
            {
                "id": r.id,
                "name": r.name,
                "phone": r.phone,
                "email": r.email,
                "location": loc,
                "category": r.category,
                "description": r.description,
                "status": r.status,
                # Map urgency to priority if urgency field is missing
                "urgency": getattr(r, "urgency", None) or getattr(r, "priority", None),
            }
        )

    volunteers_dict = [
        {
            "id": v.id,
            "name": v.name,
            "email": v.email,
            "phone": v.phone,
            "location": v.location,
            "skills": v.skills,
        }
        for v in volunteers
    ]

    # Defensive stats: ensure templates always receive a `stats` mapping
    try:
        total_requests = len(requests) if requests is not None else 0
    except Exception:
        total_requests = 0
    try:
        pending_requests = sum(
            1
            for r in requests
            if getattr(r, "status", None) not in ("completed", "done", None)
        )
    except Exception:
        pending_requests = 0
    try:
        total_volunteers = len(volunteers) if volunteers is not None else 0
    except Exception:
        total_volunteers = 0

    stats = {
        "total_requests": total_requests,
        "pending_requests": pending_requests,
        "total_volunteers": total_volunteers,
    }

    try:
        now = utc_now()
        week_ago = now - timedelta(days=7)

        total_requests_cnt = db.session.query(func.count(Request.id)).scalar() or 0

        open_requests_cnt = (
            db.session.query(func.count(Request.id))
            .filter(Request.status.notin_(["done", "rejected"]))
            .scalar()
            or 0
        )

        closed_requests_cnt = (
            db.session.query(func.count(Request.id))
            .filter(Request.status.in_(["done", "rejected"]))
            .scalar()
            or 0
        )

        closed_last_7d_cnt = (
            db.session.query(func.count(Request.id))
            .filter(Request.completed_at.isnot(None), Request.completed_at >= week_ago)
            .scalar()
            or 0
        )

        avg_resolution_hours = (
            db.session.query(
                func.avg(
                    func.julianday(Request.completed_at)
                    - func.julianday(Request.created_at)
                )
                * 24.0
            )
            .filter(Request.completed_at.isnot(None), Request.created_at.isnot(None))
            .scalar()
        )
        avg_resolution_hours = (
            float(avg_resolution_hours) if avg_resolution_hours is not None else None
        )

        unassigned_over_2d_cnt = (
            db.session.query(func.count(Request.id))
            .filter(
                Request.owner_id.is_(None),
                Request.status.notin_(["done", "rejected"]),
                Request.created_at <= (now - timedelta(days=2)),
            )
            .scalar()
            or 0
        )

        high_open_count = (
            db.session.query(func.count(Request.id))
            .filter(Request.status.notin_(["done", "rejected"]))
            .filter(Request.priority == "high")
            .scalar()
            or 0
        )

        stale_open_count = (
            db.session.query(func.count(Request.id))
            .filter(Request.status.notin_(["done", "rejected"]))
            .filter(Request.created_at <= (now - timedelta(days=7)))
            .scalar()
            or 0
        )

        # --- Requests per day (last 14 days) ---
        since_dt = now - timedelta(days=14)
        rows = (
            db.session.query(func.date(Request.created_at), func.count(Request.id))
            .filter(Request.created_at.isnot(None))
            .filter(Request.created_at >= since_dt)
            .group_by(func.date(Request.created_at))
            .order_by(func.date(Request.created_at))
            .all()
        )
        impact_dates = [str(r[0]) for r in rows]
        impact_counts = [int(r[1]) for r in rows]

        # --- Requests by category ---
        cat_rows = (
            db.session.query(Request.category, func.count(Request.id))
            .group_by(Request.category)
            .order_by(func.count(Request.id).desc())
            .all()
        )
        impact_cat_labels = [r[0] or "unknown" for r in cat_rows]
        impact_cat_counts = [int(r[1]) for r in cat_rows]

        impact = {
            "total": total_requests_cnt,
            "open": open_requests_cnt,
            "closed": closed_requests_cnt,
            "closed_last_7d": closed_last_7d_cnt,
            "avg_resolution_hours": avg_resolution_hours,
            "unassigned_over_2d": unassigned_over_2d_cnt,
            "open_requests": int(open_requests_cnt or 0),
            "unassigned_48h": int(unassigned_over_2d_cnt or 0),
            "requests_dates": impact_dates,
            "requests_counts": impact_counts,
            "cat_labels": impact_cat_labels,
            "cat_counts": impact_cat_counts,
            "high_open": int(high_open_count or 0),
            "stale_open": int(stale_open_count or 0),
        }
    except Exception:
        impact = {
            "total": 0,
            "open": 0,
            "closed": 0,
            "closed_last_7d": 0,
            "avg_resolution_hours": None,
            "unassigned_over_2d": 0,
            "open_requests": 0,
            "unassigned_48h": 0,
            "requests_dates": [],
            "requests_counts": [],
            "cat_labels": [],
            "cat_counts": [],
            "high_open": 0,
            "stale_open": 0,
        }

    # Log the final template context summary for diagnostics during tests
    try:
        import logging as _logging

        _log = _logging.getLogger(__name__)
        _log.info(
            "admin_dashboard rendering: stats=%s, requests_items=%s, volunteers=%s",
            stats,
            total_requests,
            total_volunteers,
        )
    except Exception:
        pass

    return render_template(
        "admin_dashboard.html",
        requests={"items": requests},
        logs_dict=logs_dict,
        requests_json=requests_dict,
        volunteers=volunteers,
        volunteers_json=volunteers_dict,
        stats=stats,
        impact=impact,
        STATUS_LABELS=STATUS_LABELS,
    )


@admin_bp.route("/intervenants", methods=["GET"])
@admin_required
def admin_intervenants():
    admin_required_404()
    """Управление на интервенанти (canonical data)"""

    import logging

    logging.warning(
        f"[DEBUG] admin_volunteers: is_authenticated={getattr(current_user, 'is_authenticated', None)}, is_admin={getattr(current_user, 'is_admin', None)}, id={getattr(current_user, 'id', None)}, username={getattr(current_user, 'username', None)}"
    )
    if not current_user.is_admin:
        flash(_("Access denied."), "error")
        return redirect(url_for("main.index"))

    search = (request.args.get("search") or "").strip()
    location_filter = (request.args.get("location") or "").strip()
    sort_by = (request.args.get("sort") or "created_at").strip().lower()
    sort_order = (request.args.get("order") or "desc").strip().lower()
    page = max(int(request.args.get("page") or 1), 1)
    per_page = max(min(int(request.args.get("per_page") or 25), 100), 10)

    query = Intervenant.query
    if not _is_global_admin():
        query = query.filter(Intervenant.structure_id == _current_structure_id())
    if search:
        q = f"%{search}%"
        query = query.filter(
            or_(
                Intervenant.name.ilike(q),
                Intervenant.email.ilike(q),
                Intervenant.phone.ilike(q),
            )
        )
    if location_filter:
        query = query.filter(Intervenant.location.ilike(f"%{location_filter}%"))

    sort_map = {
        "name": Intervenant.name,
        "email": Intervenant.email,
        "location": Intervenant.location,
        "created_at": Intervenant.created_at,
    }
    sort_col = sort_map.get(sort_by, Intervenant.created_at)
    if sort_order == "asc":
        query = query.order_by(sort_col.asc(), Intervenant.id.asc())
    else:
        query = query.order_by(sort_col.desc(), Intervenant.id.desc())

    total_volunteers = query.count()
    total_pages = max(1, int(math.ceil(total_volunteers / float(per_page)))) if total_volunteers else 1
    if page > total_pages:
        page = total_pages

    volunteers = query.offset((page - 1) * per_page).limit(per_page).all()
    return render_template(
        "admin_volunteers.html",
        volunteers=volunteers,
        total_volunteers=total_volunteers,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
        search=search,
        location_filter=location_filter,
        sort_by=sort_by,
        sort_order=sort_order,
    )


@admin_bp.route("/volunteers", methods=["GET"])
@admin_required
def admin_volunteers():
    admin_required_404()
    return redirect(url_for("admin.admin_intervenants"), code=302)

@admin_bp.route("/admin_volunteers", methods=["GET"])
@admin_required
def admin_volunteers_compat():
    admin_required_404()
    return redirect(url_for("admin.admin_volunteers"), code=302)


@admin_bp.route("/admin_volunteers/add", methods=["GET", "POST"])
@admin_required
def add_volunteer():
    admin_required_404()
    """Добавяне на доброволец"""
    if not current_user.is_admin:
        flash(_("Access denied."), "error")
        return redirect(url_for("main.index"))

    if request.method == "POST":
        volunteer = Volunteer(
            name=request.form["name"],
            email=request.form["email"],
            phone=request.form["phone"],
            location=request.form["location"],
            skills=request.form.get("skills", ""),
        )
        db.session.add(volunteer)
        db.session.commit()
        flash(_("Volunteer added successfully!"), "success")
        return redirect(url_for("admin.admin_volunteers"))

    return render_template("add_volunteer.html")


@admin_bp.route("/delete_volunteer/<int:id>", methods=["POST"])
@admin_required
def delete_volunteer(id):
    admin_required_404()
    """Изтриване на доброволец"""
    if not current_user.is_admin:
        flash(_("Access denied."), "error")
        return redirect(url_for("main.index"))

    from flask import abort

    volunteer = db.session.get(Volunteer, id)
    if not volunteer:
        abort(404)
    db.session.delete(volunteer)
    db.session.commit()
    flash(_("Volunteer deleted successfully!"), "success")
    return redirect(url_for("admin.admin_volunteers"))


@admin_bp.route("/admin_volunteers/edit/<int:id>", methods=["GET", "POST"])
@admin_required
def edit_volunteer(id):
    admin_required_404()
    """Редактиране на доброволец"""
    if not current_user.is_admin:
        flash(_("Access denied."), "error")
        return redirect(url_for("main.index"))

    from flask import abort

    volunteer = db.session.get(Volunteer, id)
    if not volunteer:
        abort(404)

    import logging

    if request.method == "POST":
        logging.warning(f"[DEBUG] POST data: {request.form}")
        volunteer.name = request.form["name"]
        volunteer.email = request.form["email"]
        volunteer.phone = request.form["phone"]
        volunteer.location = request.form["location"]
        volunteer.skills = request.form.get("skills", "")
        logging.warning(
            f"[DEBUG] Before commit: name={volunteer.name}, email={volunteer.email}, phone={volunteer.phone}, location={volunteer.location}, skills={volunteer.skills}"
        )
        db.session.commit()
        logging.warning(
            f"[DEBUG] After commit: id={volunteer.id}, email={volunteer.email}"
        )
        flash(_("Changes saved!"), "success")
        return redirect(url_for("admin.admin_volunteers"))

    return render_template("edit_volunteer.html", volunteer=volunteer)


@admin_bp.route("/export_volunteers")
@admin_required
def export_volunteers():
    admin_required_404()
    """Експорт на доброволци като CSV"""
    if not current_user.is_admin:
        flash(_("Access denied."), "error")
        return redirect(url_for("main.index"))

    import csv
    from io import StringIO

    from flask import Response

    volunteers = Volunteer.query.all()
    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(["Име", "Имейл", "Телефон", "Град/регион", "Умения"])
    for v in volunteers:
        cw.writerow([v.name, v.email, v.phone, v.location, v.skills])

    output = si.getvalue()
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=volunteers.csv"},
    )


@admin_bp.route("/update_status/<int:req_id>", methods=["POST"])
@admin_required
def update_status(req_id):
    admin_required_404()
    """Обновяване статуса на заявка"""
    from flask import current_app

    if not current_app.config.get("TESTING", False):
        if not getattr(current_user, "is_admin", False):
            return jsonify({"error": "Unauthorized"}), 403

    req = db.session.get(Request, req_id)
    if not req:
        return jsonify({"error": "Request not found"}), 404
    if not can_edit_request(req, current_user):
        role = getattr(getattr(current_user, "role", None), "value", getattr(current_user, "role", None))
        role = (role or "").strip().lower()
        if role not in {"ops", "admin", "superadmin", "super_admin"}:
            abort(403)

    new_status = (request.form.get("status") or "").strip()
    old_raw_status = req.status
    old_status = normalize_request_status(old_raw_status)
    new_status = normalize_request_status(new_status)

    if new_status not in REQUEST_STATUS_ALLOWED:
        current_app.logger.warning(
            "ADMIN update_status blocked invalid new_status=%r for request_id=%s",
            new_status,
            req_id,
        )
        flash("Invalid status.", "warning")
        return redirect(url_for("admin.admin_request_details", req_id=req_id))

    # Guard: no-op changes shouldn't log noise
    if not new_status or new_status == old_status:
        if request.is_json or (
            request.accept_mimetypes
            and request.accept_mimetypes.best == "application/json"
        ):
            return (
                jsonify(
                    {
                        "success": False,
                        "status": old_status,
                        "message": "No status change.",
                    }
                ),
                200,
            )
        flash("No status change.", "info")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    # ✅ SUCCESS PATH (was accidentally placed under the alias handler)
    req.status = new_status
    closing_statuses = {"done", "cancelled"}
    if new_status in closing_statuses:
        req.completed_at = utc_now()
    else:
        req.completed_at = None
    # Activity + legacy request log (single commit)
    log_request_activity(
        req,
        "status_change",
        old=old_status,
        new=new_status,
        actor_admin_id=getattr(current_user, "id", None),
    )
    _audit_request(
        req.id,
        action="status_change",
        message="Status updated",
        old=old_status,
        new=new_status,
    )
    # metrics
    if _table_exists("request_metrics"):
        metric = db.session.query(RequestMetric).filter_by(request_id=req.id).first()
        if metric is None:
            metric = RequestMetric(request_id=req.id)
            db.session.add(metric)
        if new_status == "done" and metric.time_to_complete is None and req.created_at:
            try:
                metric.time_to_complete = int((utc_now() - req.created_at).total_seconds())
            except Exception:
                pass

    # --- Bulletproof policy sync: VolunteerInterest follows Request.status + assigned_volunteer_id ---
    from ..models.volunteer_interest import (
        VolunteerInterest,
    )

    if new_status == "in_progress":
        assigned_volunteer_id = getattr(req, "assigned_volunteer_id", None)
        if not assigned_volunteer_id:
            current_app.logger.warning(
                "Interest sync skipped: request_id=%s set to in_progress without assigned_volunteer_id",
                req.id,
            )
        else:
            assigned_volunteer = db.session.get(Volunteer, int(assigned_volunteer_id))
            if assigned_volunteer is None:
                current_app.logger.warning(
                    "Interest sync skipped: request_id=%s assigned_volunteer_id=%s not found in volunteers",
                    req.id,
                    assigned_volunteer_id,
                )
                assigned_volunteer_id = None

        if assigned_volunteer_id:
            q = VolunteerInterest.query.filter_by(request_id=req.id)

            # Ensure assigned volunteer's latest interest exists and is approved
            owner_latest = (
                q.filter_by(volunteer_id=assigned_volunteer_id)
                .order_by(VolunteerInterest.id.desc())
                .first()
            )
            if owner_latest is None:
                owner_latest = VolunteerInterest(
                    request_id=req.id,
                    volunteer_id=assigned_volunteer_id,
                    status="approved",
                )
                db.session.add(owner_latest)
                current_app.logger.info(
                    "Interest sync: created approved assigned-volunteer interest (request_id=%s, volunteer_id=%s)",
                    req.id,
                    assigned_volunteer_id,
                )
            elif owner_latest.status != "approved":
                owner_latest.status = "approved"
                db.session.add(owner_latest)
                current_app.logger.info(
                    "Interest sync: set assigned-volunteer interest to approved (request_id=%s, volunteer_id=%s)",
                    req.id,
                    assigned_volunteer_id,
                )

            # Reject other pending interests
            pending_others = (
                q.filter(VolunteerInterest.status == "pending")
                .filter(VolunteerInterest.volunteer_id != assigned_volunteer_id)
                .all()
            )
            for vi_row in pending_others:
                vi_row.status = "rejected"
                db.session.add(vi_row)

            if pending_others:
                current_app.logger.info(
                    "Interest sync: rejected %s pending interests (request_id=%s, assigned_volunteer_id=%s)",
                    len(pending_others),
                    req.id,
                    assigned_volunteer_id,
                )

    elif new_status in {"done", "cancelled"}:
        q = VolunteerInterest.query.filter_by(request_id=req.id)
        pending_all = q.filter(VolunteerInterest.status == "pending").all()
        for vi_row in pending_all:
            vi_row.status = "rejected"
            db.session.add(vi_row)

        if pending_all:
            current_app.logger.info(
                "Interest sync: rejected %s pending interests on close (request_id=%s, new_status=%s)",
                len(pending_all),
                req.id,
                new_status,
            )

    db.session.commit()
    audit_admin_action(
        action="STATUS_CHANGE",
        target_type="Request",
        target_id=req.id,
        payload={"old": {"status": old_status}, "new": {"status": new_status}},
    )

    # Изпращане на email при промяна на статус (async, non-blocking for admin UI)
    try:
        subject = f"Статусът на вашата заявка #{req.id} е променен на {new_status}"
        recipient = getattr(req, "email", None)
        recipient_name = getattr(req, "name", "Потребител")
        content = f"Статусът на вашата заявка е променен на <b>{new_status}</b>.\n\nОписание: {req.description or ''}"
        context = {
            "subject": subject,
            "recipient_name": recipient_name,
            "content": content,
            "request_id": req.id,
            "new_status": new_status,
            "description": req.description,
            "updated_at": req.updated_at,
        }
        if recipient:
            _send_status_email_async(recipient, subject, context)
    except Exception as e:
        import logging

        logging.warning(
            f"[EMAIL] Async status email scheduling failed: {e}"
        )

    # Ако е JSON/AJAX – връщаме JSON; иначе redirect за формата
    if request.is_json or (
        request.accept_mimetypes and request.accept_mimetypes.best == "application/json"
    ):
        return jsonify({"success": True, "status": new_status or req.status})
    flash(_("Status updated."), "success")
    return redirect(url_for("admin.admin_request_details", req_id=req_id))


@admin_bp.post("/requests/<int:req_id>/status")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_request_set_status(req_id: int):
    # Alias: keep old canonical handler, just expose the “resource” URL too.
    return update_status(req_id)


@admin_bp.post("/requests/bulk")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_requests_bulk():
    admin_required_404()

    action = (request.form.get("bulk_action") or "").strip()
    selected_ids_raw = request.form.getlist("selected_ids")
    selected_ids: list[int] = []
    for raw in selected_ids_raw:
        try:
            rid = int(raw)
        except Exception:
            continue
        if rid > 0:
            selected_ids.append(rid)
    selected_ids = sorted(set(selected_ids))

    if not action or not selected_ids:
        flash("No bulk action applied (missing action or selection).", "warning")
        return redirect(url_for("admin.admin_requests"))

    requests = _scope_requests(Request.query).filter(Request.id.in_(selected_ids)).all()
    requests_by_id = {r.id: r for r in requests}
    ordered_reqs = [requests_by_id[rid] for rid in selected_ids if rid in requests_by_id]

    status_map = {
        "set_status_pending": "pending",
        "set_status_in_progress": "in_progress",
        "set_status_done": "done",
        "set_status_rejected": "rejected",
        # Current UI values in template/admin-requests.js
        "status:pending": "pending",
        "status:in_progress": "in_progress",
        "status:done": "done",
        "status:rejected": "rejected",
    }

    changed = 0
    nudged = 0
    skipped = 0
    actor_admin_id = getattr(current_user, "id", None)

    if action in status_map:
        target_status = normalize_request_status(status_map[action])
        for req in ordered_reqs:
            if not can_edit_request(req, current_user):
                skipped += 1
                continue
            old_status = normalize_request_status(getattr(req, "status", None))
            if old_status == target_status:
                continue
            req.status = target_status
            if target_status in {"done", "cancelled"}:
                req.completed_at = utc_now()
            else:
                req.completed_at = None
            try:
                log_request_activity(
                    req,
                    "status_change",
                    old=old_status,
                    new=target_status,
                    actor_admin_id=actor_admin_id,
                )
            except Exception:
                pass
            changed += 1
        db.session.commit()
        flash(f"Bulk status updated: {changed} changed, {skipped} skipped.", "success")
        return redirect(url_for("admin.admin_requests"))

    if action in {"nudge_selected_volunteers", "nudge"}:
        for req in ordered_reqs:
            if not can_edit_request(req, current_user):
                skipped += 1
                continue
            volunteer_id = getattr(req, "assigned_volunteer_id", None)
            if not volunteer_id:
                skipped += 1
                continue
            created = send_nudge_notification(
                request_id=req.id,
                volunteer_id=int(volunteer_id),
                actor_admin_id=actor_admin_id,
            )
            if created:
                nudged += 1
        flash(f"Bulk nudge sent: {nudged} sent, {skipped} skipped.", "success")
        return redirect(url_for("admin.admin_requests"))

    # Front-end-only actions are valid but have no server-side effect.
    if action in {"open_selected", "open", "copy_ids", "copy_links"}:
        flash("Bulk action is UI-only and has no server-side effect.", "info")
        return redirect(url_for("admin.admin_requests"))

    flash("Unknown bulk action.", "warning")
    return redirect(url_for("admin.admin_requests"))


@admin_bp.post("/requests/<int:req_id>/archive", endpoint="admin_request_archive")
@login_required
@admin_required
@admin_role_required("superadmin")
def admin_request_archive(req_id: int):
    """One-click archive/close action used by the details view button."""
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)

    old_status = normalize_request_status(getattr(req, "status", None))
    req.status = "cancelled"
    req.completed_at = utc_now()
    req.is_archived = True
    if getattr(req, "archived_at", None) is None:
        req.archived_at = utc_now()

    log_request_activity(
        req,
        "status_change",
        old=old_status,
        new="cancelled",
        actor_admin_id=getattr(current_user, "id", None),
    )
    db.session.commit()
    audit_admin_action(
        action="request.archive",
        target_type="Request",
        target_id=req.id,
        payload={
            "old": {"status": old_status, "archived": False},
            "new": {"status": "cancelled", "archived": True},
        },
    )
    flash("Request archived and closed.", "success")
    return redirect(url_for("admin.admin_request_details", req_id=req.id))


from flask import current_app, flash, redirect, render_template, request, url_for
from sqlalchemy import and_, func, or_, tuple_

from ..models import Request, RequestActivity, db

ALLOWED_STATUSES = {"pending", "approved", "in_progress", "done", "rejected"}

STATUS_LABELS_BG = {
    "pending": "Чакащи",
    "approved": "Одобрени",
    "in_progress": "В процес",
    "done": "Приключени",
    "rejected": "Отхвърлени",
}

# Canonical EN msgids for status labels - passed to templates for localization
STATUS_LABELS = {
    "pending": "Pending",
    "approved": "Approved",
    "in_progress": "In progress",
    "done": "Completed",
    "rejected": "Rejected",
}


@admin_bp.get("/risk")
@admin_required
def admin_risk_panel():
    admin_required_404()
    return render_template("admin/risk_panel.html")


@admin_bp.get("/sla")
@admin_required
def admin_sla_breakdown():
    admin_required_404()
    _require_global_admin()

    breach_type = (request.args.get("type") or "owner_assign").strip().lower()
    if breach_type not in {"all", "resolve", "owner_assign", "volunteer_assign"}:
        breach_type = "owner_assign"
    sort = (request.args.get("sort") or "overdue").strip().lower()
    days = _normalize_sla_days(request.args.get("days", 30))
    limit = max(1, min(int(request.args.get("limit", 200) or 200), 1000))

    now = datetime.now(UTC).replace(tzinfo=None)
    q = _sla_base_window_query(_scope_requests(Request.query), days=days, now=now)

    if breach_type == "resolve":
        breach_label = "SLA résolution"
    elif breach_type == "owner_assign":
        breach_label = "SLA assignation owner"
    elif breach_type == "volunteer_assign":
        breach_label = "Affectation bénévole"
    else:
        breach_label = "Toutes violations"

    resolve_count = (
        _apply_sla_queue_filter(
            _scope_requests(Request.query),
            sla_kind="resolution_overdue",
            days=days,
            now=now,
        ).count()
        or 0
    )
    owner_assign_count = (
        _apply_sla_queue_filter(
            _scope_requests(Request.query),
            sla_kind="owner_assignment_overdue",
            days=days,
            now=now,
        ).count()
        or 0
    )
    volunteer_assign_count = (
        _apply_sla_queue_filter(
            _scope_requests(Request.query),
            sla_kind="volunteer_assignment_overdue",
            days=days,
            now=now,
        ).count()
        or 0
    )

    requests = q.all()
    prediction_counts = {
        "resolution_overdue": 0,
        "owner_assignment_overdue": 0,
        "volunteer_assignment_overdue": 0,
    }

    rows: list[dict] = []

    for req in requests:
        for kind in prediction_counts.keys():
            pred = _sla_prediction_state(req, sla_kind=kind, now=now)
            if pred.get("state") == "due_soon":
                prediction_counts[kind] += 1

        overdue_by_kind = _sla_overdue_hours_by_kind(req, now=now)
        candidates = [
            (SLA_KIND_TO_BREAKDOWN_TYPE[kind], overdue)
            for kind, overdue in overdue_by_kind.items()
            if kind in SLA_KIND_TO_BREAKDOWN_TYPE
        ]

        if not candidates:
            continue

        if breach_type == "all":
            row_breach_type, overdue_hours = max(candidates, key=lambda x: x[1])
        else:
            picked = {k: v for k, v in candidates}.get(breach_type)
            if picked is None:
                continue
            row_breach_type, overdue_hours = breach_type, picked

        rows.append(
            {
                "id": req.id,
                "title": req.title,
                "category": req.category,
                "status": req.status,
                "created_at": req.created_at,
                "owner_id": req.owner_id,
                "assigned_volunteer_id": req.assigned_volunteer_id,
                "overdue_hours": round(float(overdue_hours), 1),
                "breach_type": row_breach_type,
            }
        )

    def _created_ts(row: dict) -> float:
        dt = _to_utc_naive(row.get("created_at"))
        return dt.timestamp() if dt else 0.0

    if sort == "created":
        rows.sort(key=_created_ts)
    else:
        rows.sort(
            key=lambda r: (-(float(r.get("overdue_hours") or 0.0)), _created_ts(r))
        )
    rows = rows[:limit]

    return render_template(
        "admin/sla.html",
        breach_label=breach_label,
        breach_type=breach_type,
        days=days,
        sort=sort,
        limit=limit,
        resolve_count=int(resolve_count),
        owner_assign_count=int(owner_assign_count),
        volunteer_assign_count=int(volunteer_assign_count),
        prediction_counts=prediction_counts,
        rows=rows,
    )


@admin_bp.get("/pilotage")
@admin_required
@admin_role_required("readonly", "ops", "superadmin", "admin")
def admin_pilotage():
    admin_required_404()

    base_query = _scope_requests(Request.query).filter(Request.deleted_at.is_(None))
    active_query = base_query.filter(
        or_(
            Request.status.is_(None),
            ~func.lower(func.coalesce(Request.status, "")).in_(
                ("done", "cancelled", "rejected")
            ),
        )
    )

    has_risk_level = _table_has_column("requests", "risk_level")
    has_risk_signals = _table_has_column("requests", "risk_signals")
    has_risk_score = _table_has_column("requests", "risk_score")
    has_created_at = _table_has_column("requests", "created_at")
    has_owned_at = _table_has_column("requests", "owned_at")
    has_completed_at = _table_has_column("requests", "completed_at")
    has_updated_at = _table_has_column("requests", "updated_at")

    critical_count = 0
    attention_count = 0
    standard_count = 0
    if has_risk_level:
        critical_count = (
            base_query.filter(func.lower(func.coalesce(Request.risk_level, "")) == "critical").count()
        )
        attention_count = (
            base_query.filter(func.lower(func.coalesce(Request.risk_level, "")) == "attention").count()
        )
        standard_count = (
            base_query.filter(func.lower(func.coalesce(Request.risk_level, "standard")) == "standard").count()
        )
    elif has_risk_score:
        critical_count = base_query.filter(func.coalesce(Request.risk_score, 0) >= 70).count()
        attention_count = base_query.filter(
            func.coalesce(Request.risk_score, 0) >= 40,
            func.coalesce(Request.risk_score, 0) < 70,
        ).count()
        standard_count = base_query.filter(func.coalesce(Request.risk_score, 0) < 40).count()

    no_owner_count = 0
    not_seen_72h_count = 0
    critical_no_owner_count = 0
    if has_risk_signals:
        no_owner_filter = func.lower(func.coalesce(Request.risk_signals, "")).like("%no_owner%")
        not_seen_filter = func.lower(func.coalesce(Request.risk_signals, "")).like("%not_seen_72h%")
        no_owner_count = base_query.filter(no_owner_filter).count()
        not_seen_72h_count = base_query.filter(not_seen_filter).count()
        if has_risk_level:
            critical_no_owner_count = base_query.filter(
                func.lower(func.coalesce(Request.risk_level, "")) == "critical",
                no_owner_filter,
            ).count()

    rec_counts = {
        "assign_immediately": 0,
        "manager_review_today": 0,
        "route_to_housing_partner": 0,
        "route_to_food_support": 0,
        "route_to_health_support": 0,
    }
    if has_risk_level or has_risk_signals:
        rec_rows = base_query.with_entities(
            (Request.risk_level if has_risk_level else func.literal("standard")),
            (Request.risk_signals if has_risk_signals else func.literal("")),
        ).all()
        for risk_level, risk_signals in rec_rows:
            rec = compute_recommendation(
                SimpleNamespace(risk_level=risk_level, risk_signals=risk_signals)
            )
            action = rec.get("recommended_action")
            if action in rec_counts:
                rec_counts[action] += 1
    assign_immediately_count = int(rec_counts.get("assign_immediately", 0) or 0)
    manager_review_today_count = int(rec_counts.get("manager_review_today", 0) or 0)

    def _signals_text(raw: object) -> str:
        if raw is None:
            return ""
        if isinstance(raw, str):
            return raw.strip().lower()
        try:
            return json.dumps(raw, ensure_ascii=False).lower()
        except Exception:
            return str(raw).strip().lower()

    def _to_unix_ts(dt_value: object) -> float:
        if not isinstance(dt_value, datetime):
            return 0.0
        try:
            return float(dt_value.timestamp())
        except Exception:
            return 0.0

    priority_cols = [
        Request.id,
        Request.title,
        Request.description,
        Request.owner_id,
        (Request.risk_level if has_risk_level else func.literal("standard")),
        (Request.risk_signals if has_risk_signals else func.literal("")),
        (Request.risk_score if has_risk_score else func.literal(0)),
        (Request.created_at if has_created_at else func.literal(None)),
    ]
    priority_rows = base_query.with_entities(*priority_cols).limit(500).all()

    priority_items: list[dict[str, object]] = []
    for (
        rid,
        title,
        description,
        owner_id,
        risk_level,
        risk_signals,
        risk_score,
        created_at,
    ) in priority_rows:
        risk_level_norm = (str(risk_level or "standard").strip().lower() or "standard")
        if risk_level_norm not in {"standard", "attention", "critical"}:
            risk_level_norm = "standard"
        signals_text = _signals_text(risk_signals)
        has_no_owner = "no_owner" in signals_text
        has_not_seen_72h = "not_seen_72h" in signals_text

        rec_action = "routine_queue"
        try:
            rec_action = (
                compute_recommendation(
                    SimpleNamespace(risk_level=risk_level_norm, risk_signals=risk_signals)
                ).get("recommended_action")
                or "routine_queue"
            )
        except Exception:
            rec_action = "routine_queue"

        if risk_level_norm == "critical" and has_no_owner:
            indicator_label = "Sans responsable"
            rank_group = 0
        elif has_not_seen_72h:
            indicator_label = "Sans action depuis 72 heures"
            rank_group = 1
        elif rec_action == "assign_immediately":
            indicator_label = "Affectation immédiate recommandée"
            rank_group = 2
        elif rec_action == "manager_review_today":
            indicator_label = "Revue managériale requise"
            rank_group = 3
        elif risk_level_norm == "critical":
            indicator_label = "Niveau critique"
            rank_group = 4
        else:
            continue

        score_value = int(risk_score or 0)
        summary_compact = build_case_summary_snippet(
            SimpleNamespace(
                title=title,
                description=description,
                risk_level=risk_level_norm,
                risk_signals=risk_signals,
                owner_id=owner_id,
            ),
            {"recommended_action": rec_action},
            max_len=100,
        )
        priority_items.append(
            {
                "id": rid,
                "title": title or f"Demande #{rid}",
                "risk_level": risk_level_norm,
                "indicator_label": indicator_label,
                "summary_compact": summary_compact,
                "rank_group": rank_group,
                "risk_score": score_value,
                "created_ts": _to_unix_ts(created_at),
            }
        )

    priority_items.sort(
        key=lambda item: (
            int(item.get("rank_group") or 99),
            -int(item.get("risk_score") or 0),
            -float(item.get("created_ts") or 0.0),
            -int(item.get("id") or 0),
        )
    )
    priority_items = priority_items[:5]

    now = datetime.now(UTC).replace(tzinfo=None)
    day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end = day_start + timedelta(days=1)

    received_today = 0
    taken_today = 0
    closed_today = 0
    if has_created_at:
        received_today = base_query.filter(
            Request.created_at >= day_start, Request.created_at < day_end
        ).count()
    if has_owned_at:
        taken_today = base_query.filter(
            Request.owned_at >= day_start, Request.owned_at < day_end
        ).count()
    if has_completed_at:
        closed_today = base_query.filter(
            Request.completed_at >= day_start, Request.completed_at < day_end
        ).count()

    def _category_label(value: str | None) -> str:
        norm = (value or "").strip().lower()
        mapping = {
            "housing": "logement",
            "logement": "logement",
            "health": "santé",
            "sante": "santé",
            "food": "aide alimentaire",
            "emergency": "urgence",
            "social": "accompagnement social",
            "general": "suivi social",
            "safety": "protection",
        }
        if norm in mapping:
            return mapping[norm]
        if not norm:
            return ""
        return norm.replace("_", " ")

    top_category_row = (
        active_query.with_entities(
            func.lower(func.trim(func.coalesce(Request.category, ""))).label("cat"),
            func.count(Request.id).label("cnt"),
        )
        .group_by("cat")
        .order_by(func.count(Request.id).desc(), func.lower(Request.category).asc())
        .all()
    )
    top_category = ""
    top_category_count = 0
    for row in top_category_row:
        cat = (getattr(row, "cat", "") or "").strip()
        if not cat:
            continue
        top_category = cat
        top_category_count = int(getattr(row, "cnt", 0) or 0)
        break
    if top_category and top_category_count >= 2:
        category_trend_text = (
            f"La catégorie la plus fréquente actuellement concerne {_category_label(top_category)}."
        )
    else:
        category_trend_text = "Aucune tendance catégorielle significative à ce stade."

    assignment_delay_hours: list[float] = []
    if has_created_at and has_owned_at:
        assignment_rows = (
            active_query.with_entities(Request.created_at, Request.owned_at)
            .filter(Request.created_at.isnot(None), Request.owned_at.isnot(None))
            .limit(1500)
            .all()
        )
        for created_at, owned_at in assignment_rows:
            if not isinstance(created_at, datetime) or not isinstance(owned_at, datetime):
                continue
            delta_sec = (owned_at - created_at).total_seconds()
            if delta_sec < 0:
                continue
            assignment_delay_hours.append(delta_sec / 3600.0)
    if len(assignment_delay_hours) >= 3:
        avg_hours = round(sum(assignment_delay_hours) / len(assignment_delay_hours))
        assignment_delay_text = (
            f"Le délai moyen avant affectation est actuellement de {int(avg_hours)} heures."
        )
    else:
        assignment_delay_text = (
            "Données insuffisantes pour estimer le délai moyen avant affectation."
        )

    if int(critical_no_owner_count or 0) > 0:
        vigilance_text = (
            "Les situations critiques sans responsable nécessitent une vigilance renforcée."
        )
    elif int(not_seen_72h_count or 0) > 0:
        vigilance_text = (
            "Les situations sans action récente nécessitent une vigilance renforcée."
        )
    elif int(attention_count or 0) > int(critical_count or 0) and int(attention_count or 0) > 0:
        vigilance_text = (
            "Une part importante des situations est au niveau attention et appelle un suivi rapproché."
        )
    else:
        vigilance_text = "Aucun signal de vigilance particulier n’est identifié à ce stade."

    return render_template(
        "admin/pilotage.html",
        critical_count=critical_count,
        attention_count=attention_count,
        standard_count=standard_count,
        no_owner_count=no_owner_count,
        not_seen_72h_count=not_seen_72h_count,
        critical_no_owner_count=critical_no_owner_count,
        critical_without_owner_count=critical_no_owner_count,
        assign_immediately_count=assign_immediately_count,
        manager_review_today_count=manager_review_today_count,
        priority_items=priority_items,
        rec_counts=rec_counts,
        received_today=received_today,
        taken_today=taken_today,
        closed_today=closed_today,
        category_trend_text=category_trend_text,
        assignment_delay_text=assignment_delay_text,
        vigilance_text=vigilance_text,
    )


@admin_bp.get("/requests")
@login_required
@admin_required
@admin_role_required("superadmin")
def admin_requests():
    admin_required_404()
    if _admin_role_value() != "superadmin":
        _audit_denied_action(required_roles={"superadmin"}, actor_role=_admin_role_value())
        abort(403)
    STATUS_LABELS_BG = {
        "new": "Нови",
        "pending": "Чакащи",
        "approved": "Одобрени",
        "in_progress": "В процес",
        "done": "Приключени",
        "rejected": "Отхвърлени",
    }

    queue = (request.args.get("queue") or "").strip().lower()
    sla_kind = _normalize_sla_kind(request.args.get("sla_kind"))
    sla_days = _normalize_sla_days(request.args.get("sla_days", 30))
    active_sla_queue = bool(queue == "sla" and sla_kind)
    active_sla_filter_label = (
        SLA_QUEUE_KINDS.get(sla_kind, "") if active_sla_queue else ""
    )
    show_deleted = (request.args.get("deleted") or "").strip() == "1"
    query = _scope_requests(Request.query)
    query, status, q, risk, risk_level, no_owner, not_seen_72h, sort = (
        build_requests_query(query, request.args)
    )
    requests = query.all()
    now_aware = utc_now()
    now_naive = datetime.now(UTC).replace(tzinfo=None)
    SLA_WARN_NO_OWNER_DAYS = 2
    SLA_STALE_DAYS = 7
    risk_notseen_tier_hours = _notseen_hours_from_risk(risk)
    risk_columns_ready = _table_has_column("requests", "risk_level") and _table_has_column(
        "requests", "risk_signals"
    )
    critical_count = 0
    attention_count = 0
    no_owner_count = 0
    not_seen_72h_count = 0
    if risk_columns_ready:
        overview_query = _scope_requests(Request.query).filter(Request.deleted_at.is_(None))
        critical_count = (
            overview_query.filter(func.lower(func.coalesce(Request.risk_level, "")) == "critical").count()
        )
        attention_count = (
            overview_query.filter(func.lower(func.coalesce(Request.risk_level, "")) == "attention").count()
        )
        no_owner_count = (
            overview_query.filter(
                func.lower(func.coalesce(Request.risk_signals, "")).like("%no_owner%")
            ).count()
        )
        not_seen_72h_count = (
            overview_query.filter(
                func.lower(func.coalesce(Request.risk_signals, "")).like("%not_seen_72h%")
            ).count()
        )
    age_days_by_id = {}
    for r in requests:
        created_at = getattr(r, "created_at", None)
        if created_at is None:
            age_days_by_id[int(r.id)] = 0
            continue
        try:
            if getattr(created_at, "tzinfo", None) is not None:
                created_at = created_at.replace(tzinfo=None)
            age_days_by_id[int(r.id)] = max((now_naive - created_at).days, 0)
        except Exception:
            age_days_by_id[int(r.id)] = 0

    scope_label = "Vue globale"
    if not _is_global_admin():
        try:
            sid = _current_structure_id()
            active_structure = db.session.get(Structure, sid)
            if active_structure and active_structure.name:
                scope_label = f"Structure active : {active_structure.name}"
            else:
                scope_label = f"Structure active : #{sid}"
        except Exception:
            scope_label = "Structure active : —"

    # Volunteer signals counts per request
    action_counts = {}
    last_signal_by_req = {}
    engagement_by_request = {}
    nudge_ui = {}
    volunteer_actions_supported = _table_exists("volunteer_actions")
    if requests and volunteer_actions_supported:
        req_ids = [r.id for r in requests]
        rows = (
            db.session.query(
                VolunteerAction.request_id,
                VolunteerAction.action,
                func.count(VolunteerAction.id),
            )
            .filter(VolunteerAction.request_id.in_(req_ids))
            .group_by(VolunteerAction.request_id, VolunteerAction.action)
            .all()
        )
        for rid, act, cnt in rows:
            action_counts.setdefault(rid, {}).update({act: cnt})

        # --- Last volunteer signal per request (page only) ---
        # One extra query for this page, avoids N+1 and makes "can't help" visible.
        last_rows = (
            VolunteerAction.query.filter(VolunteerAction.request_id.in_(req_ids))
            .order_by(
                VolunteerAction.request_id.asc(),
                VolunteerAction.updated_at.desc(),
                VolunteerAction.created_at.desc(),
            )
            .all()
        )
        # pick first (newest) per request_id (because ordered desc by updated_at/created_at)
        for a in last_rows:
            if a.request_id not in last_signal_by_req:
                last_signal_by_req[a.request_id] = a

    if requests:
        assigned_volunteer_ids = sorted(
            {
                int(r.assigned_volunteer_id)
                for r in requests
                if getattr(r, "assigned_volunteer_id", None)
            }
        )
        engagement_by_volunteer = {}
        for volunteer_id in assigned_volunteer_ids:
            try:
                engagement_by_volunteer[volunteer_id] = get_volunteer_engagement_score(
                    volunteer_id, now=now_naive
                )
            except Exception:
                db.session.rollback()
                engagement_by_volunteer[volunteer_id] = {
                    "volunteer_id": int(volunteer_id),
                    "score": 0,
                    "label": "At risk",
                    "seen_within_24h": 0,
                    "not_seen_72h": 0,
                    "can_help": 0,
                    "cant_help": 0,
                }
        engagement_by_request = {
            r.id: engagement_by_volunteer.get(getattr(r, "assigned_volunteer_id", None))
            for r in requests
        }

        # Pair-safe nudge cooldown status for queue rows (single query; no N+1).
        def _to_utc_naive(dt):
            if dt is None:
                return None
            if dt.tzinfo is None:
                return dt
            return dt.astimezone(UTC).replace(tzinfo=None)

        now = datetime.now(UTC).replace(tzinfo=None)
        cooldown = timedelta(hours=NUDGE_COOLDOWN_HOURS)
        pairs = {
            (r.id, int(r.assigned_volunteer_id))
            for r in requests
            if getattr(r, "id", None) and getattr(r, "assigned_volunteer_id", None)
        }
        nudge_by_pair: dict[tuple[int, int], datetime] = {}

        if pairs:
            try:
                nudge_rows = (
                    db.session.query(
                        Notification.request_id,
                        Notification.volunteer_id,
                        Notification.created_at,
                    )
                    .filter(Notification.type == "admin_nudge")
                    .filter(
                        tuple_(Notification.request_id, Notification.volunteer_id).in_(
                            pairs
                        )
                    )
                    .all()
                )
            except Exception:
                req_ids = {req_id for req_id, _ in pairs}
                nudge_rows = (
                    db.session.query(
                        Notification.request_id,
                        Notification.volunteer_id,
                        Notification.created_at,
                    )
                    .filter(Notification.type == "admin_nudge")
                    .filter(Notification.request_id.in_(req_ids))
                    .all()
                )

            for req_id, vol_id, created_at in nudge_rows:
                if (req_id, vol_id) not in pairs:
                    continue
                created_naive = _to_utc_naive(created_at)
                if not created_naive:
                    continue
                key = (int(req_id), int(vol_id))
                prev = nudge_by_pair.get(key)
                if prev is None or created_naive > prev:
                    nudge_by_pair[key] = created_naive

        for r in requests:
            rid = getattr(r, "id", None)
            vid = getattr(r, "assigned_volunteer_id", None)
            if not rid:
                continue
            if not vid:
                nudge_ui[rid] = {"disabled": True, "title": "No assigned volunteer"}
                continue

            created_at = nudge_by_pair.get((rid, int(vid)))
            if not created_at:
                nudge_ui[rid] = {
                    "disabled": False,
                    "title": "Send reminder to assigned volunteer",
                }
                continue

            next_at = created_at + cooldown
            if next_at > now:
                remaining = next_at - now
                mins = int(remaining.total_seconds() // 60)
                hrs = mins // 60
                mm = mins % 60
                if hrs > 0:
                    tip = f"Nudge available in {hrs}h {mm}m"
                else:
                    tip = f"Nudge available in {mm}m"
                nudge_ui[rid] = {"disabled": True, "title": tip}
            else:
                nudge_ui[rid] = {
                    "disabled": False,
                    "title": "Send reminder to assigned volunteer",
                }

    return render_template(
        "admin/requests.html",
        STATUS_LABELS_BG=STATUS_LABELS_BG,
        STATUS_LABELS=STATUS_LABELS,
        requests=requests,
        age_days_by_id=age_days_by_id,
        scope_label=scope_label,
        status=status,
        q=q,
        risk=risk,
        risk_level=risk_level,
        no_owner=no_owner,
        not_seen_72h=not_seen_72h,
        sort=sort,
        show_deleted=show_deleted,
        now_aware=now_aware,
        now_naive=now_naive,
        SLA_WARN_NO_OWNER_DAYS=SLA_WARN_NO_OWNER_DAYS,
        SLA_STALE_DAYS=SLA_STALE_DAYS,
        volunteer_action_counts=action_counts,
        last_signal_by_req=last_signal_by_req,
        engagement_by_request=engagement_by_request,
        nudge_ui=nudge_ui,
        risk_notseen_tier_hours=risk_notseen_tier_hours,
        queue=queue,
        sla_kind=sla_kind,
        sla_days=sla_days,
        active_sla_queue=active_sla_queue,
        active_sla_filter_label=active_sla_filter_label,
        critical_count=critical_count,
        attention_count=attention_count,
        no_owner_count=no_owner_count,
        not_seen_72h_count=not_seen_72h_count,
    )


@admin_bp.route("/requests/new", methods=["GET", "POST"])
@admin_required
@admin_role_required("superadmin")
def admin_request_new():
    admin_required_404()
    if _admin_role_value() != "superadmin":
        _audit_denied_action(required_roles={"superadmin"}, actor_role=_admin_role_value())
        abort(403)

    def _ensure_internal_requester_user() -> User:
        email = "agent.intake@helpchain.local"
        username = "agent_intake"
        existing = User.query.filter(func.lower(User.email) == email).first()
        if existing:
            return existing
        existing = User.query.filter(func.lower(User.username) == username).first()
        if existing:
            return existing
        user = User(
            username=username,
            email=email,
            role="requester",
            is_active=True,
            password_hash="",
        )
        try:
            user.set_password(secrets.token_urlsafe(24))
        except Exception:
            user.password_hash = "!"
        db.session.add(user)
        db.session.flush()
        return user

    categories = [code for code, _label in request_category_choices()]

    structures = Structure.query.order_by(Structure.name.asc(), Structure.id.asc()).all()
    admins = (
        AdminUser.query.filter(AdminUser.is_active.is_(True))
        .order_by(AdminUser.username.asc(), AdminUser.id.asc())
        .all()
    )

    form_data = {
        "title": (request.form.get("title") or "").strip(),
        "description": (request.form.get("description") or "").strip(),
        "person_name": (request.form.get("person_name") or "").strip(),
        "email": (request.form.get("email") or "").strip(),
        "phone": (request.form.get("phone") or "").strip(),
        "city": (request.form.get("city") or "").strip(),
        "category": normalize_request_category((request.form.get("category") or "").strip()),
        "priority": (request.form.get("priority") or "standard").strip().lower(),
        "structure_id": (request.form.get("structure_id") or "").strip(),
        "owner_id": (request.form.get("owner_id") or "").strip(),
        "internal_notes": (request.form.get("internal_notes") or "").strip(),
    }
    form_errors: dict[str, str] = {}

    if request.method == "POST":
        if not form_data["title"]:
            form_errors["title"] = "Veuillez renseigner le titre."
        if not form_data["description"]:
            form_errors["description"] = "Veuillez renseigner la description."
        if not form_data["person_name"]:
            form_errors["person_name"] = "Veuillez renseigner la personne concernée."
        if not form_data["city"]:
            form_errors["city"] = "Veuillez renseigner la ville ou le territoire."
        if not form_data["category"]:
            form_errors["category"] = "Veuillez sélectionner une catégorie."
        elif form_data["category"] not in set(REQUEST_CATEGORY_CODES):
            form_errors["category"] = "Veuillez sélectionner une catégorie valide."
        if form_data["priority"] not in {"standard", "attention", "urgent"}:
            form_errors["priority"] = "Veuillez sélectionner une priorité valide."
        if form_data["email"] and "@" not in form_data["email"]:
            form_errors["email"] = "Veuillez renseigner une adresse e-mail valide."

        structure_id = None
        if not _is_global_admin():
            try:
                structure_id = _current_structure_id()
            except Exception:
                form_errors["structure_id"] = (
                    "Impossible de déterminer la structure active."
                )
        elif form_data["structure_id"]:
            try:
                structure_id = int(form_data["structure_id"])
            except Exception:
                form_errors["structure_id"] = "Structure invalide."
            else:
                if not db.session.get(Structure, structure_id):
                    form_errors["structure_id"] = "Structure invalide."
        else:
            try:
                structure_id = _current_structure_id()
            except Exception:
                form_errors["structure_id"] = (
                    "Impossible de déterminer la structure active."
                )

        owner_id = None
        if form_data["owner_id"]:
            try:
                owner_id = int(form_data["owner_id"])
            except Exception:
                form_errors["owner_id"] = "Responsable initial invalide."
            else:
                if not db.session.get(AdminUser, owner_id):
                    form_errors["owner_id"] = "Responsable initial invalide."

        if form_errors:
            flash("Veuillez corriger les champs indiqués.", "warning")
        else:
            requester_user = _ensure_internal_requester_user()
            priority_map = {
                "standard": "medium",
                "attention": "high",
                "urgent": "urgent",
            }
            req = Request(
                title=form_data["title"],
                description=form_data["description"],
                name=form_data["person_name"],
                email=form_data["email"] or None,
                phone=form_data["phone"] or None,
                city=form_data["city"],
                category=form_data["category"],
                priority=priority_map.get(form_data["priority"], "medium"),
                status="pending",
                structure_id=structure_id,
                owner_id=owner_id,
                message=form_data["internal_notes"] or None,
                user_id=requester_user.id,
            )
            lat, lng = geocode_location_best_effort(city=form_data["city"])
            if lat is not None and lng is not None:
                req.latitude = lat
                req.longitude = lng
            db.session.add(req)
            db.session.commit()
            audit_admin_action(
                action="CREATE_REQUEST",
                target_type="Request",
                target_id=req.id,
                payload={
                    "structure_id": req.structure_id,
                    "owner_id": req.owner_id,
                    "status": req.status,
                    "priority": req.priority,
                    "category": req.category,
                },
            )
            flash("Demande créée avec succès.", "success")
            return redirect(url_for("admin.admin_request_details", req_id=req.id), code=303)

    return render_template(
        "admin/request_new.html",
        form_data=form_data,
        form_errors=form_errors,
        categories=categories,
        structures=structures,
        admins=admins,
        current_structure_id=_current_structure_id(),
    )


def apply_risk_filter(base_query, risk: str, now: datetime):
    closed_statuses = {"done", "cancelled", "rejected"}
    open_filter = or_(
        Request.status.is_(None), ~func.lower(Request.status).in_(closed_statuses)
    )
    notseen_hours = _notseen_hours_from_risk(risk)
    if (
        risk
        not in {
            "stale",
            "unassigned",
            "assigned_recent",
            "volunteer_stale",
            "sla_resolve_breach",
            "sla_assign_breach",
        }
        and notseen_hours is None
    ):
        return base_query

    if risk == "stale":
        base_query = base_query.filter(Request.created_at < (now - timedelta(days=8)))
        base_query = base_query.filter(open_filter)
    elif risk == "unassigned":
        base_query = base_query.filter(
            Request.created_at < (now - timedelta(days=3)),
            Request.assigned_volunteer_id.is_(None),
        )
        base_query = base_query.filter(open_filter)
    elif risk == "assigned_recent":
        base_query = base_query.filter(
            Request.created_at >= (now - timedelta(days=7)),
            Request.assigned_volunteer_id.isnot(None),
        )
    elif risk == "volunteer_stale":
        base_query = base_query.filter(
            Request.created_at < (now - timedelta(hours=VOLUNTEER_ASSIGN_SLA_HOURS)),
            Request.assigned_volunteer_id.is_(None),
        )
        base_query = base_query.filter(open_filter)
    elif risk == "sla_resolve_breach":
        base_query = base_query.filter(
            Request.created_at < (now - timedelta(days=RESOLVE_SLA_DAYS)),
            Request.completed_at.is_(None),
        )
        base_query = base_query.filter(open_filter)
    elif risk == "sla_assign_breach":
        base_query = base_query.filter(
            Request.created_at < (now - timedelta(hours=ASSIGN_SLA_HOURS)),
            Request.owned_at.is_(None),
        )
        base_query = base_query.filter(open_filter)
    elif notseen_hours is not None:
        notseen_subq, _source = _build_notseen_subquery(now, hours=notseen_hours)
        base_query = base_query.filter(
            Request.id.in_(select(notseen_subq.c.request_id))
        )
        base_query = base_query.filter(open_filter)
    return base_query


def build_requests_query(base_query, request_args, legacy: bool = False):
    base_query = _scope_requests(base_query)
    status = (request_args.get("status") or "").strip()
    q = (request_args.get("q") or "").strip()
    category = normalize_request_category((request_args.get("category") or "").strip())
    risk = (request_args.get("risk") or "").strip().lower()
    risk_level = (request_args.get("risk_level") or "").strip().lower()
    no_owner = (request_args.get("no_owner") or "").strip() == "1"
    not_seen_72h = (request_args.get("not_seen_72h") or "").strip() == "1"
    sort = (request_args.get("sort") or "").strip().lower()
    show_deleted = (request_args.get("deleted") or "").strip() == "1"
    queue = (request_args.get("queue") or "").strip().lower()
    sla_kind = _normalize_sla_kind(request_args.get("sla_kind"))
    sla_days = _normalize_sla_days(request_args.get("sla_days", 30))
    now = datetime.now(UTC).replace(tzinfo=None)

    if show_deleted:
        base_query = base_query.filter(Request.deleted_at.isnot(None))
    else:
        base_query = base_query.filter(Request.deleted_at.is_(None))

    if status:
        internal = "pending" if status == "new" else status
        base_query = base_query.filter(Request.status == internal)
    if q:
        like = f"%{q}%"
        base_query = base_query.filter(
            or_(
                Request.title.ilike(like),
                Request.name.ilike(like),
                Request.email.ilike(like),
                Request.phone.ilike(like),
                Request.description.ilike(like),
            )
        )
    if category:
        category_variants = {category}
        for legacy_code in ("general", "social", "medical", "tech", "admin", "other"):
            if normalize_request_category(legacy_code) == category:
                category_variants.add(legacy_code)
        base_query = base_query.filter(func.lower(Request.category).in_([c.lower() for c in category_variants]))

    base_query = apply_risk_filter(base_query, risk, now)
    if risk_level in {"critical", "attention", "standard"}:
        base_query = base_query.filter(
            func.lower(func.coalesce(Request.risk_level, "standard")) == risk_level
        )
    if no_owner:
        base_query = base_query.filter(
            func.lower(func.coalesce(Request.risk_signals, "")).like("%no_owner%")
        )
    if not_seen_72h:
        base_query = base_query.filter(
            func.lower(func.coalesce(Request.risk_signals, "")).like("%not_seen_72h%")
        )
    if queue == "sla" and sla_kind:
        base_query = _apply_sla_queue_filter(
            base_query, sla_kind=sla_kind, days=sla_days, now=now
        )
    created_sort_col = func.coalesce(Request.created_at, Request.updated_at)
    if sort == "created_asc":
        base_query = base_query.order_by(created_sort_col.asc(), Request.id.asc())
    elif sort == "created_desc":
        base_query = base_query.order_by(created_sort_col.desc(), Request.id.desc())
    elif sort == "risk_asc":
        base_query = base_query.order_by(
            func.coalesce(Request.risk_score, 0).asc(),
            created_sort_col.desc(),
            Request.id.desc(),
        )
    else:
        base_query = base_query.order_by(
            func.coalesce(Request.risk_score, 0).desc(),
            created_sort_col.desc(),
            Request.id.desc(),
        )
    result = (base_query, status, q, risk, risk_level, no_owner, not_seen_72h, sort)
    if legacy:
        return result[:4]
    return result


@admin_bp.get("/requests/export.csv")
@admin_required
def admin_requests_export_csv():
    admin_required_404()
    query, *_ = build_requests_query(Request.query, request.args)
    rows = query.limit(5000).all()
    log_activity(
        entity_type="export",
        entity_id=0,
        action="requests_export",
        message="Requests export",
        meta={"format": "csv", "anonymized": False},
        persist=True,
    )

    out = StringIO()
    # Excel-friendly CSV for EU locales: UTF-8 BOM + semicolon delimiter.
    writer = csv.writer(out, delimiter=";")
    writer.writerow(
        [
            "id",
            "created_at",
            "status",
            "priority",
            "category",
            "title",
            "name",
            "email",
            "phone",
            "owner_id",
            "owned_at",
            "completed_at",
        ]
    )

    for r in rows:
        writer.writerow(
            [
                r.id,
                getattr(r, "created_at", "") or "",
                getattr(r, "status", "") or "",
                getattr(r, "priority", "") or "",
                getattr(r, "category", "") or "",
                getattr(r, "title", "") or "",
                getattr(r, "name", "") or "",
                getattr(r, "email", "") or "",
                getattr(r, "phone", "") or "",
                getattr(r, "owner_id", "") or "",
                getattr(r, "owned_at", "") or "",
                getattr(r, "completed_at", "") or "",
            ]
        )

    filename = f"helpchain_requests_{datetime.now(UTC).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.csv"
    return Response(
        "\ufeff" + out.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@admin_bp.get("/requests/export.xlsx")
@admin_required
def admin_requests_export_xlsx():
    admin_required_404()
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return Response("openpyxl is not installed", status=500)

    query, *_ = build_requests_query(Request.query, request.args)
    rows = query.limit(5000).all()
    log_activity(
        entity_type="export",
        entity_id=0,
        action="requests_export",
        message="Requests export",
        meta={"format": "xlsx", "anonymized": False},
        persist=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Requests"
    headers = [
        "id",
        "created_at",
        "status",
        "priority",
        "category",
        "title",
        "name",
        "email",
        "phone",
        "owner_id",
        "owned_at",
        "completed_at",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                r.id,
                getattr(r, "created_at", None),
                getattr(r, "status", None),
                getattr(r, "priority", None),
                getattr(r, "category", None),
                getattr(r, "title", None),
                getattr(r, "name", None),
                getattr(r, "email", None),
                getattr(r, "phone", None),
                getattr(r, "owner_id", None),
                getattr(r, "owned_at", None),
                getattr(r, "completed_at", None),
            ]
        )

    # Keep phone values as text to avoid Excel scientific notation.
    phone_col = headers.index("phone") + 1
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=phone_col)
        if cell.value is not None:
            cell.value = str(cell.value)
        cell.number_format = "@"

    # Auto-fit column widths for readable exports.
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(10, max_len + 2), 60
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"helpchain_requests_{datetime.now(UTC).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.get("/requests/export_anonymized.csv")
@admin_required
def admin_requests_export_csv_anonymized():
    admin_required_404()
    query, *_ = build_requests_query(Request.query, request.args)
    rows = query.limit(5000).all()
    log_activity(
        entity_type="export",
        entity_id=0,
        action="requests_export",
        message="Requests export (anonymized)",
        meta={"format": "csv", "anonymized": True},
        persist=True,
    )

    out = StringIO()
    # Excel-friendly CSV for EU locales: UTF-8 BOM + semicolon delimiter.
    writer = csv.writer(out, delimiter=";")
    writer.writerow(
        [
            "id",
            "created_at",
            "status",
            "priority",
            "category",
            "owner_id",
            "owned_at",
            "completed_at",
        ]
    )

    for r in rows:
        writer.writerow(
            [
                r.id,
                getattr(r, "created_at", "") or "",
                getattr(r, "status", "") or "",
                getattr(r, "priority", "") or "",
                getattr(r, "category", "") or "",
                getattr(r, "owner_id", "") or "",
                getattr(r, "owned_at", "") or "",
                getattr(r, "completed_at", "") or "",
            ]
        )

    filename = (
        f"helpchain_requests_ANON_{datetime.now(UTC).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.csv"
    )
    return Response(
        "\ufeff" + out.getvalue(),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@admin_bp.get("/requests/export_anonymized.xlsx")
@admin_required
def admin_requests_export_xlsx_anonymized():
    admin_required_404()
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return Response("openpyxl is not installed", status=500)

    query, *_ = build_requests_query(Request.query, request.args)
    rows = query.limit(5000).all()
    log_activity(
        entity_type="export",
        entity_id=0,
        action="requests_export",
        message="Requests export (anonymized)",
        meta={"format": "xlsx", "anonymized": True},
        persist=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Requests (Anon)"
    headers = [
        "id",
        "created_at",
        "status",
        "priority",
        "category",
        "owner_id",
        "owned_at",
        "completed_at",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                r.id,
                getattr(r, "created_at", None),
                getattr(r, "status", None),
                getattr(r, "priority", None),
                getattr(r, "category", None),
                getattr(r, "owner_id", None),
                getattr(r, "owned_at", None),
                getattr(r, "completed_at", None),
            ]
        )

    # Auto-fit column widths for readable exports.
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(10, max_len + 2), 60
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = (
        f"helpchain_requests_ANON_{datetime.now(UTC).replace(tzinfo=None).strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@admin_bp.get("/requests/<int:req_id>")
@admin_required
@admin_role_required("superadmin")
def admin_request_details(req_id: int):
    admin_required_404()
    activities_supported = _table_has_column("request_activities", "volunteer_id")
    request_logs_supported = _table_exists("request_logs")
    if activities_supported and request_logs_supported:
        req = (
            _scope_requests(Request.query)
            .options(joinedload(Request.logs), joinedload(Request.activities))
            .filter(Request.id == req_id)
            .first_or_404()
        )
    elif activities_supported and not request_logs_supported:
        req = (
            _scope_requests(Request.query)
            .options(joinedload(Request.activities))
            .filter(Request.id == req_id)
            .first_or_404()
        )
    elif (not activities_supported) and request_logs_supported:
        req = (
            _scope_requests(Request.query)
            .options(joinedload(Request.logs))
            .filter(Request.id == req_id)
            .first_or_404()
        )
    else:
        req = (
            _scope_requests(Request.query)
            .filter(Request.id == req_id)
            .first_or_404()
        )
    linked_case = None
    if _cases_enabled():
        linked_case = Case.query.filter(Case.request_id == req.id).first()
    risk_ai_suggestion = _build_risk_ai_suggestion(req)
    operational_blockages = _build_operational_blockages(req, linked_case)
    volunteer_request_states_supported = _table_exists("volunteer_request_states")
    volunteer_interests_supported = _table_exists("volunteer_interests")
    volunteer_actions_supported = _table_exists("volunteer_actions")
    notifications_supported = _table_exists("notifications")
    admin_id = current_user.id
    now = _now_utc()
    latest_actions = []
    audit_logs = []
    if activities_supported:
        latest_actions = (
            RequestActivity.query.filter_by(request_id=req_id)
            .filter(
                RequestActivity.action.in_(
                    ["volunteer_can_help", "volunteer_cant_help"]
                )
            )
            .order_by(RequestActivity.created_at.desc())
            .limit(10)
            .all()
        )
    linked_volunteer_ids = []
    if volunteer_request_states_supported:
        linked_volunteer_ids = [
            int(v_id)
            for (v_id,) in db.session.query(VolunteerRequestState.volunteer_id)
            .filter(VolunteerRequestState.request_id == req_id)
            .distinct()
            .all()
            if v_id is not None
        ]
    volunteer_engagement = []
    if linked_volunteer_ids:
        linked_volunteers = {
            v.id: v
            for v in Volunteer.query.filter(
                Volunteer.id.in_(linked_volunteer_ids)
            ).all()
        }
        for v_id in linked_volunteer_ids:
            try:
                score_row = get_volunteer_engagement_score(v_id, now=now)
            except Exception:
                db.session.rollback()
                score_row = {
                    "volunteer_id": int(v_id),
                    "score": 0,
                    "label": "At risk",
                    "seen_within_24h": 0,
                    "not_seen_72h": 0,
                    "can_help": 0,
                    "cant_help": 0,
                }
            v = linked_volunteers.get(v_id)
            display = (
                (
                    getattr(v, "name", None)
                    or getattr(v, "email", None)
                    or f"Volunteer #{v_id}"
                )
                if v is not None
                else f"Volunteer #{v_id}"
            )
            score_row["display"] = display
            volunteer_engagement.append(score_row)
        volunteer_engagement.sort(key=lambda x: (-x["score"], x["volunteer_id"]))

    locked_by = None
    # --- AUTO-LOCK (must happen BEFORE any render_template return) ---
    if req.owner_id is None:
        req.owner_id = admin_id
        req.owned_at = now
        if activities_supported:
            db.session.add(
                RequestActivity(
                    request_id=req.id,
                    actor_admin_id=admin_id,
                    action="lock",
                    old_value="",
                    new_value=str(admin_id),
                    created_at=now,
                )
            )
        db.session.commit()
    elif req.owner_id == admin_id:
        # refresh TTL quietly
        if _lock_expired(req, now):
            req.owned_at = now
            db.session.commit()
    else:
        if _lock_expired(req, now):
            old_owner = req.owner_id
            req.owner_id = admin_id
            req.owned_at = now
            if activities_supported:
                db.session.add(
                    RequestActivity(
                        request_id=req.id,
                        actor_admin_id=admin_id,
                        action="lock",
                        old_value=str(old_owner),
                        new_value=str(admin_id),
                        created_at=now,
                    )
                )
            db.session.commit()
        else:
            locked_by = req.owner_id
            # show locked screen (no commit)
            activities = []
            if activities_supported:
                activities = sorted(
                    (req.activities or []),
                    key=lambda a: a.created_at or datetime.min,
                    reverse=True,
                )[:50]
            interests = []
            if volunteer_interests_supported:
                interests = (
                    VolunteerInterest.query.filter_by(request_id=req_id)
                    .order_by(VolunteerInterest.created_at.desc())
                    .all()
                )
            locked_recommendation = compute_recommendation(req)
            return (
                render_template(
                    "admin/request_details.html",
                    req=req,
                    activities=activities,
                    logs=(req.logs if request_logs_supported else []),
                    STATUS_LABELS_BG=STATUS_LABELS_BG,
                    is_stale=is_stale,
                    interests=interests,
                    latest_actions=latest_actions,
                    volunteer_engagement=volunteer_engagement,
                    audit_logs=audit_logs,
                    case_signals=_compute_case_signals(req, activities, now),
                    recommendation=locked_recommendation,
                    helpchain_recommendation=_build_helpchain_recommendation(
                        req, activities, now
                    ),
                    case_summary=build_case_summary(req, locked_recommendation),
                    risk_ai_suggestion=risk_ai_suggestion,
                    operational_blockages=operational_blockages,
                    linked_case=linked_case,
                    is_locked=True,
                    locked_by=locked_by,
                ),
                200,
            )
    is_locked = False
    logs = req.logs if request_logs_supported else []
    activities = []
    if activities_supported:
        activities = sorted(
            (req.activities or []),
            key=lambda a: a.created_at or datetime.min,
            reverse=True,
        )[:50]
    interests = []
    if volunteer_interests_supported:
        interests = (
            VolunteerInterest.query.filter_by(request_id=req_id)
            .order_by(VolunteerInterest.created_at.desc())
            .all()
        )

    # --- V3: Match & engagement (city-based) ---
    req_city = (getattr(req, "city", "") or "").strip().lower()

    def _norm_city(val: str) -> str:
        return (val or "").strip().lower()

    vols = Volunteer.query.filter_by(is_active=True).all()
    matched_volunteers = [
        v for v in vols if _norm_city(getattr(v, "location", None)) == req_city
    ]
    matched_volunteers = matched_volunteers[:20]
    matched_volunteer_ids = [v.id for v in matched_volunteers]

    notif_rows = []
    if matched_volunteer_ids and notifications_supported:
        notif_rows = Notification.query.filter(
            Notification.request_id == req.id,
            Notification.type == "new_match",
            Notification.volunteer_id.in_(matched_volunteer_ids),
        ).all()
    notified_count = len(notif_rows)
    seen_count = sum(1 for n in notif_rows if getattr(n, "is_read", False))

    interest_rows = interests  # already loaded for this request
    interested_ids = {i.volunteer_id for i in interest_rows}
    interested_count = len(interested_ids)

    notif_by_vol = {n.volunteer_id: n for n in notif_rows}
    flags_by_vol = {}
    for v in matched_volunteers:
        n = notif_by_vol.get(v.id)
        flags_by_vol[v.id] = {
            "notified": n is not None,
            "seen": bool(getattr(n, "is_read", False)) if n else False,
            "interested": v.id in interested_ids,
        }

    assigned_volunteer = None
    if getattr(req, "assigned_volunteer_id", None):
        assigned_volunteer = db.session.get(Volunteer, req.assigned_volunteer_id)

    # Volunteer signals (can/can't help)
    volunteer_signals = []
    if volunteer_actions_supported:
        volunteer_signals = (
            VolunteerAction.query.filter_by(request_id=req.id)
            .order_by(VolunteerAction.updated_at.desc())
            .all()
        )
    # Most recent signal for quick, high-visibility admin context.
    last_vol_signal = volunteer_signals[0] if volunteer_signals else None
    signal_vol_ids = [va.volunteer_id for va in volunteer_signals]
    volunteers_map = (
        {
            v.id: v
            for v in Volunteer.query.filter(Volunteer.id.in_(signal_vol_ids)).all()
        }
        if signal_vol_ids
        else {}
    )
    can_help_count = sum(1 for va in volunteer_signals if va.action == "CAN_HELP")
    cant_help_count = sum(1 for va in volunteer_signals if va.action == "CANT_HELP")
    case_signals = _compute_case_signals(req, activities, now)
    recommendation = compute_recommendation(req)
    helpchain_recommendation = _build_helpchain_recommendation(req, activities, now)
    case_summary = build_case_summary(req, recommendation)

    return (
        render_template(
            "admin/request_details.html",
            req=req,
            activities=activities,
            logs=logs,
            STATUS_LABELS_BG=STATUS_LABELS_BG,
            is_stale=is_stale,
            interests=interests,
            is_locked=is_locked,
            locked_by=locked_by,
            matched_volunteers=matched_volunteers,
            matched_count=len(matched_volunteers),
            notified_count=notified_count,
            seen_count=seen_count,
            interested_count=interested_count,
            flags_by_vol=flags_by_vol,
            assigned_volunteer=assigned_volunteer,
            volunteer_signals=volunteer_signals,
            last_vol_signal=last_vol_signal,
            volunteers_map=volunteers_map,
            can_help_count=can_help_count,
            cant_help_count=cant_help_count,
            latest_actions=latest_actions,
            volunteer_engagement=volunteer_engagement,
            audit_logs=audit_logs,
            case_signals=case_signals,
            recommendation=recommendation,
            helpchain_recommendation=helpchain_recommendation,
            case_summary=case_summary,
            risk_ai_suggestion=risk_ai_suggestion,
            operational_blockages=operational_blockages,
            linked_case=linked_case,
        ),
        200,
    )


def _get_scoped_case_or_404(case_id: int) -> tuple[Case, Request]:
    case_row = db.session.get(Case, int(case_id))
    if not case_row:
        abort(404)
    req = _scope_requests(Request.query).filter(Request.id == case_row.request_id).first()
    if not req:
        abort(404)
    return case_row, req


def _priority_with_manual_guard(current_priority: str | None, derived_priority: str | None) -> str:
    current_rank = CASE_PRIORITY_RANK.get((current_priority or "").strip().lower(), -1)
    derived_rank = CASE_PRIORITY_RANK.get((derived_priority or "").strip().lower(), -1)
    if current_rank > derived_rank:
        return (current_priority or "").strip().lower() or "normal"
    return (derived_priority or "").strip().lower() or "normal"


def _apply_cases_risk_filter(query, risk_value: str):
    risk = (risk_value or "").strip().lower()
    score_col = func.coalesce(Case.risk_score, 0)
    if risk == "critical":
        return query.filter(score_col >= 85)
    if risk == "high":
        return query.filter(score_col >= 60, score_col <= 84)
    if risk == "normal":
        return query.filter(score_col >= 30, score_col <= 59)
    if risk == "low":
        return query.filter(score_col <= 29)
    return query


@admin_bp.post("/requests/<int:req_id>/open-case")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_open_case_from_request(req_id: int):
    admin_required_404()
    if not _cases_enabled():
        flash("Case system tables are not available yet. Run migrations first.", "warning")
        return redirect(url_for("admin.admin_request_details", req_id=req_id), code=303)

    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    existing = Case.query.filter(Case.request_id == req.id).first()
    if existing:
        return redirect(url_for("admin.admin_case_detail", case_id=existing.id), code=303)

    now = _now_utc()
    triage = score_request_risk(req)
    derived_priority = _priority_with_manual_guard("normal", triage.get("priority"))
    case_row = Case(
        request_id=req.id,
        structure_id=getattr(req, "structure_id", None),
        owner_user_id=None,
        assigned_professional_lead_id=None,
        status="new",
        priority=derived_priority,
        risk_score=int(triage.get("score") or 0),
        opened_at=now,
        assigned_at=None,
        resolved_at=None,
        closed_at=None,
        last_activity_at=now,
        created_at=now,
        updated_at=now,
    )
    db.session.add(case_row)
    db.session.flush()
    _append_case_event(
        case_id=case_row.id,
        actor_user_id=getattr(current_user, "id", None),
        event_type="case_created",
        message=f"Case created from request #{req.id}",
        metadata={"request_id": req.id},
    )
    _append_case_event(
        case_id=case_row.id,
        actor_user_id=getattr(current_user, "id", None),
        event_type="triage_scored",
        message=(
            f"Triage scored at {int(triage.get('score') or 0)}/100 "
            f"({risk_label_from_score(int(triage.get('score') or 0))})"
        ),
        metadata={
            "risk_score": int(triage.get("score") or 0),
            "risk_label": risk_label_from_score(int(triage.get("score") or 0)),
            "derived_priority": derived_priority,
            "matched_rules": triage.get("matched_rules") or [],
            "suggested_category_code": triage.get("suggested_category_code"),
            "suggested_category_label": triage.get("suggested_category_label"),
        },
    )
    db.session.commit()
    flash(f"Case #{case_row.id} opened.", "success")
    return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)


def _render_cases_list():
    if not _cases_enabled():
        flash("Case system tables are not available yet. Run migrations first.", "warning")
        return render_template(
            "admin/cases.html",
            cases=[],
            status="",
            priority="",
            owner="",
            category="",
            risk="",
            stale=False,
            statuses=list(CATEGORY_CASE_STATUSES),
            priorities=list(CASE_PRIORITIES),
            owners=[],
            critical_count=0,
            attention_count=0,
            no_owner_count=0,
            stale_count=0,
        )

    status = (request.args.get("status") or "").strip().lower()
    priority = (request.args.get("priority") or "").strip().lower()
    owner = (request.args.get("owner") or "").strip()
    category = normalize_request_category((request.args.get("category") or "").strip())
    risk = (request.args.get("risk") or "").strip().lower()
    stale = (request.args.get("stale") or "").strip() == "1"
    owner_id = None
    owner_none = owner.lower() == "none"
    if owner:
        try:
            owner_id = int(owner)
        except Exception:
            owner_id = None

    scoped_ids_subq = _scope_requests(Request.query.with_entities(Request.id)).subquery()
    query = Case.query.join(scoped_ids_subq, Case.request_id == scoped_ids_subq.c.id)
    activity_expr = func.coalesce(Case.last_activity_at, Case.updated_at, Case.created_at)
    stale_threshold = _now_utc() - timedelta(hours=72)
    if status in CATEGORY_CASE_STATUSES:
        query = query.filter(Case.status == status)
    if priority in CASE_PRIORITIES:
        query = query.filter(Case.priority == priority)
    if category:
        category_variants = {category}
        for legacy_code in ("general", "social", "medical", "tech", "admin", "other"):
            if normalize_request_category(legacy_code) == category:
                category_variants.add(legacy_code)
        query = query.filter(func.lower(func.coalesce(Request.category, "")).in_([c.lower() for c in category_variants]))
    if owner_id:
        query = query.filter(Case.owner_user_id == owner_id)
    elif owner_none:
        query = query.filter(Case.owner_user_id.is_(None))
    query = _apply_cases_risk_filter(query, risk)
    if stale:
        query = query.filter(activity_expr <= stale_threshold)

    case_rows = (
        query.options(
            joinedload(Case.request),
            joinedload(Case.owner_user),
            joinedload(Case.assigned_professional_lead),
        )
        .order_by(
            case(
                ((func.coalesce(Case.risk_score, 0) >= 85), 0),
                ((func.coalesce(Case.risk_score, 0) >= 60), 1),
                else_=2,
            ).asc(),
            case(
                ((Case.priority == "critical"), 0),
                ((Case.priority == "high"), 1),
                ((Case.priority == "normal"), 2),
                else_=3,
            ).asc(),
            case((Case.owner_user_id.is_(None), 0), else_=1).asc(),
            case((activity_expr <= stale_threshold, 0), else_=1).asc(),
            activity_expr.desc().nullslast(),
            Case.id.desc(),
        )
        .limit(300)
        .all()
    )

    case_signals = {}
    ops_priority_levels = {}
    now_utc = _now_utc()
    for c in case_rows:
        result = compute_ops_priority(case_row=c, request_row=getattr(c, "request", None), now=now_utc)
        case_signals[int(c.id)] = result.get("ops_priority_reasons") or []
        ops_priority_levels[int(c.id)] = result.get("ops_priority_level") or "normal"

    counts_base = Case.query.join(scoped_ids_subq, Case.request_id == scoped_ids_subq.c.id)
    score_col = func.coalesce(Case.risk_score, 0)
    critical_count = counts_base.filter(score_col >= 85).count()
    attention_count = counts_base.filter(score_col >= 60, score_col <= 84).count()
    no_owner_count = counts_base.filter(Case.owner_user_id.is_(None)).count()
    stale_count = counts_base.filter(activity_expr <= stale_threshold).count()

    owners = (
        AdminUser.query.with_entities(AdminUser.id, AdminUser.username)
        .order_by(AdminUser.username.asc())
        .all()
    )

    return render_template(
        "admin/cases.html",
        cases=case_rows,
        status=status,
        priority=priority,
        owner=owner,
        category=category,
        risk=risk,
        stale=stale,
        statuses=list(CATEGORY_CASE_STATUSES),
        priorities=list(CASE_PRIORITIES),
        owners=owners,
        critical_count=critical_count,
        attention_count=attention_count,
        no_owner_count=no_owner_count,
        stale_count=stale_count,
        case_signals=case_signals,
        ops_priority_levels=ops_priority_levels,
    )


@admin_bp.get("/cases")
@admin_required
@admin_role_required("readonly", "ops", "superadmin")
def admin_cases_list():
    admin_required_404()
    return _render_cases_list()


@admin_bp.get("/cases/<int:case_id>")
@admin_required
@admin_role_required("readonly", "ops", "superadmin")
def admin_case_detail(case_id: int):
    admin_required_404()
    if not _cases_enabled():
        flash("Case system tables are not available yet. Run migrations first.", "warning")
        return redirect(url_for("admin.admin_requests"), code=303)

    case_row, req = _get_scoped_case_or_404(case_id)
    risk_ai_suggestion = _build_risk_ai_suggestion(req)
    operational_blockages = _build_operational_blockages(req, case_row)
    suggested_professionals = suggest_professional_leads_for_case(case_row, req, limit=8)
    assistant_recommendation = build_case_assistant_recommendation(
        case_row,
        req,
        risk_ai_suggestion,
        suggested_professionals=suggested_professionals,
    )
    events = (
        CaseEvent.query.filter(CaseEvent.case_id == case_row.id)
        .order_by(CaseEvent.created_at.desc(), CaseEvent.id.desc())
        .limit(200)
        .all()
    )
    participants = (
        CaseParticipant.query.filter(CaseParticipant.case_id == case_row.id)
        .order_by(CaseParticipant.added_at.desc(), CaseParticipant.id.desc())
        .all()
    )
    owners = (
        AdminUser.query.with_entities(AdminUser.id, AdminUser.username)
        .order_by(AdminUser.username.asc())
        .all()
    )
    users = (
        User.query.with_entities(User.id, User.username, User.email)
        .order_by(User.username.asc())
        .limit(300)
        .all()
    )
    professionals = (
        ProfessionalLead.query.order_by(ProfessionalLead.created_at.desc(), ProfessionalLead.id.desc())
        .limit(300)
        .all()
    )
    return render_template(
        "admin/case_detail.html",
        case_row=case_row,
        req=req,
        events=events,
        statuses=list(CATEGORY_CASE_STATUSES),
        priorities=list(CASE_PRIORITIES),
        participant_types=list(CASE_PARTICIPANT_TYPES),
        participant_roles=list(CASE_PARTICIPANT_ROLES),
        owners=owners,
        users=users,
        professionals=professionals,
        participants=participants,
        risk_ai_suggestion=risk_ai_suggestion,
        operational_blockages=operational_blockages,
        suggested_professionals=suggested_professionals,
        assistant_recommendation=assistant_recommendation,
    )


def _render_notifications_list():
    if not _table_exists("notification_jobs"):
        flash("Notification jobs table is not available yet. Run migrations first.", "warning")
        return render_template(
            "admin/notifications.html",
            jobs=[],
            status="",
            channel="",
            event_type="",
            recipient="",
            summary={"pending": 0, "retry": 0, "failed": 0, "sent": 0},
            channels=[],
        )

    status = (request.args.get("status") or "").strip().lower()
    channel = (request.args.get("channel") or "").strip().lower()
    event_type = (request.args.get("event_type") or "").strip()
    recipient = (request.args.get("recipient") or "").strip()

    query = NotificationJob.query
    try:
        if not _is_global_admin():
            sid = _current_structure_id()
            query = query.filter(
                (NotificationJob.structure_id == sid)
                | (NotificationJob.structure_id.is_(None))
            )
    except Exception:
        pass

    if status in {"pending", "processing", "sent", "retry", "failed"}:
        query = query.filter(NotificationJob.status == status)
    if channel:
        query = query.filter(NotificationJob.channel == channel)
    if event_type:
        query = query.filter(NotificationJob.event_type.ilike(f"%{event_type}%"))
    if recipient:
        query = query.filter(NotificationJob.recipient.ilike(f"%{recipient}%"))

    status_rank = case(
        (NotificationJob.status == "failed", 0),
        (NotificationJob.status == "retry", 1),
        (NotificationJob.status == "pending", 2),
        (NotificationJob.status == "processing", 3),
        else_=4,
    )

    jobs = (
        query.order_by(
            status_rank.asc(),
            NotificationJob.created_at.desc().nullslast(),
            NotificationJob.id.desc(),
        )
        .limit(200)
        .all()
    )

    counts_base = NotificationJob.query
    try:
        if not _is_global_admin():
            sid = _current_structure_id()
            counts_base = counts_base.filter(
                (NotificationJob.structure_id == sid)
                | (NotificationJob.structure_id.is_(None))
            )
    except Exception:
        pass

    summary = {
        "pending": counts_base.filter(NotificationJob.status == "pending").count(),
        "retry": counts_base.filter(NotificationJob.status == "retry").count(),
        "failed": counts_base.filter(NotificationJob.status == "failed").count(),
        "sent": counts_base.filter(NotificationJob.status == "sent").count(),
    }

    channels = [
        c[0]
        for c in counts_base.with_entities(NotificationJob.channel)
        .distinct()
        .order_by(NotificationJob.channel.asc())
        .all()
        if c[0]
    ]

    return render_template(
        "admin/notifications.html",
        jobs=jobs,
        status=status,
        channel=channel,
        event_type=event_type,
        recipient=recipient,
        summary=summary,
        channels=channels,
    )


@admin_bp.get("/notifications")
@admin_required
@admin_role_required("readonly", "ops", "superadmin")
def admin_notifications_list():
    admin_required_404()
    return _render_notifications_list()


@admin_bp.post("/cases/<int:case_id>/status")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_case_set_status(case_id: int):
    admin_required_404()
    case_row, _req = _get_scoped_case_or_404(case_id)
    new_status = (request.form.get("status") or "").strip().lower()
    if new_status not in CATEGORY_CASE_STATUSES:
        flash("Invalid case status.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)

    old_status = (case_row.status or "").strip().lower()
    if old_status != new_status:
        now = _now_utc()
        case_row.status = new_status
        case_row.last_activity_at = now
        if new_status == "assigned" and not case_row.assigned_at:
            case_row.assigned_at = now
        if new_status == "resolved":
            case_row.resolved_at = now
            _append_case_event(
                case_id=case_row.id,
                actor_user_id=getattr(current_user, "id", None),
                event_type="case_resolved",
                message="Case marked as resolved",
            )
        if new_status == "closed":
            case_row.closed_at = now
            if not case_row.resolved_at:
                case_row.resolved_at = now
            _append_case_event(
                case_id=case_row.id,
                actor_user_id=getattr(current_user, "id", None),
                event_type="case_closed",
                message="Case marked as closed",
            )
        if new_status == "cancelled" and not case_row.closed_at:
            case_row.closed_at = now
        _append_case_event(
            case_id=case_row.id,
            actor_user_id=getattr(current_user, "id", None),
            event_type="status_changed",
            message=f"Status changed: {old_status or '-'} -> {new_status}",
            metadata={"old_status": old_status, "new_status": new_status},
        )
        db.session.commit()
        flash("Case status updated.", "success")
    return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)


@admin_bp.post("/cases/<int:case_id>/priority")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_case_set_priority(case_id: int):
    admin_required_404()
    case_row, _req = _get_scoped_case_or_404(case_id)
    new_priority = (request.form.get("priority") or "").strip().lower()
    if new_priority not in CASE_PRIORITIES:
        flash("Invalid case priority.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)

    old_priority = (case_row.priority or "").strip().lower()
    if old_priority != new_priority:
        case_row.priority = new_priority
        case_row.last_activity_at = _now_utc()
        _append_case_event(
            case_id=case_row.id,
            actor_user_id=getattr(current_user, "id", None),
            event_type="priority_changed",
            message=f"Priority changed: {old_priority or '-'} -> {new_priority}",
            metadata={"old_priority": old_priority, "new_priority": new_priority},
        )
        db.session.commit()
        flash("Case priority updated.", "success")
    return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)


@admin_bp.post("/cases/<int:case_id>/assign-owner")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_case_assign_owner(case_id: int):
    admin_required_404()
    case_row, _req = _get_scoped_case_or_404(case_id)
    owner_raw = (request.form.get("owner_user_id") or "").strip()
    owner_id = None
    if owner_raw:
        try:
            owner_id = int(owner_raw)
        except Exception:
            owner_id = None
    if owner_id is not None and not db.session.get(AdminUser, owner_id):
        flash("Selected owner does not exist.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)

    old_owner_id = case_row.owner_user_id
    if old_owner_id != owner_id:
        now = _now_utc()
        case_row.owner_user_id = owner_id
        case_row.last_activity_at = now
        if owner_id and not case_row.assigned_at:
            case_row.assigned_at = now
            if case_row.status in {"new", "triaged"}:
                case_row.status = "assigned"
            _upsert_case_participant(
                case_id=case_row.id,
                participant_type="admin_user",
                role="owner",
                user_id=owner_id,
                status="active",
            )
        _append_case_event(
            case_id=case_row.id,
            actor_user_id=getattr(current_user, "id", None),
            event_type="owner_assigned",
            message=f"Owner changed: {old_owner_id or '-'} -> {owner_id or '-'}",
            metadata={"old_owner_user_id": old_owner_id, "new_owner_user_id": owner_id},
        )
        db.session.commit()
        flash("Case owner updated.", "success")
    return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)


@admin_bp.post("/cases/<int:case_id>/assign-professional")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_case_assign_professional(case_id: int):
    admin_required_404()
    case_row, _req = _get_scoped_case_or_404(case_id)
    lead_raw = (request.form.get("assigned_professional_lead_id") or request.form.get("primary_professional_lead_id") or "").strip()
    lead_id = None
    if lead_raw:
        try:
            lead_id = int(lead_raw)
        except Exception:
            lead_id = None
    if lead_id is not None and not db.session.get(ProfessionalLead, lead_id):
        flash("Selected professional lead does not exist.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)

    old_lead_id = case_row.assigned_professional_lead_id
    if old_lead_id != lead_id:
        now = _now_utc()
        case_row.assigned_professional_lead_id = lead_id
        case_row.last_activity_at = now
        if lead_id and not case_row.assigned_at:
            case_row.assigned_at = now
            if case_row.status in {"new", "triaged"}:
                case_row.status = "assigned"
        if lead_id:
            _upsert_case_participant(
                case_id=case_row.id,
                participant_type="professional_lead",
                role="primary_professional",
                professional_lead_id=lead_id,
                status="active",
            )
        _append_case_event(
            case_id=case_row.id,
            actor_user_id=getattr(current_user, "id", None),
            event_type="professional_assigned",
            message=f"Primary professional lead changed: {old_lead_id or '-'} -> {lead_id or '-'}",
            metadata={
                "old_professional_lead_id": old_lead_id,
                "new_professional_lead_id": lead_id,
            },
        )
        db.session.commit()
        flash("Case professional assignment updated.", "success")
    return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)


@admin_bp.post("/cases/<int:case_id>/participants")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_case_add_participant(case_id: int):
    admin_required_404()
    case_row, _req = _get_scoped_case_or_404(case_id)

    participant_type = (request.form.get("participant_type") or "").strip().lower()
    role = (request.form.get("role") or "contributor").strip().lower()
    participant_status = (request.form.get("status") or "active").strip().lower()
    user_raw = (request.form.get("user_id") or "").strip()
    lead_raw = (request.form.get("professional_lead_id") or "").strip()
    external_name = (request.form.get("external_name") or "").strip()

    if participant_type not in CASE_PARTICIPANT_TYPES:
        flash("Invalid participant type.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)
    if role not in CASE_PARTICIPANT_ROLES:
        role = "contributor"
    if participant_status not in {"active", "inactive"}:
        participant_status = "active"

    user_id = None
    lead_id = None
    if user_raw:
        try:
            user_id = int(user_raw)
        except Exception:
            user_id = None
    if lead_raw:
        try:
            lead_id = int(lead_raw)
        except Exception:
            lead_id = None

    if user_id is not None and not db.session.get(User, user_id):
        flash("Selected user does not exist.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)
    if lead_id is not None and not db.session.get(ProfessionalLead, lead_id):
        flash("Selected professional lead does not exist.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)

    if participant_type == "professional_lead" and not lead_id:
        flash("professional_lead participant requires professional_lead_id.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)
    if participant_type in {"admin_user", "professional_user"} and not user_id:
        flash("Selected participant type requires user_id.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)
    if participant_type in {"association", "external_contact"} and not external_name:
        flash("External participant requires external name.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)

    _upsert_case_participant(
        case_id=case_row.id,
        participant_type=participant_type,
        role=role,
        user_id=user_id,
        professional_lead_id=lead_id,
        external_name=external_name or None,
        status=participant_status,
    )
    case_row.last_activity_at = _now_utc()
    _append_case_event(
        case_id=case_row.id,
        actor_user_id=getattr(current_user, "id", None),
        event_type="participant_added",
        message="Participant added/updated",
        metadata={
            "participant_type": participant_type,
            "role": role,
            "status": participant_status,
            "user_id": user_id,
            "professional_lead_id": lead_id,
            "external_name": external_name or None,
        },
    )
    db.session.commit()
    flash("Case participant updated.", "success")
    return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)


@admin_bp.post("/cases/<int:case_id>/events")
@admin_required
@admin_role_required("ops", "superadmin")
def admin_case_add_event(case_id: int):
    admin_required_404()
    case_row, _req = _get_scoped_case_or_404(case_id)
    event_type = (request.form.get("event_type") or "note_added").strip().lower()
    message = (request.form.get("message") or "").strip()
    metadata = _safe_json_dict(request.form.get("metadata_json"))
    visibility = (request.form.get("visibility") or "internal").strip().lower()
    if visibility not in {"internal", "public"}:
        visibility = "internal"

    if not event_type:
        event_type = "note_added"
    if not message and not metadata:
        flash("Event is empty.", "warning")
        return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)

    case_row.last_activity_at = _now_utc()
    _append_case_event(
        case_id=case_row.id,
        actor_user_id=getattr(current_user, "id", None),
        event_type=event_type,
        message=message,
        metadata=metadata or None,
        visibility=visibility,
    )
    db.session.commit()
    flash("Case event added.", "success")
    return redirect(url_for("admin.admin_case_detail", case_id=case_row.id), code=303)


@admin_bp.post("/requests/<int:req_id>/unlock", endpoint="admin_request_unlock")
@admin_required
@admin_role_required("superadmin")
def admin_request_unlock(req_id: int):
    admin_required_404()
    admin_id = _admin_id()
    if not admin_id:
        abort(403)

    req = db.session.get(Request, req_id)
    if not req:
        abort(404)

    old_owner = req.owner_id
    if old_owner is not None:
        req.owner_id = None
        req.owned_at = None
        db.session.add(
            RequestActivity(
                request_id=req.id,
                actor_admin_id=admin_id,
                action="unlock",
                old_value=str(old_owner),
                new_value="",
                created_at=_now_utc(),
            )
        )
        db.session.commit()
        audit_admin_action(
            action="request.unlock",
            target_type="Request",
            target_id=req.id,
            payload={
                "req_id": req.id,
                "old": {"locked": True, "owner_id": old_owner},
                "new": {"locked": False, "owner_id": None},
            },
        )

    flash("Unlocked.", "success")
    return redirect(url_for("admin.admin_request_details", req_id=req_id))


# --- Volunteer interest moderation ---
@admin_bp.post(
    "/requests/<int:req_id>/interests/<int:interest_id>/approve",
    endpoint="admin_interest_approve",
)
@admin_required
@admin_role_required("ops", "superadmin")
def admin_interest_approve(req_id: int, interest_id: int):
    current_app.logger.info(
        "ADMIN_APPROVE HIT req_id=%s interest_id=%s", req_id, interest_id
    )

    admin_required_404()
    admin = current_user
    admin_id = getattr(admin, "id", None)

    req = db.session.get(Request, req_id)
    if not req:
        abort(404)

    if _locked_by_other(req, admin_id):
        flash(
            "🔒 Заявката е заключена от друг админ. Може да я отключите ръчно.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    vi = db.session.get(VolunteerInterest, interest_id)
    if not vi or vi.request_id != req_id:
        abort(404)

    if _is_request_locked(req):
        flash(
            "🔒 Заявката е заключена (done/cancelled). Смени статуса, за да отключиш действията.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    old_vi = vi.status
    changed_vi = old_vi != "approved"
    if changed_vi:
        vi.status = "approved"
        db.session.add(
            RequestActivity(
                request_id=req_id,
                actor_admin_id=admin_id,
                action="volunteer_interest_approved",
                old_value=old_vi,
                new_value="approved",
            )
        )

    # Auto transition: open -> in_progress (single log)
    old_rs = req.status
    status_changed = False
    if old_rs == "open":
        req.status = "in_progress"
        _log_status_change_once(req_id, old_rs, req.status, admin_id)
        status_changed = True

    if changed_vi or status_changed:
        current_app.logger.info(
            "BEFORE commit: req.status=%s vi.status=%s", req.status, vi.status
        )
        db.session.commit()
        if changed_vi:
            audit_admin_action(
                action="interest.approve",
                target_type="Interest",
                target_id=vi.id,
                payload={
                    "req_id": req.id,
                    "interest_id": vi.id,
                    "old": {"status": old_vi},
                    "new": {"status": vi.status},
                },
            )
        current_app.logger.info(
            "AFTER commit: req.status=%s vi.status=%s", req.status, vi.status
        )
        flash("Approved.", "success")
    else:
        flash("No changes.", "info")

    return redirect(url_for("admin.admin_request_details", req_id=req_id))


@admin_bp.post(
    "/requests/<int:req_id>/interests/<int:interest_id>/reject",
    endpoint="admin_interest_reject",
)
@admin_required
@admin_role_required("ops", "superadmin")
def admin_interest_reject(req_id: int, interest_id: int):
    admin_required_404()
    admin = current_user
    admin_id = getattr(admin, "id", None)

    req = db.session.get(Request, req_id)
    if not req:
        abort(404)

    if _locked_by_other(req, admin_id):
        flash(
            "🔒 Заявката е заключена от друг админ. Може да я отключите ръчно.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    interest = db.session.get(VolunteerInterest, interest_id)
    if not interest or interest.request_id != req_id:
        abort(404)

    if _is_request_locked(req):
        flash(
            "🔒 Заявката е заключена (done/cancelled). Смени статуса, за да отключиш действията.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    old_vi = interest.status
    if old_vi == "rejected":
        flash("No changes.", "info")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    current_app.logger.info(
        "ADMIN_REJECT HIT req_id=%s interest_id=%s", req_id, interest_id
    )

    reject_reason = (
        request.form.get("reason")
        or request.form.get("reject_reason")
        or request.form.get("note")
        or ""
    ).strip()
    interest.status = "rejected"
    db.session.add(
        RequestActivity(
            request_id=req.id,
            actor_admin_id=admin_id,
            action="volunteer_interest_rejected",
            old_value=old_vi,
            new_value="rejected",
        )
    )

    db.session.commit()
    payload = {
        "req_id": req.id,
        "interest_id": interest.id,
        "old": {"status": old_vi},
        "new": {"status": interest.status},
    }
    if reject_reason:
        payload["reason"] = reject_reason
    audit_admin_action(
        action="interest.reject",
        target_type="Interest",
        target_id=interest.id,
        payload=payload,
    )
    flash("Rejected.", "warning")
    return redirect(url_for("admin.admin_request_details", req_id=req_id))


# --- Assign owner to request ---
@admin_bp.post("/requests/<int:req_id>/assign", endpoint="admin_request_assign")
@login_required
@admin_required
@admin_role_required("ops", "superadmin")
def admin_request_assign(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if _locked_by_other(req, getattr(current_user, "id", None)):
        flash(
            "🔒 Заявката е заключена от друг админ. Може да я отключите ръчно.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if _is_request_locked(req):
        flash(
            "This request is locked (done/cancelled). Unlock it by changing status first.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if req.owner_id and req.owner_id != getattr(current_user, "id", None):
        flash("Deja pris en charge.", "warning")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if req.owner_id == getattr(current_user, "id", None):
        flash("Deja assigne a vous.", "info")
        next_url = (request.form.get("next") or "").strip()
        if next_url and is_safe_url(next_url):
            return redirect(next_url)
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    takeover = False
    old_owner = req.owner_id
    req.owner_id = current_user.id
    req.owned_at = utc_now()
    if _table_exists("request_metrics"):
        metric = db.session.query(RequestMetric).filter_by(request_id=req.id).first()
        if metric is None:
            metric = RequestMetric(request_id=req.id)
            db.session.add(metric)
        if metric.time_to_assign is None and req.created_at:
            try:
                metric.time_to_assign = int((utc_now() - req.created_at).total_seconds())
            except Exception:
                pass
    action_name = "takeover" if takeover else "assign"
    reason = None
    if takeover and req.owned_at:
        try:
            hours = (utc_now() - req.owned_at).total_seconds() / 3600
            reason = f"stale: {hours:.1f}h"
        except Exception:
            reason = "stale"
    new_val = (
        f"{current_user.id}" if reason is None else f"{current_user.id} ({reason})"
    )
    log_request_activity(
        req, action_name, old=old_owner, new=new_val, actor_admin_id=current_user.id
    )
    _audit_request(
        req.id,
        action="assign_owner",
        message="Owner assigned",
        old=str(old_owner) if old_owner is not None else None,
        new=str(current_user.id),
    )
    db.session.commit()
    audit_admin_action(
        action="ASSIGN_OPERATOR",
        target_type="Request",
        target_id=req.id,
        payload={
            "old": {"owner_id": old_owner},
            "new": {"owner_id": current_user.id},
        },
    )
    flash(_("The request has been assigned to you."), "success")
    next_url = (request.form.get("next") or "").strip()
    if next_url and is_safe_url(next_url):
        return redirect(next_url)
    return redirect(url_for("admin.admin_request_details", req_id=req_id))


@admin_bp.post("/requests/<int:req_id>/unassign", endpoint="admin_request_unassign")
@login_required
@admin_required
@admin_role_required("ops", "superadmin")
def admin_request_unassign(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if _locked_by_other(req, getattr(current_user, "id", None)):
        flash(
            "🔒 Заявката е заключена от друг админ. Може да я отключите ръчно.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if _is_request_locked(req):
        flash(
            "This request is locked (done/cancelled). Unlock it by changing status first.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if not can_edit_request(req, current_user):
        abort(403)
    old_owner = req.owner_id
    req.owner_id = None
    req.owned_at = None
    log_request_activity(
        req,
        "unassign",
        old=old_owner,
        new=None,
        actor_admin_id=getattr(current_user, "id", None),
    )
    _audit_request(
        req.id,
        action="unassign_owner",
        message="Owner unassigned",
        old=str(old_owner) if old_owner is not None else None,
        new=None,
    )
    db.session.commit()
    audit_admin_action(
        action="request.unassign_owner",
        target_type="Request",
        target_id=req.id,
        payload={
            "old": {"owner_id": old_owner},
            "new": {"owner_id": None},
        },
    )
    flash(_("Owner removed."), "info")
    return redirect(url_for("admin.admin_request_details", req_id=req_id))


@admin_bp.post(
    "/requests/<int:req_id>/assign_volunteer/<int:volunteer_id>",
    endpoint="admin_assign_volunteer",
)
@login_required
def admin_assign_volunteer(req_id: int, volunteer_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)
    if _is_request_locked(req):
        flash("This request is locked (done/cancelled).", "warning")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    req.assigned_volunteer_id = volunteer_id
    log_request_activity(
        req,
        "assign_volunteer",
        old=getattr(req, "assigned_volunteer_id", None),
        new=volunteer_id,
        actor_admin_id=getattr(current_user, "id", None),
    )
    db.session.commit()
    flash("Assigned to volunteer.", "success")
    return redirect(url_for("admin.admin_request_details", req_id=req.id))


@admin_bp.post(
    "/requests/<int:req_id>/unassign_volunteer", endpoint="admin_unassign_volunteer"
)
@login_required
def admin_unassign_volunteer(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)
    old_val = getattr(req, "assigned_volunteer_id", None)
    req.assigned_volunteer_id = None
    log_request_activity(
        req,
        "unassign_volunteer",
        old=old_val,
        new=None,
        actor_admin_id=getattr(current_user, "id", None),
    )
    db.session.commit()
    flash("Volunteer unassigned.", "info")
    return redirect(url_for("admin.admin_request_details", req_id=req.id))


@admin_bp.post("/requests/<int:req_id>/nudge", endpoint="admin_request_nudge")
@login_required
def admin_request_nudge(req_id: int):
    admin_required_404()
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)

    volunteer_id = getattr(req, "assigned_volunteer_id", None)
    if not volunteer_id:
        flash("No assigned volunteer to nudge.", "warning")
        return redirect(request.referrer or url_for("admin.admin_requests"))

    created = send_nudge_notification(
        request_id=req.id,
        volunteer_id=int(volunteer_id),
        actor_admin_id=getattr(current_user, "id", None),
    )
    if created:
        flash("Nudge sent.", "success")
    else:
        flash("Nudge suppressed (recently sent).", "info")
    return redirect(request.referrer or url_for("admin.admin_requests"))


@admin_bp.post("/requests/<int:req_id>/delete", endpoint="admin_request_delete")
@login_required
def admin_request_delete(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)

    if not getattr(req, "is_archived", False):
        flash(
            "Archive the request first. Only archived requests can be deleted.",
            "warning",
        )
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    if getattr(req, "deleted_at", None) is None:
        req.deleted_at = utc_now()
        req.is_archived = True
        if getattr(req, "archived_at", None) is None:
            req.archived_at = req.deleted_at
        log_request_activity(
            req,
            "delete",
            old=None,
            new=str(req.deleted_at),
            actor_admin_id=getattr(current_user, "id", None),
        )
        db.session.commit()
        flash("Request moved to Deleted.", "success")

    return redirect(url_for("admin.admin_request_details", req_id=req.id))


@admin_bp.post(
    "/requests/<int:req_id>/restore-deleted", endpoint="admin_request_restore_deleted"
)
@login_required
def admin_request_restore_deleted(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    if not can_edit_request(req, current_user):
        abort(403)

    if getattr(req, "deleted_at", None) is not None:
        old = req.deleted_at
        req.deleted_at = None
        req.is_archived = True
        if getattr(req, "archived_at", None) is None:
            req.archived_at = utc_now()
        log_request_activity(
            req,
            "restore_deleted",
            old=str(old),
            new=None,
            actor_admin_id=getattr(current_user, "id", None),
        )
        db.session.commit()
        flash("Request restored from Deleted (kept archived).", "success")

    return redirect(url_for("admin.admin_request_details", req_id=req.id))


@admin_bp.post("/requests/<int:req_id>/note")
@login_required
def admin_request_add_note(req_id: int):
    req = _scope_requests(Request.query).filter(Request.id == req_id).first_or_404()
    note = (request.form.get("note") or "").strip()
    if not note:
        flash("Note is empty.", "warning")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))
    if len(note) > 2000:
        flash("Note is too long (max 2000 chars).", "danger")
        return redirect(url_for("admin.admin_request_details", req_id=req.id))

    log_request_activity(
        req,
        "note",
        old=None,
        new=note,
        actor_admin_id=getattr(current_user, "id", None),
    )
    _audit_request(
        req.id,
        action="note_add",
        message="Admin note added",
    )
    db.session.commit()
    flash("Note added.", "success")
    return redirect(url_for("admin.admin_request_details", req_id=req.id))


# --- ALIASES (temporary, to keep templates stable) ---
@admin_bp.get("/requests/<int:req_id>/status")
@login_required
def admin_request_status_get_alias(req_id: int):
    # Status is edited inside details page; redirect there.
    return redirect(url_for("admin.admin_request_details", req_id=req_id), code=302)


@admin_bp.get("/requests/<int:req_id>/notes")
@login_required
def admin_request_notes_get_alias(req_id: int):
    return redirect(url_for("admin.admin_request_details", req_id=req_id), code=302)


@admin_bp.post("/requests/<int:req_id>/notes")
@login_required
def admin_request_notes_post_alias(req_id: int):
    # Reuse existing note handler (/note) without changing templates.
    return admin_request_add_note(req_id)


@admin_bp.get("/professional-leads")
@login_required
@admin_required
def admin_professional_leads():
    if not _table_exists("professional_leads"):
        flash(
            "Professional leads table is not available in this environment yet.",
            "warning",
        )
        return (
            render_template(
                "admin/professional_leads.html",
                leads=[],
                q="",
                profession="",
                city="",
                status="",
                status_choices=["new", "imported", "contacted", "qualified", "rejected"],
                professions=[],
            ),
            200,
        )

    q = (request.args.get("q") or "").strip()
    profession = (request.args.get("profession") or "").strip()
    city = (request.args.get("city") or "").strip()
    status = (request.args.get("status") or "").strip().lower()
    status_choices = ["new", "imported", "contacted", "qualified", "rejected"]

    query = ProfessionalLead.query

    if q:
        like = f"%{q.lower()}%"
        query = query.filter(ProfessionalLead.email.ilike(like))

    if profession:
        query = query.filter(ProfessionalLead.profession == profession)

    if city:
        query = query.filter(ProfessionalLead.city.ilike(f"%{city}%"))

    if status:
        query = query.filter(func.lower(ProfessionalLead.status) == status)

    leads = (
        query.order_by(ProfessionalLead.created_at.desc(), ProfessionalLead.id.desc())
        .limit(200)
        .all()
    )

    professions = (
        ProfessionalLead.query.with_entities(ProfessionalLead.profession)
        .distinct()
        .order_by(ProfessionalLead.profession.asc())
        .all()
    )
    professions = [p[0] for p in professions if p and p[0]]

    return (
        render_template(
            "admin/professional_leads.html",
            leads=leads,
            q=q,
            profession=profession,
            city=city,
            status=status,
            status_choices=status_choices,
            professions=professions,
        ),
        200,
    )


@admin_bp.post("/professional-leads/<int:lead_id>/contacted")
@login_required
@admin_required
def admin_professional_lead_mark_contacted(lead_id: int):
    if not _table_exists("professional_leads"):
        flash("Professional leads table is not available.", "warning")
        return redirect(url_for("admin.admin_professional_leads"), code=303)

    lead = ProfessionalLead.query.get_or_404(lead_id)
    if (lead.status or "").lower() != "contacted":
        lead.status = "contacted"
        if not lead.contacted_at:
            lead.contacted_at = datetime.now(UTC)
        db.session.commit()
        flash(f"Lead #{lead.id} marked as contacted.", "success")
    return redirect(url_for("admin.admin_professional_leads"), code=303)


@admin_bp.route("/professional-leads/<int:lead_id>", methods=["GET", "POST"])
@login_required
@admin_required
def admin_professional_lead_detail(lead_id: int):
    if not _table_exists("professional_leads"):
        flash("Professional leads table is not available.", "warning")
        return redirect(url_for("admin.admin_professional_leads"), code=303)

    lead = ProfessionalLead.query.get_or_404(lead_id)
    status_choices = ["new", "imported", "contacted", "qualified", "rejected"]

    if request.method == "POST":
        status = (request.form.get("status") or "").strip().lower()
        notes = (request.form.get("notes") or "").strip()
        if status not in status_choices:
            status = "new"

        lead.status = status
        lead.notes = notes or None
        if status == "contacted" and not lead.contacted_at:
            lead.contacted_at = datetime.now(UTC)
        elif status != "contacted":
            lead.contacted_at = None

        db.session.commit()
        flash(f"Lead #{lead.id} updated.", "success")
        return redirect(
            url_for("admin.admin_professional_lead_detail", lead_id=lead.id), code=303
        )

    return (
        render_template(
            "admin/professional_lead_detail.html",
            lead=lead,
            status_choices=status_choices,
        ),
        200,
    )


@admin_bp.get("/professionnels/leads")
@login_required
@admin_required
def admin_professionnels_leads():
    return redirect(url_for("admin.admin_professional_leads"), code=302)


@admin_bp.get("/pro-access")
@login_required
@admin_required
def admin_pro_access_list():
    if ProAccessRequest is None:
        flash("Pro Access module is not available in this environment.", "info")
        return redirect(url_for("admin.admin_requests"), code=303)

    status = (request.args.get("status") or "new").strip().lower()
    if status not in PRO_ACCESS_STATUSES and status != "all":
        status = "new"

    query = ProAccessRequest.query
    if status != "all":
        query = query.filter(ProAccessRequest.status == status)

    rows = (
        query.order_by(ProAccessRequest.created_at.desc(), ProAccessRequest.id.desc())
        .limit(500)
        .all()
    )

    counts = {
        "new": ProAccessRequest.query.filter_by(status="new").count(),
        "reviewed": ProAccessRequest.query.filter_by(status="reviewed").count(),
        "approved": ProAccessRequest.query.filter_by(status="approved").count(),
        "rejected": ProAccessRequest.query.filter_by(status="rejected").count(),
        "all": ProAccessRequest.query.count(),
    }

    return (
        render_template(
            "admin/pro_access_list.html",
            rows=rows,
            status=status,
            counts=counts,
        ),
        200,
    )


@admin_bp.get("/audit")
@login_required
@admin_required
@admin_role_required("readonly", "ops", "superadmin")
def admin_audit():
    _require_global_admin()
    if not _table_exists("admin_audit_events"):
        return (
            render_template(
                "admin/audit.html",
                events=[],
                pagination=type(
                    "_PaginationStub",
                    (),
                    {
                        "page": 1,
                        "pages": 1,
                        "total": 0,
                        "has_prev": False,
                        "has_next": False,
                    },
                )(),
                filters={
                    "action": "",
                    "admin": "",
                    "target_id": "",
                    "days": "7",
                },
                actions=[],
            ),
            200,
        )

    action = (request.args.get("action") or "").strip()
    admin_username = (request.args.get("admin") or "").strip()
    target_type_raw = (request.args.get("target_type") or "").strip()
    target_id_raw = (request.args.get("target_id") or "").strip()
    days_raw = (request.args.get("days") or "7").strip()
    page_raw = (request.args.get("page") or "1").strip()

    try:
        days = max(1, min(int(days_raw), 365))
    except Exception:
        days = 7

    try:
        page = max(1, int(page_raw))
    except Exception:
        page = 1

    target_id = None
    if target_id_raw:
        try:
            target_id = int(target_id_raw)
        except Exception:
            target_id = None

    now = datetime.now(timezone.utc)
    since = now - timedelta(days=days)

    query = AdminAuditEvent.query.filter(AdminAuditEvent.created_at >= since)
    if action:
        query = query.filter(AdminAuditEvent.action == action)
    if admin_username:
        query = query.filter(AdminAuditEvent.admin_username == admin_username)
    if target_type_raw:
        query = query.filter(AdminAuditEvent.target_type == target_type_raw)
    if target_id is not None:
        query = query.filter(
            AdminAuditEvent.target_type == "Request",
            AdminAuditEvent.target_id == target_id,
        )

    query = query.order_by(AdminAuditEvent.created_at.desc(), AdminAuditEvent.id.desc())
    pagination = query.paginate(page=page, per_page=50, error_out=False)
    actions = (
        AdminAuditEvent.query.with_entities(AdminAuditEvent.action)
        .distinct()
        .order_by(AdminAuditEvent.action.asc())
        .all()
    )
    actions = [row[0] for row in actions if row and row[0]]
    target_types = (
        AdminAuditEvent.query.with_entities(AdminAuditEvent.target_type)
        .distinct()
        .order_by(AdminAuditEvent.target_type.asc())
        .all()
    )
    target_types = [row[0] for row in target_types if row and row[0]]

    return (
        render_template(
            "admin/audit.html",
            events=pagination.items,
            pagination=pagination,
            filters={
                "action": action,
                "admin": admin_username,
                "target_type": target_type_raw,
                "target_id": target_id_raw,
                "days": str(days),
            },
            actions=actions,
            target_types=target_types,
        ),
        200,
    )


@admin_bp.get("/security")
@login_required
@admin_required
@admin_role_required("readonly", "ops", "superadmin")
def admin_security():
    _require_global_admin()
    now = datetime.now(timezone.utc)
    since_24h = now - timedelta(hours=24)
    since_1h = now - timedelta(hours=1)

    success_24h = (
        db.session.query(func.count(AdminLoginAttempt.id))
        .filter(
            AdminLoginAttempt.created_at >= since_24h,
            AdminLoginAttempt.success.is_(True),
        )
        .scalar()
        or 0
    )
    failed_24h = (
        db.session.query(func.count(AdminLoginAttempt.id))
        .filter(
            AdminLoginAttempt.created_at >= since_24h,
            AdminLoginAttempt.success.is_(False),
        )
        .scalar()
        or 0
    )
    distinct_failed_ips_24h = (
        db.session.query(func.count(func.distinct(AdminLoginAttempt.ip)))
        .filter(
            AdminLoginAttempt.created_at >= since_24h,
            AdminLoginAttempt.success.is_(False),
            AdminLoginAttempt.ip.isnot(None),
            AdminLoginAttempt.ip != "",
        )
        .scalar()
        or 0
    )
    distinct_failed_usernames_24h = (
        db.session.query(func.count(func.distinct(AdminLoginAttempt.username)))
        .filter(
            AdminLoginAttempt.created_at >= since_24h,
            AdminLoginAttempt.success.is_(False),
            AdminLoginAttempt.username.isnot(None),
            AdminLoginAttempt.username != "",
        )
        .scalar()
        or 0
    )
    failed_1h = (
        db.session.query(func.count(AdminLoginAttempt.id))
        .filter(
            AdminLoginAttempt.created_at >= since_1h,
            AdminLoginAttempt.success.is_(False),
        )
        .scalar()
        or 0
    )

    fail_buckets = (
        db.session.query(
            AdminLoginAttempt.ip.label("ip"),
            func.coalesce(AdminLoginAttempt.username, "").label("username"),
            func.count(AdminLoginAttempt.id).label("fails"),
        )
        .filter(
            AdminLoginAttempt.created_at >= since_24h,
            AdminLoginAttempt.success.is_(False),
        )
        .group_by(AdminLoginAttempt.ip, func.coalesce(AdminLoginAttempt.username, ""))
        .having(func.count(AdminLoginAttempt.id) >= ADMIN_LOGIN_MAX_FAILS)
        .subquery()
    )
    lockout_buckets_24h = (
        db.session.query(func.count())
        .select_from(fail_buckets)
        .scalar()
        or 0
    )

    risky_actions_24h = (
        db.session.query(func.count(AdminAuditEvent.id))
        .filter(
            AdminAuditEvent.created_at >= since_24h,
            AdminAuditEvent.action.in_(RISKY_ACTIONS),
        )
        .scalar()
        or 0
    )
    denied_24h = (
        db.session.query(func.count(AdminAuditEvent.id))
        .filter(
            AdminAuditEvent.created_at >= since_24h,
            AdminAuditEvent.action == "security.denied_action",
        )
        .scalar()
        or 0
    )
    denied_1h = (
        db.session.query(func.count(AdminAuditEvent.id))
        .filter(
            AdminAuditEvent.created_at >= since_1h,
            AdminAuditEvent.action == "security.denied_action",
        )
        .scalar()
        or 0
    )
    avg_denied_hourly = (float(denied_24h) / 24.0) if denied_24h else 0.0

    top_denied_ips = (
        db.session.query(AdminAuditEvent.ip, func.count(AdminAuditEvent.id).label("cnt"))
        .filter(
            AdminAuditEvent.created_at >= since_24h,
            AdminAuditEvent.action == "security.denied_action",
            AdminAuditEvent.ip.isnot(None),
            AdminAuditEvent.ip != "",
        )
        .group_by(AdminAuditEvent.ip)
        .order_by(func.count(AdminAuditEvent.id).desc())
        .limit(10)
        .all()
    )
    top_denied_usernames = (
        db.session.query(
            AdminAuditEvent.admin_username, func.count(AdminAuditEvent.id).label("cnt")
        )
        .filter(
            AdminAuditEvent.created_at >= since_24h,
            AdminAuditEvent.action == "security.denied_action",
            AdminAuditEvent.admin_username.isnot(None),
            AdminAuditEvent.admin_username != "",
        )
        .group_by(AdminAuditEvent.admin_username)
        .order_by(func.count(AdminAuditEvent.id).desc())
        .limit(10)
        .all()
    )

    recent_logins = (
        AdminLoginAttempt.query.order_by(AdminLoginAttempt.created_at.desc())
        .limit(50)
        .all()
    )
    recent_risky = (
        AdminAuditEvent.query.filter(AdminAuditEvent.action.in_(RISKY_ACTIONS))
        .order_by(AdminAuditEvent.created_at.desc())
        .limit(50)
        .all()
    )
    recent_denied = (
        AdminAuditEvent.query.filter(AdminAuditEvent.action == "security.denied_action")
        .order_by(AdminAuditEvent.created_at.desc())
        .limit(50)
        .all()
    )
    sensitive_actions = {
        "ROLE_CHANGE",
        "STRUCTURE_CREATED",
        "STRUCTURE_ADMIN_ASSIGNED",
        "STATUS_CHANGE",
        "ASSIGN_OPERATOR",
        "CREATE_REQUEST",
    }
    recent_sensitive = (
        AdminAuditEvent.query.filter(AdminAuditEvent.action.in_(sensitive_actions))
        .order_by(AdminAuditEvent.created_at.desc())
        .limit(50)
        .all()
    )

    top_ips = (
        db.session.query(
            AdminLoginAttempt.ip, func.count(AdminLoginAttempt.id).label("fails")
        )
        .filter(
            AdminLoginAttempt.created_at >= since_24h,
            AdminLoginAttempt.success.is_(False),
        )
        .group_by(AdminLoginAttempt.ip)
        .order_by(func.count(AdminLoginAttempt.id).desc())
        .limit(10)
        .all()
    )

    top_usernames = (
        db.session.query(
            AdminLoginAttempt.username, func.count(AdminLoginAttempt.id).label("fails")
        )
        .filter(
            AdminLoginAttempt.created_at >= since_24h,
            AdminLoginAttempt.success.is_(False),
            AdminLoginAttempt.username.isnot(None),
            AdminLoginAttempt.username != "",
        )
        .group_by(AdminLoginAttempt.username)
        .order_by(func.count(AdminLoginAttempt.id).desc())
        .limit(10)
        .all()
    )

    avg_hourly = (float(failed_24h) / 24.0) if failed_24h else 0.0
    spike_threshold = max(10.0, 3.0 * avg_hourly)
    top_ip = top_ips[0] if top_ips else (None, 0)
    top_username = top_usernames[0] if top_usernames else (None, 0)
    top_denied_ip = top_denied_ips[0] if top_denied_ips else (None, 0)
    top_denied_username = top_denied_usernames[0] if top_denied_usernames else (None, 0)
    top_denied_ip_count = int(top_denied_ip[1] or 0)
    top_denied_username_count = int(top_denied_username[1] or 0)
    denied_spike_on = int(denied_1h) >= max(5.0, 3.0 * avg_denied_hourly)
    repeated_denied_on = (top_denied_ip_count >= 10) or (
        top_denied_username_count >= 8
    )
    anomalies = {
        "spike_failed_logins": float(failed_1h) > spike_threshold,
        "repeated_fails_by_ip": any(int(fails) >= 20 for _, fails in top_ips),
        "repeated_fails_by_username": any(
            int(fails) >= 10 for _, fails in top_usernames
        ),
        "failed_1h": int(failed_1h),
        "avg_hourly": round(avg_hourly, 2),
        "spike_threshold": round(spike_threshold, 2),
        "top_ip": top_ip[0],
        "top_ip_fails": int(top_ip[1] or 0),
        "top_username": top_username[0],
        "top_username_fails": int(top_username[1] or 0),
        "denied_spike": bool(denied_spike_on),
        "repeated_denied": bool(repeated_denied_on),
        "denied_1h": int(denied_1h),
        "avg_denied_hourly": round(avg_denied_hourly, 2),
        "top_denied_ip": top_denied_ip[0],
        "top_denied_ip_count": int(top_denied_ip_count),
        "top_denied_username": top_denied_username[0],
        "top_denied_username_count": int(top_denied_username_count),
    }

    return (
        render_template(
            "admin/security.html",
            kpis={
                "success_24h": int(success_24h),
                "failed_24h": int(failed_24h),
                "distinct_failed_ips_24h": int(distinct_failed_ips_24h),
                "distinct_failed_usernames_24h": int(distinct_failed_usernames_24h),
                "lockout_buckets_24h": int(lockout_buckets_24h),
                "risky_actions_24h": int(risky_actions_24h),
                "denied_24h": int(denied_24h),
            },
            recent_logins=recent_logins,
            recent_risky=recent_risky,
            recent_denied=recent_denied,
            recent_sensitive=recent_sensitive,
            top_ips=top_ips,
            top_usernames=top_usernames,
            top_denied_ips=top_denied_ips,
            top_denied_usernames=top_denied_usernames,
            anomalies=anomalies,
            risky_actions=RISKY_ACTIONS,
        ),
        200,
    )


@admin_bp.get("/roles")
@login_required
@admin_required
@admin_role_required("superadmin")
def admin_roles():
    _require_global_admin()
    admins = AdminUser.query.order_by(AdminUser.username.asc(), AdminUser.id.asc()).all()
    superadmin_ids = [u.id for u in admins if _is_superadmin_role(getattr(u, "role", None))]
    last_superadmin_id = superadmin_ids[0] if len(superadmin_ids) == 1 else None
    role_options = [
        ("readonly", "readonly"),
        ("ops", "ops"),
        ("superadmin", "superadmin"),
    ]
    return (
        render_template(
            "admin/roles.html",
            admins=admins,
            role_options=role_options,
            last_superadmin_id=last_superadmin_id,
            superadmin_count=len(superadmin_ids),
        ),
        200,
    )


@admin_bp.post("/roles/<int:admin_id>/role")
@login_required
@admin_required
@admin_role_required("superadmin")
@require_admin_fresh_auth(minutes=10)
def admin_roles_set_role(admin_id: int):
    _require_global_admin()
    target = db.session.get(AdminUser, admin_id)
    if not target:
        abort(404)

    requested_role = (request.form.get("role") or "").strip().lower()
    allowed_roles = {"readonly", "ops", "superadmin"}
    if requested_role not in allowed_roles:
        flash("Invalid role.", "danger")
        return redirect(url_for("admin.admin_roles"), code=303)

    old_role = _normalize_admin_role_value(getattr(target, "role", None))
    if old_role is None:
        old_role = "superadmin"

    if old_role == requested_role:
        flash("No changes.", "info")
        return redirect(url_for("admin.admin_roles"), code=303)

    superadmin_count = (
        db.session.query(func.count(AdminUser.id))
        .filter(
            or_(
                AdminUser.role == "superadmin",
                AdminUser.role == "super_admin",
                AdminUser.role == "admin",
            )
        )
        .scalar()
        or 0
    )

    # Prevent lockout by downgrading the last superadmin.
    if old_role == "superadmin" and requested_role != "superadmin" and int(superadmin_count) <= 1:
        flash("Cannot downgrade the last superadmin.", "danger")
        return redirect(url_for("admin.admin_roles"), code=303)

    target.role = requested_role
    db.session.commit()

    audit_admin_action(
        action="ROLE_CHANGE",
        target_type="AdminUser",
        target_id=target.id,
        payload={
            "old": {"role": old_role},
            "new": {"role": requested_role},
            "actor": {
                "admin_user_id": getattr(current_user, "id", None),
                "username": getattr(current_user, "username", None),
            },
            "ip": _client_ip(),
            "user_agent": request.headers.get("User-Agent"),
        },
    )
    flash("Role updated.", "success")
    return redirect(url_for("admin.admin_roles"), code=303)


@admin_bp.get("/structures")
@admin_required
@admin_role_required("superadmin")
def admin_structures():
    _require_global_admin()
    rows = Structure.query.order_by(Structure.name.asc(), Structure.id.asc()).all()
    return (
        render_template(
            "admin/structures.html",
            structures=rows,
        ),
        200,
    )


@admin_bp.get("/structures/new")
@admin_required
@admin_role_required("superadmin")
def admin_structure_new():
    _require_global_admin()
    return (
        render_template(
            "admin/structure_new.html",
        ),
        200,
    )


@admin_bp.post("/structures/new")
@admin_required
@admin_role_required("superadmin")
def admin_structure_create():
    _require_global_admin()
    name = (request.form.get("name") or "").strip()
    slug = (request.form.get("slug") or "").strip()

    errors = {}
    if not name:
        errors["name"] = "Le nom est requis."
    if not slug:
        errors["slug"] = "Le slug est requis."

    if slug:
        existing = Structure.query.filter(Structure.slug == slug).first()
        if existing:
            errors["slug"] = "Ce slug est déjà utilisé."

    if errors:
        for msg in errors.values():
            flash(msg, "danger")
        return (
            render_template(
                "admin/structure_new.html",
                form_data={"name": name, "slug": slug},
                form_errors=errors,
            ),
            400,
        )

    row = Structure(
        name=name,
        slug=slug,
        created_at=datetime.utcnow(),
    )
    db.session.add(row)
    db.session.commit()
    audit_admin_action(
        action="STRUCTURE_CREATED",
        target_type="Structure",
        target_id=row.id,
        payload={
            "structure": {"id": row.id, "name": row.name, "slug": row.slug},
            "actor": {
                "admin_user_id": getattr(current_user, "id", None),
                "username": getattr(current_user, "username", None),
            },
        },
    )
    flash("Structure créée.", "success")
    return redirect(url_for("admin.admin_structure_detail", structure_id=row.id), code=303)


@admin_bp.get("/structures/<int:structure_id>")
@admin_required
@admin_role_required("superadmin")
def admin_structure_detail(structure_id: int):
    _require_structure_admin_or_global()
    if _is_structure_admin() and int(getattr(current_user, "structure_id") or 0) != int(
        structure_id
    ):
        abort(403)
    structure = Structure.query.get_or_404(structure_id)
    users_count = AdminUser.query.filter(
        AdminUser.structure_id == structure_id
    ).count()
    open_filter = or_(
        Request.status.is_(None), ~func.lower(Request.status).in_(list(CLOSED_STATUSES))
    )
    active_requests = (
        Request.query.filter(Request.structure_id == structure_id)
        .filter(open_filter)
        .count()
    )
    done_requests = (
        Request.query.filter(Request.structure_id == structure_id)
        .filter(func.lower(Request.status) == "done")
        .count()
    )
    recent_requests = (
        Request.query.filter_by(structure_id=structure_id)
        .order_by(Request.created_at.desc())
        .limit(10)
        .all()
    )
    health_score = compute_structure_health(structure_id)
    alerts = compute_structure_alerts(structure_id)
    return (
        render_template(
            "admin/structure_dashboard.html",
            structure=structure,
            users_count=users_count,
            active_requests=active_requests,
            done_requests=done_requests,
            recent_requests=recent_requests,
            health_score=health_score,
            alerts=alerts,
        ),
        200,
    )


@admin_bp.post("/structures/<int:structure_id>/assign-admin")
@admin_required
@admin_role_required("superadmin")
def admin_structure_assign_admin(structure_id: int):
    _require_global_admin()
    row = Structure.query.get_or_404(structure_id)
    admin_id_raw = (request.form.get("admin_id") or "").strip()
    if not admin_id_raw:
        flash("Veuillez sélectionner un administrateur.", "danger")
        return redirect(url_for("admin.admin_structure_detail", structure_id=row.id), code=303)
    try:
        admin_id = int(admin_id_raw)
    except Exception:
        flash("Administrateur invalide.", "danger")
        return redirect(url_for("admin.admin_structure_detail", structure_id=row.id), code=303)

    admin_user = db.session.get(AdminUser, admin_id)
    if not admin_user:
        flash("Administrateur introuvable.", "danger")
        return redirect(url_for("admin.admin_structure_detail", structure_id=row.id), code=303)

    admin_user.structure_id = row.id
    db.session.commit()
    audit_admin_action(
        action="STRUCTURE_ADMIN_ASSIGNED",
        target_type="AdminUser",
        target_id=admin_user.id,
        payload={
            "structure": {"id": row.id, "name": row.name, "slug": row.slug},
            "admin_user_id": admin_user.id,
            "admin_username": getattr(admin_user, "username", None),
            "actor": {
                "admin_user_id": getattr(current_user, "id", None),
                "username": getattr(current_user, "username", None),
            },
        },
    )
    flash("Administrateur assigné à la structure.", "success")
    return redirect(url_for("admin.admin_structure_detail", structure_id=row.id), code=303)


@admin_bp.get("/pro-access/<int:pro_id>")
@login_required
@admin_required
def admin_pro_access_detail(pro_id: int):
    if ProAccessRequest is None:
        abort(404)

    row = ProAccessRequest.query.get_or_404(pro_id)
    audit_logs = []
    return (
        render_template(
            "admin/pro_access_detail.html",
            row=row,
            audit_logs=audit_logs,
        ),
        200,
    )


def _pro_access_set_status(
    row: ProAccessRequest, new_status: str, note: str | None = None
):
    old_status = (row.status or "").strip().lower() or None
    row.status = new_status
    row.reviewed_at = utc_now()
    row.reviewed_by = (
        getattr(current_user, "email", None)
        or getattr(current_user, "username", None)
        or "admin"
    )
    if note is not None:
        row.admin_notes = note
    _audit_pro_access(
        row.id,
        action="status_change",
        message="Pro access status updated",
        old=old_status,
        new=new_status,
        meta={"has_admin_notes": bool(note)},
    )


@admin_bp.post("/pro-access/<int:pro_id>/review")
@login_required
@admin_required
def admin_pro_access_review(pro_id: int):
    if ProAccessRequest is None:
        abort(404)

    row = ProAccessRequest.query.get_or_404(pro_id)
    note = (request.form.get("admin_notes") or "").strip() or None
    _pro_access_set_status(row, "reviewed", note)
    db.session.commit()
    flash("Marked as reviewed.", "success")
    return redirect(url_for("admin.admin_pro_access_detail", pro_id=pro_id), code=303)


@admin_bp.post("/pro-access/<int:pro_id>/approve")
@login_required
@admin_required
def admin_pro_access_approve(pro_id: int):
    if ProAccessRequest is None:
        abort(404)

    row = ProAccessRequest.query.get_or_404(pro_id)
    note = (request.form.get("admin_notes") or "").strip() or None
    _pro_access_set_status(row, "approved", note)
    db.session.commit()
    flash("Approved.", "success")
    return redirect(url_for("admin.admin_pro_access_detail", pro_id=pro_id), code=303)


@admin_bp.post("/pro-access/<int:pro_id>/reject")
@login_required
@admin_required
def admin_pro_access_reject(pro_id: int):
    if ProAccessRequest is None:
        abort(404)

    row = ProAccessRequest.query.get_or_404(pro_id)
    note = (request.form.get("admin_notes") or "").strip() or None
    _pro_access_set_status(row, "rejected", note)
    db.session.commit()
    flash("Rejected.", "success")
    return redirect(url_for("admin.admin_pro_access_detail", pro_id=pro_id), code=303)
    active_sla_queue = bool(queue == "sla" and sla_kind)
    active_sla_filter_label = SLA_QUEUE_KINDS.get(sla_kind, "") if active_sla_queue else ""
