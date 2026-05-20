from __future__ import annotations

from datetime import UTC, datetime, timedelta

from backend.extensions import db
from backend.models import Request, Structure
from backend.helpchain_backend.src.services.reporting.operations_report import (
    build_operational_report,
)


def _make_structure(name: str = "Test Structure") -> Structure:
    structure = Structure(name=name, slug=name.lower().replace(" ", "-"), status="active")
    db.session.add(structure)
    db.session.flush()
    return structure


def _make_request(
    *,
    structure_id: int,
    title: str,
    status: str = "new",
    created_at: datetime,
    completed_at: datetime | None = None,
    owned_at: datetime | None = None,
    category: str = "coordination",
    deleted_at: datetime | None = None,
    is_archived: bool = False,
) -> Request:
    req = Request(
        title=title,
        description=title,
        category=category,
        status=status,
        city="Paris",
        structure_id=structure_id,
        created_at=created_at,
        completed_at=completed_at,
        owned_at=owned_at,
        deleted_at=deleted_at,
        is_archived=is_archived,
    )
    db.session.add(req)
    db.session.flush()
    return req


def test_operations_report_counts_and_scope(app, db_schema):
    now = datetime(2026, 5, 17, 12, 0, tzinfo=UTC)

    with app.app_context():
        s1 = _make_structure("Alpha")
        s2 = _make_structure("Beta")

        _make_request(
            structure_id=s1.id,
            title="new in scope",
            status="new",
            created_at=now - timedelta(days=1),
            category="logement",
        )
        _make_request(
            structure_id=s1.id,
            title="resolved in scope",
            status="done",
            created_at=now - timedelta(days=3),
            completed_at=now - timedelta(days=1),
            owned_at=now - timedelta(days=2),
            category="sante",
        )
        _make_request(
            structure_id=s1.id,
            title="stale in scope",
            status="in_progress",
            created_at=now - timedelta(days=5),
            category="coordination",
        )
        _make_request(
            structure_id=s2.id,
            title="other structure",
            status="new",
            created_at=now - timedelta(days=1),
            category="other",
        )

        report = build_operational_report(structure_id=s1.id, days=7, now=now)

        assert report["scope"]["structure_id"] == s1.id
        assert report["scope"]["structure_name"] == "Alpha"
        assert report["requests"]["new"] == 3
        assert report["requests"]["resolved"] == 1
        assert report["requests"]["open"] == 2
        assert report["requests"]["stale"] == 1
        assert report["requests"]["unassigned"] == 2


def test_operations_report_excludes_deleted_and_archived(app, db_schema):
    now = datetime(2026, 5, 17, 12, 0, tzinfo=UTC)

    with app.app_context():
        s1 = _make_structure("Alpha")

        _make_request(
            structure_id=s1.id,
            title="visible",
            status="new",
            created_at=now - timedelta(days=1),
        )
        _make_request(
            structure_id=s1.id,
            title="deleted",
            status="new",
            created_at=now - timedelta(days=1),
            deleted_at=now - timedelta(hours=1),
        )
        _make_request(
            structure_id=s1.id,
            title="archived",
            status="new",
            created_at=now - timedelta(days=1),
            is_archived=True,
        )

        report = build_operational_report(structure_id=s1.id, days=7, now=now)

        assert report["requests"]["new"] == 1
        assert report["requests"]["open"] == 1
        assert report["breakdowns"]["by_category"][0]["count"] == 1


def test_operations_report_payload_shape(app, db_schema):
    now = datetime(2026, 5, 17, 12, 0, tzinfo=UTC)

    with app.app_context():
        s1 = _make_structure("Alpha")

        _make_request(
            structure_id=s1.id,
            title="visible",
            status="new",
            created_at=now - timedelta(days=1),
        )

        report = build_operational_report(structure_id=s1.id, days=7, now=now)

        assert set(report.keys()) == {
            "generated_at",
            "period",
            "scope",
            "requests",
            "sla",
            "sla_alerts",
            "breakdowns",
            "timeline",
            "items",
            "timeline_charts",
            "executive_summary",
            "executive_snapshot",
            "priority_actions",
            "operational_severity",
            "recommendations",
            "territorial_pressure",
            "automatic_analysis",
            "operational_conclusion",
            "trends",
            "definition",
        }
        assert "by_category" in report["breakdowns"]
        assert "by_status" in report["breakdowns"]
        assert "avg_assignment_hours" in report["sla"]
        assert "avg_resolution_hours" in report["sla"]
        assert len(report["executive_snapshot"]) == 6
        assert isinstance(report["priority_actions"], list)
        assert isinstance(report["territorial_pressure"]["zones"], list)
        assert isinstance(report["automatic_analysis"], list)
        assert "primary_recommendation" in report["operational_conclusion"]

def test_report_ignores_negative_or_invalid_resolution_durations(app, db_session):
    from datetime import UTC, datetime, timedelta

    from backend.models import Request
    from backend.helpchain_backend.src.services.reporting.operations_report import (
        build_operational_report,
    )

    now = datetime.now(UTC)

    broken = Request(
        title="Broken legacy request",
        status="closed",
        created_at=now,
        completed_at=now - timedelta(days=7),
    )

    db_session.add(broken)
    db_session.commit()

    with app.app_context():
        report = build_operational_report(days=30)

    assert report["sla"]["avg_resolution_hours"] >= 0


def test_operations_report_priority_actions_and_territorial_pressure(app, db_schema):
    now = datetime(2026, 5, 17, 12, 0, tzinfo=UTC)

    with app.app_context():
        structure = _make_structure("Gamma")

        for index in range(4):
            _make_request(
                structure_id=structure.id,
                title=f"pressure-{index}",
                status="open",
                created_at=now - timedelta(days=4, hours=index),
                category="coordination",
            )

        report = build_operational_report(structure_id=structure.id, days=7, now=now)

        assert report["priority_actions"]
        assert report["priority_actions"][0]["severity"] in {"critical", "high", "moderate", "stable"}
        assert report["territorial_pressure"]["zones"]
        assert report["territorial_pressure"]["zones"][0]["city"] == "Paris"


def test_operations_report_xlsx_workbook_structure(app, db_schema):
    from io import BytesIO

    from openpyxl import load_workbook

    from backend.helpchain_backend.src.routes.admin import _build_operational_report_xlsx_response

    now = datetime(2026, 5, 17, 12, 0, tzinfo=UTC)

    with app.app_context():
        structure = _make_structure("Workbook Structure")
        _make_request(
            structure_id=structure.id,
            title="DEMO OPS 12",
            status="in_progress",
            created_at=now - timedelta(days=2),
            category="housing",
        )

        report = build_operational_report(structure_id=structure.id, days=7, now=now)
        response = _build_operational_report_xlsx_response(report)
        workbook = load_workbook(BytesIO(response.get_data()))

    assert workbook.sheetnames == [
        "Synthèse exécutive",
        "Indicateurs opérationnels",
        "Tensions territoriales",
        "Actions prioritaires",
        "Situations opérationnelles",
        "Analyse automatique",
        "Définitions",
    ]
    assert workbook["Synthèse exécutive"]["A1"].value == "HelpChain - Rapport opérationnel"
    assert workbook["Situations opérationnelles"]["B5"].value == "Situation #1"
